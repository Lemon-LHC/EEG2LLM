import os
import json
import time
import numpy as np
import pandas as pd
import traceback
import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm
from openai import OpenAI
from transformers.utils.versions import require_version
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix

require_version("openai>=1.5.0", "To fix: pip install openai>=1.5.0")

# 定义睡眠阶段标签，避免重复硬编码
SLEEP_STAGE_LABELS = [
    'Wake (W)', 
    'NREM Stage 1 (N1)', 
    'NREM Stage 2 (N2)', 
    'NREM Stage 3 (N3)', 
    'NREM Stage 4 (N4)', 
    'REM Sleep (R)'
]

def load_test_data(file_path, max_samples=None):
    """加载测试数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        test_data = json.load(f)
    
    # 如果设置了最大样本数，随机选择样本
    if max_samples and max_samples < len(test_data):
        indices = np.random.choice(len(test_data), max_samples, replace=False)
        test_data = [test_data[i] for i in indices]
    
    return test_data

def load_dataset_info(test_data_path):
    """从information目录加载数据集信息
    
    Args:
        test_data_path: 测试数据文件路径
        
    Returns:
        dataset_info: 数据集信息，如果无法加载则返回None
    """
    try:
        # 从测试数据路径提取关键信息
        import re
        
        file_name = os.path.basename(test_data_path)
        # 提取edf文件数和采样率
        match_files = re.search(r'edf(\d+)_(\d+)hz', file_name)
        if match_files:
            max_files = match_files.group(1)
            sampling_rate = match_files.group(2)
            
            # 构建可能的信息文件路径
            info_dir = "/data/lhc/datasets_new/sleep/information"
            json_path = os.path.join(info_dir, f"sleep_stage_statistics_{max_files}_{sampling_rate}hz.json")
            excel_path = os.path.join(info_dir, f"sleep_stage_statistics_{max_files}_{sampling_rate}hz.xlsx")
            
            # 尝试读取JSON格式的信息
            if os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    dataset_info = json.load(f)
                print(f"已加载数据集信息: {json_path}")
                return dataset_info
            
            # 如果JSON不存在，尝试读取Excel格式
            elif os.path.exists(excel_path):
                print(f"发现Excel格式的数据集信息，但当前未实现从Excel读取功能: {excel_path}")
                return None
            
            else:
                print(f"未找到匹配的数据集信息文件 ({json_path} 或 {excel_path})")
                
                # 尝试在information目录下找到任何可能的统计文件
                if os.path.exists(info_dir):
                    for file in os.listdir(info_dir):
                        if file.endswith('.json') and 'statistics' in file:
                            json_path = os.path.join(info_dir, file)
                            print(f"发现其他数据集信息文件: {json_path}")
                            with open(json_path, 'r', encoding='utf-8') as f:
                                dataset_info = json.load(f)
                            print(f"已加载替代数据集信息: {json_path}")
                            return dataset_info
        
        # 尝试提取训练集信息，从测试集文件名
        train_file_path = test_data_path.replace("_test.json", "_train.json")
        if os.path.exists(train_file_path):
            print(f"找到对应的训练集文件: {train_file_path}")
            
            # 统计训练集各阶段样本数
            with open(train_file_path, 'r', encoding='utf-8') as f:
                train_data = json.load(f)
            
            # 计算训练集阶段分布
            train_counts = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0}
            for item in train_data:
                if "output" in item:
                    try:
                        stage = int(item["output"])
                        train_counts[stage] = train_counts.get(stage, 0) + 1
                    except (ValueError, KeyError):
                        pass
            
            # 计算测试集阶段分布
            test_counts = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0}
            with open(test_data_path, 'r', encoding='utf-8') as f:
                test_data = json.load(f)
                
            for item in test_data:
                if "output" in item:
                    try:
                        stage = int(item["output"])
                        test_counts[stage] = test_counts.get(stage, 0) + 1
                    except (ValueError, KeyError):
                        pass
            
            # 创建数据集信息
            dataset_info = {
                "dataset_name": os.path.basename(test_data_path).split('.')[0],
                "train_samples": len(train_data),
                "test_samples": len(test_data),
                "train_stage_counts": train_counts,
                "test_stage_counts": test_counts
            }
            
            print(f"已手动生成数据集信息")
            return dataset_info
                
        print("未能从测试数据路径中提取足够的信息以找到对应的数据集统计文件")
        return None
    
    except Exception as e:
        print(f"加载数据集信息时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def save_results_to_excel(results, metrics, test_data, save_dir, model_name="unknown_model", test_set_name="unknown_dataset", dataset_info=None):
    """保存测试结果为Excel表格文件，包含训练集和测试集的信息
    
    Args:
        results: 详细的测试结果
        metrics: 性能指标
        test_data: 测试数据
        save_dir: 保存目录 (与混淆矩阵相同的目录)
        model_name: 模型名称
        test_set_name: 测试集名称
        dataset_info: 数据集信息
    """
    try:
        # 导入必要的模块
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from datetime import datetime

        # 直接使用传入的save_dir作为结果目录
        # 将Excel表格保存在与混淆矩阵相同的目录下，并按模型和数据集命名
        excel_path = os.path.join(save_dir, f'{model_name}_{test_set_name}_results.xlsx')
        
        # 创建一个Excel工作簿
        wb = Workbook()
        
        # 设置单元格样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 计算各睡眠阶段的样本数
        class_counts = {}
        for r in results:
            label = r['true_label']
            if label not in class_counts:
                class_counts[label] = 0
            class_counts[label] += 1
        
        # 1. 总体指标摘要工作表
        ws_summary = wb.active
        ws_summary.title = "总体指标"
        
        # 添加标题和基本信息
        ws_summary.append(["睡眠阶段分类测试结果摘要"])
        ws_summary.append(["测试时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append(["模型名称", model_name])
        ws_summary.append(["测试集名称", test_set_name])
        ws_summary.append([])  # 空行
        
        # 如果有数据集信息，添加数据集统计信息
        if dataset_info:
            ws_summary.append(["数据集完整统计"])
            
            # 如果有训练集和测试集的比例信息
            if 'train_samples' in dataset_info and 'test_samples' in dataset_info:
                train_samples = dataset_info['train_samples']
                test_samples = dataset_info['test_samples']
                total_samples = train_samples + test_samples
                
                train_ratio = train_samples / total_samples if total_samples > 0 else 0
                test_ratio = test_samples / total_samples if total_samples > 0 else 0
                
                ws_summary.append(["训练集样本数", train_samples, f"{train_ratio:.2%}"])
                ws_summary.append(["测试集样本数", test_samples, f"{test_ratio:.2%}"])
                ws_summary.append(["总样本数", total_samples])
            
            ws_summary.append([])  # 空行
        
        # 数据集统计信息
        ws_summary.append(["当前测试数据统计"])
        ws_summary.append(["总样本数", len(test_data)])
        ws_summary.append(["有效预测样本数", len([r for r in results if r['prediction'] != -1])])
        ws_summary.append([])  # 空行
        
        # 整体性能指标
        ws_summary.append(["整体性能指标"])
        ws_summary.append(["总体准确率", metrics['accuracy']])
        ws_summary.append(["宏平均F1分数", metrics['f1_macro']])
        ws_summary.append(["平均推理时间(秒/样本)", metrics['avg_inference_time']])
        ws_summary.append([])  # 空行
        
        # 各睡眠阶段准确率
        ws_summary.append(["各睡眠阶段准确率"])
        headers = ["睡眠阶段", "准确率"]
        ws_summary.append(headers)
        
        for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
            acc = metrics['class_accuracies'][class_idx]
            ws_summary.append([label_name, acc])
        
        # 应用样式
        for cell in ws_summary[1][0:1]:
            cell.font = Font(bold=True, size=14)
            
        # 2. 混淆矩阵工作表
        ws_cm = wb.create_sheet("混淆矩阵")
        
        # 添加标题
        ws_cm.append(["睡眠阶段分类混淆矩阵"])
        ws_cm.merge_cells('A1:G1')
        ws_cm['A1'].font = Font(bold=True, size=14)
        ws_cm['A1'].alignment = center_alignment
        
        # 添加列标题
        headers = ["", "预测"] + [f"{label.split(' ')[0]}" for label in SLEEP_STAGE_LABELS]
        ws_cm.append(headers)
        
        # 标记合并单元格
        ws_cm.merge_cells('B2:G2')
        ws_cm['B2'].alignment = center_alignment
        ws_cm['B2'].font = header_font
        
        # 添加行标题和数据
        ws_cm.append(["真实标签", ""])
        
        for i, label in enumerate(SLEEP_STAGE_LABELS):
            row = [label.split(' ')[0]] + [metrics['confusion_matrix'][i, j] for j in range(6)]
            ws_cm.append(row)
            
        # 应用样式
        for row in ws_cm.iter_rows(min_row=3, max_row=9, min_col=1, max_col=1):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_alignment
                
        for row in ws_cm.iter_rows(min_row=4, max_row=9, min_col=2, max_col=7):
            for cell in row:
                cell.alignment = center_alignment
                
        # 设置条件格式，高亮对角线(正确预测)
        for i in range(4, 10):
            diagonal_cell = ws_cm.cell(row=i, column=i-2)
            diagonal_cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
            
        # 3. 样本分布工作表
        ws_dist = wb.create_sheet("样本分布")
        
        # 添加标题
        ws_dist.append(["测试数据集样本分布"])
        ws_dist.merge_cells('A1:C1')
        ws_dist['A1'].font = Font(bold=True, size=14)
        ws_dist['A1'].alignment = center_alignment
        
        # 添加列标题
        ws_dist.append(["睡眠阶段", "样本数量", "百分比"])
        
        # 添加数据
        total_samples = len(test_data)
        for i in range(6):
            count = class_counts.get(i, 0)
            percentage = (count / total_samples) * 100 if total_samples > 0 else 0
            ws_dist.append([SLEEP_STAGE_LABELS[i], count, f"{percentage:.2f}%"])
            
        # 添加总计行
        ws_dist.append(["总计", total_samples, "100.00%"])
        
        # 应用样式
        for i, row in enumerate(ws_dist.iter_rows(min_row=2, max_row=9, min_col=1, max_col=3)):
            for cell in row:
                cell.border = border
                if i == 0 or i == 7:  # 标题行和总计行
                    cell.font = header_font
                    if i == 0:
                        cell.fill = header_fill
                        
                if i > 0:  # 数据行
                    cell.alignment = Alignment(horizontal='center')
        
        # 4. 如果有数据集信息，添加训练测试集分布工作表
        if dataset_info and 'train_stage_counts' in dataset_info and 'test_stage_counts' in dataset_info:
            ws_train_test = wb.create_sheet("训练测试集分布")
            
            # 添加标题
            ws_train_test.append(["训练集和测试集样本分布"])
            ws_train_test.merge_cells('A1:E1')
            ws_train_test['A1'].font = Font(bold=True, size=14)
            ws_train_test['A1'].alignment = center_alignment
            
            # 添加列标题
            ws_train_test.append(["睡眠阶段", "训练集样本数", "训练集占比", "测试集样本数", "测试集占比"])
            
            # 添加数据
            train_counts = dataset_info['train_stage_counts']
            test_counts = dataset_info['test_stage_counts']
            
            train_total = sum(train_counts.values())
            test_total = sum(test_counts.values())
            
            for i in range(6):
                stage = str(i) if isinstance(list(train_counts.keys())[0], str) else i
                train_count = train_counts.get(stage, 0)
                test_count = test_counts.get(stage, 0)
                
                train_percentage = (train_count / train_total * 100) if train_total > 0 else 0
                test_percentage = (test_count / test_total * 100) if test_total > 0 else 0
                
                ws_train_test.append([SLEEP_STAGE_LABELS[i], train_count, f"{train_percentage:.2f}%", 
                                    test_count, f"{test_percentage:.2f}%"])
            
            # 添加总计行
            ws_train_test.append(["总计", train_total, "100.00%", test_total, "100.00%"])
            
            # 应用样式
            for i, row in enumerate(ws_train_test.iter_rows(min_row=2, max_row=9, min_col=1, max_col=5)):
                for cell in row:
                    cell.border = border
                    if i == 0 or i == 7:  # 标题行和总计行
                        cell.font = header_font
                        if i == 0:
                            cell.fill = header_fill
                            
                    if i > 0:  # 数据行
                        cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = None
                
                # 获取列字母，跳过合并单元格
                for cell in col:
                    if hasattr(cell, 'column_letter'):
                        column = cell.column_letter
                        break
                
                # 如果没有找到列字母，跳过此列
                if not column:
                    continue
                    
                # 计算最大宽度
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                            
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width
        
        # 保存Excel文件
        wb.save(excel_path)
        print(f"结果摘要Excel表格已保存至: {excel_path}")
        
        # 保存混淆矩阵为图片
        try:
            plt.figure(figsize=(10, 8))
            cm = metrics['confusion_matrix']
            # 使用简化的标签用于混淆矩阵
            short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
            sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                       xticklabels=short_labels,
                       yticklabels=short_labels)
            plt.xlabel('Predicted Label')
            plt.ylabel('True Label')
            plt.title('Sleep Stage Classification Confusion Matrix')
            plt.tight_layout()
            cm_path = os.path.join(save_dir, 'confusion_matrix.png')
            plt.savefig(cm_path)
            plt.close()
            print(f"混淆矩阵已保存至: {cm_path}")
        except Exception as e:
            print(f"保存混淆矩阵图时出错: {e}")
            
    except Exception as e:
        print(f"保存Excel表格时出错: {e}")
        print(traceback.format_exc())

def update_summary_excel(results, metrics, test_data, save_dir, model_name, test_set_name, timestamp, dataset_info=None):
    """更新实验结果总表格，包含数据集信息
    
    Args:
        results: 详细的测试结果
        metrics: 性能指标
        test_data: 测试数据
        save_dir: 保存目录
        model_name: 模型名称
        test_set_name: 测试集名称
        timestamp: 时间戳
        dataset_info: 数据集信息
    """
    try:
        # 导入必要的模块
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        from datetime import datetime
        
        # 总表格文件路径
        summary_excel_path = os.path.join(save_dir, 'experiments_summary.xlsx')
        
        # 检查是否存在总表格文件
        if os.path.exists(summary_excel_path):
            # 加载现有的Excel文件
            try:
                wb = load_workbook(summary_excel_path)
                ws = wb.active
                # 检查是否需要创建表头
                create_headers = False
                if ws.max_row <= 1:  # 文件存在但可能为空
                    create_headers = True
            except Exception as e:
                print(f"加载现有总表格文件失败: {e}")
                # 创建新的Excel文件
                wb = Workbook()
                ws = wb.active
                create_headers = True
        else:
            # 创建新的Excel文件
            wb = Workbook()
            ws = wb.active
            create_headers = True
        
        # 设置工作表标题
        ws.title = "实验结果总表"
        
        # 如果是新创建的表格，添加表头
        if create_headers:
            headers = [
                "实验时间", "模型名称", "测试集名称", "总样本数", "有效预测样本数",
                "总体准确率", "宏平均F1分数", "平均推理时间(秒/样本)",
                "Wake (W) 准确率", "NREM Stage 1 (N1) 准确率", "NREM Stage 2 (N2) 准确率",
                "NREM Stage 3 (N3) 准确率", "NREM Stage 4 (N4) 准确率", "REM Sleep (R) 准确率"
            ]
            
            # 如果有数据集信息，添加训练测试集比例的列
            if dataset_info and 'train_samples' in dataset_info and 'test_samples' in dataset_info:
                headers.extend(["训练集样本数", "测试集样本数", "训练集比例", "测试集比例"])
            
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # 准备实验数据
        test_time = datetime.strptime(timestamp, '%Y%m%d_%H%M%S').strftime('%Y-%m-%d %H:%M:%S')
        total_samples = len(test_data)
        valid_predictions = len([r for r in results if r['prediction'] != -1])
        
        # 获取每个睡眠阶段的准确率
        stage_accuracies = [metrics['class_accuracies'].get(i, 0.0) for i in range(6)]
        
        # 添加本次实验数据
        row_data = [
            test_time, model_name, test_set_name, total_samples, valid_predictions,
            metrics['accuracy'], metrics['f1_macro'], metrics['avg_inference_time'],
        ] + stage_accuracies
        
        # 如果有数据集信息，添加训练测试集比例
        if dataset_info and 'train_samples' in dataset_info and 'test_samples' in dataset_info:
            train_samples = dataset_info['train_samples']
            test_samples = dataset_info['test_samples']
            total_samples_all = train_samples + test_samples
            
            train_ratio = train_samples / total_samples_all if total_samples_all > 0 else 0
            test_ratio = test_samples / total_samples_all if total_samples_all > 0 else 0
            
            # 扩展现有标题以包含新的列
            if 'train_samples' not in [cell.value for cell in ws[1]] and ws.max_row > 1:
                current_headers = [cell.value for cell in ws[1]]
                current_headers.extend(["训练集样本数", "测试集样本数", "训练集比例", "测试集比例"])
                
                # 清空行并重新添加
                for cell in ws[1]:
                    cell.value = None
                
                # 添加扩展后的标题
                for col_idx, header in enumerate(current_headers, 1):
                    ws.cell(row=1, column=col_idx).value = header
                    ws.cell(row=1, column=col_idx).font = header_font
                    ws.cell(row=1, column=col_idx).fill = header_fill
                    ws.cell(row=1, column=col_idx).alignment = center_alignment
            
            # 添加训练测试集信息
            row_data.extend([train_samples, test_samples, f"{train_ratio:.2%}", f"{test_ratio:.2%}"])
        
        ws.append(row_data)
        
        # 设置新添加行的样式
        center_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[ws.max_row]:
            cell.alignment = center_alignment
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = None
            
            # 获取列字母
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
            # 如果找到列字母，设置列宽
            if column:
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
        
        # 保存Excel文件
        wb.save(summary_excel_path)
        print(f"实验结果已追加到总表格: {summary_excel_path}")
        
    except Exception as e:
        import traceback
        print(f"更新总表格时出错: {e}")
        print(traceback.format_exc())

def get_api_prediction(client, prompt, system_prompt=None, max_retries=5, retry_delay=2):
    """通过API获取模型预测，包含重试机制"""
    for attempt in range(max_retries):
        try:
            # 构建消息
            messages = []
            if system_prompt:
                messages.append({"role": "system", "content": system_prompt})
            messages.append({"role": "user", "content": prompt})
            
            # 调用API
            start_time = time.time()
            result = client.chat.completions.create(
                messages=messages,
                model="test",
                temperature=0.1,
                max_tokens=60000,
                timeout=30  # 设置超时时间为30秒
            )
            end_time = time.time()
            
            # 获取响应
            response = result.choices[0].message.content
            
            # 提取数字
            prediction = None
            for char in response:
                if char.isdigit() and int(char) in [0, 1, 2, 3, 4, 5]:
                    prediction = int(char)
                    break
            
            return response, prediction, end_time - start_time
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"API调用出错 (尝试 {attempt+1}/{max_retries}): {e}，等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"API调用失败，已达到最大重试次数 ({max_retries}): {e}")
                return "API调用失败", None, 0

def calculate_metrics(true_labels, pred_labels, total_time):
    """计算评估指标"""
    # 过滤掉无效预测
    valid_indices = [i for i, pred in enumerate(pred_labels) if pred != -1]
    if not valid_indices:
        return {
            'accuracy': 0.0,
            'f1_macro': 0.0,
            'avg_inference_time': 0.0,
            'confusion_matrix': np.zeros((6, 6), dtype=int),
            'class_accuracies': {i: 0.0 for i in range(6)}
        }
    
    valid_true = [true_labels[i] for i in valid_indices]
    valid_pred = [pred_labels[i] for i in valid_indices]
    
    # 计算总体指标
    accuracy = accuracy_score(valid_true, valid_pred)
    f1_macro = f1_score(valid_true, valid_pred, average='macro', zero_division=0)
    avg_inference_time = total_time / len(true_labels) if true_labels else 0
    confusion_matrix_result = confusion_matrix(valid_true, valid_pred, labels=range(6))
    
    # 计算每个类别的准确率
    class_accuracies = {}
    for class_idx in range(6):
        # 找出真实标签为该类别的样本索引
        class_indices = [i for i, label in enumerate(valid_true) if label == class_idx]
        if class_indices:  # 如果有该类别的样本
            # 计算该类别的准确率
            correct = sum(1 for i in class_indices if valid_pred[i] == class_idx)
            class_accuracies[class_idx] = correct / len(class_indices)
        else:
            class_accuracies[class_idx] = 0.0
    
    return {
        'accuracy': accuracy,
        'f1_macro': f1_macro,
        'avg_inference_time': avg_inference_time,
        'confusion_matrix': confusion_matrix_result,
        'class_accuracies': class_accuracies
    }

def evaluate_model(client, test_data, print_interval=10, save_interval=500, save_dir='/data/lhc/results', 
                model_name="unknown_model", test_set_name="unknown_dataset"):
    """评估模型性能
    
    Args:
        client: OpenAI客户端
        test_data: 测试数据
        print_interval: 打印中间结果的间隔
        save_interval: 保存中间结果的间隔
        save_dir: 保存结果的目录
        model_name: 模型名称
        test_set_name: 测试集名称
        
    Returns:
        final_metrics: 包含所有指标的字典，额外包含'results'键存储详细结果
    """
    # 创建一个包含模型名称和测试集名称的文件夹来保存结果
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    result_dir = os.path.join(save_dir, f"{model_name}_{test_set_name}_{timestamp}")
    os.makedirs(result_dir, exist_ok=True)
    
    # 保存中间结果的函数
    def save_intermediate_results(current_idx, current_metrics, current_results):
        """保存中间结果
        
        Args:
            current_idx: 当前处理的样本索引
            current_metrics: 当前的评估指标
            current_results: 当前的详细结果
        """
        intermediate_dir = os.path.join(result_dir, f"intermediate_{current_idx}")
        os.makedirs(intermediate_dir, exist_ok=True)
        
        # 保存模型和测试集信息
        with open(os.path.join(intermediate_dir, "info.json"), "w") as f:
            json.dump({
                "model_name": model_name,
                "test_set_name": test_set_name,
                "samples_processed": current_idx,
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }, f, indent=4)
        
        # 保存指标
        with open(os.path.join(intermediate_dir, "metrics.json"), "w") as f:
            json.dump({k: v if not isinstance(v, np.ndarray) else v.tolist() 
                      for k, v in current_metrics.items()}, f, indent=4)
        
        # 保存结果
        with open(os.path.join(intermediate_dir, "results.json"), "w") as f:
            json.dump(current_results, f, indent=4)
        
        # 保存混淆矩阵
        plt.figure(figsize=(10, 8))
        cm = current_metrics['confusion_matrix']
        # 使用简化的标签用于混淆矩阵
        short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=short_labels,
                    yticklabels=short_labels)
        plt.xlabel('Predicted Label')
        plt.ylabel('True Label')
        plt.title(f'Intermediate Confusion Matrix (Samples: {current_idx})')
        plt.savefig(os.path.join(intermediate_dir, "confusion_matrix.png"))
        plt.close()
        
        # 保存Excel文件
        save_results_to_excel(current_results, current_metrics, test_data, intermediate_dir, model_name, test_set_name)
        
        print(f"\n中间结果已保存至: {intermediate_dir}")
    
    print("开始评估模型性能...")
    results = []
    true_labels = []
    pred_labels = []
    responses = []
    
    # 系统提示
    system_prompt = (
        "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification. "
        "Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) "
        "and determine the current sleep stage of the volunteer based on the following classification criteria:\n"
        "0: Wakefulness (W)\n"
        "1: Non-rapid eye movement sleep stage 1 (N1)\n"
        "2: Non-rapid eye movement sleep stage 2 (N2)\n"
        "3: Non-rapid eye movement sleep stage 3 (N3)\n"
        "4: Non-rapid eye movement sleep stage 4 (N4)\n"
        "5: Rapid eye movement sleep stage (R)\n"
        "The EEG data is provided in the format (time in milliseconds, Fpz-Cz voltage in μV, Pz-Oz voltage in μV). "
        "The data spans 1000ms with a sampling interval of 5ms. In your analysis, pay attention to the following "
        "characteristics of each sleep stage:\n"
        "- Wakefulness (W): High-frequency, low-amplitude waves.\n"
        "- N1: Low-amplitude, mixed-frequency waves.\n"
        "- N2: Sleep spindles and K-complexes.\n"
        "- N3: High-amplitude, low-frequency delta waves.\n"
        "- N4: Dominant delta waves.\n"
        "- REM (R): REM sleep has highly distinctive and unique characteristics. It primarily presents with rapid, irregular eye movements visible in EEG as sharp, jagged waveforms. Its core feature is low-amplitude, mixed-frequency EEG activity with prominent theta waves (4-7 Hz). While somewhat similar to N1 stage, REM has distinctive saw-tooth wave patterns, which are key diagnostic markers. Unlike N2 stage, REM lacks sleep spindles and K-complexes. The EEG in REM shows a desynchronized pattern resembling wakefulness but is accompanied by complete loss of muscle tone (muscle atonia). REM may also feature rapid, irregular transient muscle twitches, along with irregular variations in heart rate and respiration. These multiple features collectively constitute the complete picture of REM sleep, making it the most distinctive and readily identifiable among all sleep stages.\n"
        "Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. "
        "Do not include any additional text, punctuation, or explanations."
    )
    
    # 初始化指标
    metrics = {
        'total_samples': 0,
        'correct_samples': 0,
        'accuracy': 0.0,
        'avg_inference_time': 0.0,
        'confusion_matrix': np.zeros((6, 6), dtype=int)
    }
    
    # 使用tqdm创建进度条
    pbar = tqdm(test_data, desc="Test Samples", ncols=100)
    
    # 用于实时显示性能指标
    correct_count = 0
    total_count = 0
    total_time = 0.0
    
    for idx, data in enumerate(pbar):
        human_prompt = data["input"]
        true_label = int(data["output"])
        
        try:
            # 获取模型预测
            response, prediction, inference_time = get_api_prediction(
                client, human_prompt, system_prompt
            )
            
            total_time += inference_time
            
            if prediction is not None:
                correct = prediction == true_label
                if correct:
                    correct_count += 1
            else:
                correct = False
                prediction = -1  # 表示无法从回答中提取预测值
            
            total_count += 1
            
            # 更新进度条状态
            current_acc = correct_count / total_count if total_count > 0 else 0
            avg_time = total_time / total_count if total_count > 0 else 0
            
            # 设置进度条描述，实时显示准确率和平均推理时间
            pbar.set_postfix({
                'acc': f'{current_acc:.2%}',
                'avg_time': f'{avg_time:.2f}s',
                '当前预测': f'{prediction}',
                '真实标签': f'{true_label}',
                '正确': correct
            })
            
            result = {
                "id": idx,
                "true_label": true_label,
                "prediction": prediction,
                "response": response,
                "correct": correct,
                "inference_time": inference_time
            }
            
            results.append(result)
            true_labels.append(true_label)
            pred_labels.append(prediction if prediction != -1 else 0)  # 对于无法提取预测的情况，默认为0
            responses.append(response)
            
            # 更新指标
            metrics['total_samples'] += 1
            if correct:
                metrics['correct_samples'] += 1
            if prediction is not None:
                metrics['confusion_matrix'][true_label, prediction] += 1
            
            # 每隔print_interval次测试打印一次相关的测试数据
            if (idx + 1) % print_interval == 0:
                # 计算当前指标
                current_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                print("\n" + "-" * 50)
                print(f"已完成 {idx + 1} 个样本的测试:")
                print(f"总准确率: {current_metrics['accuracy']:.2%}")
                print(f"宏平均F1分数: {current_metrics['f1_macro']:.2%}")
                print(f"平均推理时间: {current_metrics['avg_inference_time']:.4f}秒/样本")
                
                # 打印每个标签的准确率
                print("\n各睡眠阶段准确率:")
                for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
                    acc = current_metrics['class_accuracies'][class_idx]
                    print(f"  {label_name}: {acc:.2%}")
                print("-" * 50 + "\n")
                
            # 每隔save_interval个样本保存一次中间结果
            if (idx + 1) % save_interval == 0:
                current_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                save_intermediate_results(idx + 1, current_metrics, results)
            
        except Exception as e:
            print(f"处理样本 {idx} 时出错: {str(e)}")
            continue  # 跳过失败样本
    
    # 计算最终指标
    final_metrics = calculate_metrics(true_labels, pred_labels, total_time)
    
    # 添加结果到指标中，方便后续处理
    final_metrics['results'] = results
    
    # 打印最终结果
    print("\n" + "=" * 50)
    print("睡眠分期测试结果")
    print("=" * 50)
    print(f"总样本数: {len(test_data)}")
    print(f"有效预测样本数: {len([i for i, pred in enumerate(pred_labels) if pred != -1])}")
    print(f"总准确率: {final_metrics['accuracy']:.2%}")
    print(f"宏平均F1分数: {final_metrics['f1_macro']:.2%}")
    print(f"平均推理时间: {final_metrics['avg_inference_time']:.4f}秒/样本")
    
    # 打印每个标签的准确率
    print("\n各睡眠阶段准确率:")
    for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
        acc = final_metrics['class_accuracies'][class_idx]
        print(f"  {label_name}: {acc:.2%}")
    
    # 打印混淆矩阵
    print("\n混淆矩阵:")
    cm = final_metrics['confusion_matrix']
    labels = ['W', 'N1', 'N2', 'N3', 'N4', 'REM']
    print("    " + " ".join(f"{label:^5s}" for label in labels))
    for i, label in enumerate(labels):
        print(f"{label:^5s}" + " ".join(f"{cm[i, j]:^5d}" for j in range(len(labels))))
    
    print("=" * 50)
    
    return final_metrics

def save_results(results, metrics, test_data, save_dir, model_name="unknown_model", test_set_name="unknown_dataset", dataset_info=None, test_data_path=None):
    """保存测试结果到指定目录，包含训练集和测试集的详细信息
    
    Args:
        results: 详细的测试结果
        metrics: 性能指标
        test_data: 测试数据
        save_dir: 保存目录
        model_name: 模型名称
        test_set_name: 测试集名称
        dataset_info: 数据集信息，如果为None则尝试加载
        test_data_path: 测试数据路径，用于加载数据集信息
    """
    import os
    import json
    import pandas as pd
    import matplotlib.pyplot as plt
    import seaborn as sns
    from datetime import datetime
    import traceback
    
    # 如果没有提供数据集信息但提供了测试数据路径，尝试加载
    if dataset_info is None and test_data_path is not None:
        dataset_info = load_dataset_info(test_data_path)
    
    # 创建保存目录，包含模型名称和测试集名称
    os.makedirs(save_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    result_dir = os.path.join(save_dir, f'{model_name}_{test_set_name}_{timestamp}')
    os.makedirs(result_dir, exist_ok=True)
    
    # 保存详细结果为CSV
    results_df = pd.DataFrame(results)
    results_csv_path = os.path.join(result_dir, 'detailed_results.csv')
    results_df.to_csv(results_csv_path, index=False)
    print(f"详细结果已保存至: {results_csv_path}")
    
    # 保存性能指标为JSON
    metrics_json = {
        'model_name': model_name,
        'test_set_name': test_set_name,
        'total_samples': len(test_data),
        'valid_predictions': len([r for r in results if r['prediction'] != -1]),
        'accuracy': metrics['accuracy'],
        'f1_macro': metrics['f1_macro'],
        'avg_inference_time': metrics['avg_inference_time'],
        'class_accuracies': {str(k): v for k, v in metrics['class_accuracies'].items()},
        'timestamp': timestamp
    }
    
    # 如果有数据集信息，添加到指标中
    if dataset_info:
        metrics_json['dataset_info'] = dataset_info
    
    metrics_path = os.path.join(result_dir, 'metrics.json')
    with open(metrics_path, 'w', encoding='utf-8') as f:
        json.dump(metrics_json, f, indent=4, ensure_ascii=False)
    print(f"性能指标已保存至: {metrics_path}")
    
    # 保存混淆矩阵为图片
    plt.figure(figsize=(10, 8))
    cm = metrics['confusion_matrix']
    # 使用简化的标签用于混淆矩阵
    short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                xticklabels=short_labels,
                yticklabels=short_labels)
    plt.xlabel('Predicted Label')
    plt.ylabel('True Label')
    plt.title('Sleep Stage Classification Confusion Matrix')
    plt.tight_layout()
    cm_path = os.path.join(result_dir, 'confusion_matrix.png')
    plt.savefig(cm_path)
    plt.close()
    print(f"混淆矩阵已保存至: {cm_path}")
    
    # 保存测试样本分布统计
    class_counts = {}
    for r in results:
        label = r['true_label']
        if label not in class_counts:
            class_counts[label] = 0
        class_counts[label] += 1
    
    # 绘制样本分布饼图
    plt.figure(figsize=(10, 8))
    sizes = [class_counts.get(i, 0) for i in range(6)]
    plt.pie(sizes, labels=SLEEP_STAGE_LABELS, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.title('Test Sample Distribution')
    dist_path = os.path.join(result_dir, 'sample_distribution.png')
    plt.savefig(dist_path)
    plt.close()
    print(f"样本分布图已保存至: {dist_path}")
    
    # 保存性能指标摘要为文本文件
    summary_path = os.path.join(result_dir, 'summary.txt')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("睡眠阶段分类测试结果摘要\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"模型: {model_name}\n")
        f.write(f"测试集: {test_set_name}\n\n")
        
        # 添加数据集信息
        if dataset_info:
            f.write("数据集信息:\n")
            f.write("-" * 60 + "\n")
            
            # 如果有训练集和测试集的比例信息
            if 'train_samples' in dataset_info and 'test_samples' in dataset_info:
                train_samples = dataset_info['train_samples']
                test_samples = dataset_info['test_samples']
                total_samples = train_samples + test_samples
                
                train_ratio = train_samples / total_samples if total_samples > 0 else 0
                test_ratio = test_samples / total_samples if total_samples > 0 else 0
                
                f.write(f"训练集样本数: {train_samples} ({train_ratio:.2%})\n")
                f.write(f"测试集样本数: {test_samples} ({test_ratio:.2%})\n")
                f.write(f"总样本数: {total_samples}\n\n")
            
            # 如果有详细的阶段分布信息
            if 'train_stage_counts' in dataset_info and 'test_stage_counts' in dataset_info:
                f.write("各睡眠阶段样本分布:\n")
                f.write(f"{'睡眠阶段':^25} | {'训练集':^10} | {'训练集占比':^10} | {'测试集':^10} | {'测试集占比':^10}\n")
                f.write("-" * 75 + "\n")
                
                train_counts = dataset_info['train_stage_counts']
                test_counts = dataset_info['test_stage_counts']
                
                train_total = sum(train_counts.values())
                test_total = sum(test_counts.values())
                
                stage_names = {
                    "0": "Wake (W)",
                    "1": "NREM Stage 1 (N1)",
                    "2": "NREM Stage 2 (N2)",
                    "3": "NREM Stage 3 (N3)",
                    "4": "NREM Stage 4 (N4)",
                    "5": "REM Sleep (R)"
                }
                
                for stage in sorted([str(i) for i in range(6)]):
                    train_count = train_counts.get(int(stage), 0)
                    test_count = test_counts.get(int(stage), 0)
                    
                    train_percentage = (train_count / train_total * 100) if train_total > 0 else 0
                    test_percentage = (test_count / test_total * 100) if test_total > 0 else 0
                    
                    stage_name = stage_names.get(stage, f"Stage {stage}")
                    f.write(f"{stage}: {stage_name:20} | {train_count:^10} | {train_percentage:^10.2f}% | {test_count:^10} | {test_percentage:^10.2f}%\n")
                
                f.write("-" * 75 + "\n")
                f.write(f"{'总计':^25} | {train_total:^10} | {'100.00':^10}% | {test_total:^10} | {'100.00':^10}%\n\n")
        
        # 当前测试情况
        f.write("当前测试数据统计:\n")
        f.write("-" * 60 + "\n")
        f.write(f"总样本数: {len(test_data)}\n")
        f.write(f"有效预测样本数: {len([r for r in results if r['prediction'] != -1])}\n\n")
        
        f.write("睡眠阶段样本分布:\n")
        for i in range(6):
            count = class_counts.get(i, 0)
            percentage = (count / len(test_data)) * 100 if len(test_data) > 0 else 0
            label_name = SLEEP_STAGE_LABELS[i]
            f.write(f"Stage {i} ({label_name}): {count} samples ({percentage:.2f}%)\n")
        f.write("\n")
        
        f.write("性能指标:\n")
        f.write("-" * 60 + "\n")
        f.write(f"总体准确率: {metrics['accuracy']:.4f}\n")
        f.write(f"宏平均F1分数: {metrics['f1_macro']:.4f}\n")
        f.write(f"平均推理时间: {metrics['avg_inference_time']:.4f} 秒/样本\n\n")
        
        f.write("各睡眠阶段准确率:\n")
        for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
            acc = metrics['class_accuracies'][class_idx]
            f.write(f"  {label_name}: {acc:.4f}\n")
    print(f"测试摘要已保存至: {summary_path}")
    
    # 调用函数保存结果为Excel表格
    save_results_to_excel(results, metrics, test_data, result_dir, model_name, test_set_name, dataset_info)
    
    # 更新总表格文件
    update_summary_excel(results, metrics, test_data, save_dir, model_name, test_set_name, timestamp, dataset_info)

if __name__ == '__main__':
    # 设置API端口
    port = os.environ.get("API_PORT", 8000)
    
    # 初始化OpenAI客户端
    client = OpenAI(
        api_key="0",
        base_url=f"http://localhost:{port}/v1",
    )
    
    # 获取测试数据路径和模型名称
    test_data_path = os.environ.get("TEST_DATA_PATH", "/data/lhc/datasets_new/sleep/test/edf100_200hz_7500ms_tok12588_balanced_0.1_sqrt_inverse_test.json")
    test_data = load_test_data(test_data_path)
    
    # 加载数据集信息
    dataset_info = load_dataset_info(test_data_path)
    
    # 从测试数据路径中提取测试集名称，并去掉数据量信息
    file_basename = os.path.basename(test_data_path).split('.')[0]
    # 假设数据量信息格式为"_n数字"，需要移除这部分
    import re
    test_set_name = re.sub(r'_n\d+', '', file_basename)
    
    # 尝试从环境变量或配置中获取模型名称，如果没有则使用默认值
    model_name = os.environ.get("MODEL_NAME", "llama_edf")
    
    print(f"模型: {model_name}")
    print(f"测试集: {test_set_name}")
    print(f"加载了 {len(test_data)} 个测试样本")
    
    # 如果有数据集信息，打印训练集和测试集信息
    if dataset_info:
        print("\n数据集信息:")
        if 'train_samples' in dataset_info and 'test_samples' in dataset_info:
            train_samples = dataset_info['train_samples']
            test_samples = dataset_info['test_samples']
            total_samples = train_samples + test_samples
            
            train_ratio = train_samples / total_samples if total_samples > 0 else 0
            test_ratio = test_samples / total_samples if total_samples > 0 else 0
            
            print(f"训练集样本数: {train_samples} ({train_ratio:.2%})")
            print(f"测试集样本数: {test_samples} ({test_ratio:.2%})")
            print(f"总样本数: {total_samples}")
            
            # 如果有阶段分布信息，打印每个阶段的样本数量和比例
            if 'train_stage_counts' in dataset_info and 'test_stage_counts' in dataset_info:
                print("\n各睡眠阶段样本分布:")
                train_counts = dataset_info['train_stage_counts']
                test_counts = dataset_info['test_stage_counts']
                
                train_total = sum(train_counts.values())
                test_total = sum(test_counts.values())
                
                print(f"{'睡眠阶段':^25} | {'训练集':^10} | {'训练集占比':^10} | {'测试集':^10} | {'测试集占比':^10}")
                print("-" * 75)
                
                for i in range(6):
                    stage = str(i) if isinstance(list(train_counts.keys())[0], str) else i
                    train_count = train_counts.get(stage, 0)
                    test_count = test_counts.get(stage, 0)
                    
                    train_percentage = (train_count / train_total * 100) if train_total > 0 else 0
                    test_percentage = (test_count / test_total * 100) if test_total > 0 else 0
                    
                    print(f"{SLEEP_STAGE_LABELS[i]:^25} | {train_count:^10} | {train_percentage:^10.2f}% | {test_count:^10} | {test_percentage:^10.2f}%")
                
                print("-" * 75)
                print(f"{'总计':^25} | {train_total:^10} | {'100.00':^10}% | {test_total:^10} | {'100.00':^10}%")
                print()
    
    # 创建保存结果的目录
    save_dir = os.environ.get("SAVE_DIR", "/data/lhc/results")
    os.makedirs(save_dir, exist_ok=True)
    
    # 设置打印和保存的间隔
    print_interval = int(os.environ.get("PRINT_INTERVAL", 10))
    save_interval = int(os.environ.get("SAVE_INTERVAL", 500))
    
    # 评估模型
    metrics = evaluate_model(client, test_data, print_interval=print_interval, save_interval=save_interval, save_dir=save_dir, 
                            model_name=model_name, test_set_name=test_set_name)
    
    # 保存结果，并传递数据集信息和测试数据路径
    save_results(results=metrics['results'] if 'results' in metrics else [], 
                metrics=metrics, 
                test_data=test_data, 
                save_dir=save_dir, 
                model_name=model_name, 
                test_set_name=test_set_name, 
                dataset_info=dataset_info, 
                test_data_path=test_data_path)
