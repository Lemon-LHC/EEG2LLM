import os
import json
import time
import numpy as np
import pandas as pd
import traceback
import datetime
from tqdm import tqdm
from openai import OpenAI
from transformers.utils.versions import require_version
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix

require_version("openai>=1.5.0", "To fix: pip install openai>=1.5.0")

# 定义睡眠阶段标签，避免重复硬编码
SLEEP_STAGE_LABELS = [
    'Wake (W)', 
    'NREM Stage 1 (N1)', 
    'NREM Stage 2 (N2)', 
    'NREM Stage 3 (N3)', 
    'NREM Stage 4 (N4)', 
    'REM Sleep (R)'
]


def load_test_data(file_path, max_samples=None):
    """加载测试数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        test_data = json.load(f)
    
    # 如果设置了最大样本数，随机选择样本
    if max_samples and max_samples < len(test_data):
        indices = np.random.choice(len(test_data), max_samples, replace=False)
        test_data = [test_data[i] for i in indices]
    
    return test_data


def get_api_prediction(client, prompt, system_prompt=None, max_retries=5, retry_delay=2):
    """通过API获取模型预测，包含重试机制"""
    for attempt in range(max_retries):
        try:
            # 构建消息
            messages = []
            if system_prompt:
                messages.append({"role": "system", "content": system_prompt})
            messages.append({"role": "user", "content": prompt})
            
            # 调用API
            start_time = time.time()
            result = client.chat.completions.create(
                messages=messages,
                model="test",
                temperature=0.1,
                max_tokens=60000,
                timeout=30  # 设置超时时间为30秒
            )
            end_time = time.time()
            
            # 获取响应
            response = result.choices[0].message.content
            
            # 提取数字
            prediction = None
            for char in response:
                if char.isdigit() and int(char) in [0, 1, 2, 3, 4, 5]:
                    prediction = int(char)
                    break
            
            return response, prediction, end_time - start_time
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"API调用出错 (尝试 {attempt+1}/{max_retries}): {e}，等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"API调用失败，已达到最大重试次数 ({max_retries}): {e}")
                return "API调用失败", None, 0


def evaluate_model(client, test_data, print_interval=10, save_interval=500, save_dir='/data/lhc/results', 
                model_name="unknown_model", test_set_name="unknown_dataset"):
    """评估模型性能
    
    Args:
        client: OpenAI客户端
        test_data: 测试数据
        print_interval: 打印中间结果的间隔
        save_interval: 保存中间结果的间隔
        save_dir: 保存结果的目录
        model_name: 模型名称
        test_set_name: 测试集名称
    """
    import matplotlib.pyplot as plt
    import seaborn as sns
    import os
    
    # 创建一个包含模型名称和测试集名称的文件夹来保存结果
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    result_dir = os.path.join(save_dir, f"{model_name}_{test_set_name}_{timestamp}")
    os.makedirs(result_dir, exist_ok=True)
    
    # 保存中间结果的函数
    def save_intermediate_results(current_idx, current_metrics, current_results):
        """保存中间结果
        
        Args:
            current_idx: 当前处理的样本索引
            current_metrics: 当前的评估指标
            current_results: 当前的详细结果
        """
        intermediate_dir = os.path.join(result_dir, f"intermediate_{current_idx}")
        os.makedirs(intermediate_dir, exist_ok=True)
        
        # 保存模型和测试集信息
        with open(os.path.join(intermediate_dir, "info.json"), "w") as f:
            json.dump({
                "model_name": model_name,
                "test_set_name": test_set_name,
                "samples_processed": current_idx,
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }, f, indent=4)
        
        # 保存指标
        with open(os.path.join(intermediate_dir, "metrics.json"), "w") as f:
            json.dump({k: v if not isinstance(v, np.ndarray) else v.tolist() 
                      for k, v in current_metrics.items()}, f, indent=4)
        
        # 保存结果
        with open(os.path.join(intermediate_dir, "results.json"), "w") as f:
            json.dump(current_results, f, indent=4)
        
        # 保存混淆矩阵
        plt.figure(figsize=(10, 8))
        cm = current_metrics['confusion_matrix']
        # 使用简化的标签用于混淆矩阵
        short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=short_labels,
                    yticklabels=short_labels)
        plt.xlabel('Predicted Label')
        plt.ylabel('True Label')
        plt.title(f'Intermediate Confusion Matrix (Samples: {current_idx})')
        plt.savefig(os.path.join(intermediate_dir, "confusion_matrix.png"))
        plt.close()
        
        # 保存Excel文件
        save_results_to_excel(current_results, current_metrics, test_data, intermediate_dir, model_name, test_set_name)
        
        print(f"\n中间结果已保存至: {intermediate_dir}")
    
    print("开始评估模型性能...")
    results = []
    true_labels = []
    pred_labels = []
    responses = []
    
    # 系统提示
    system_prompt = (
        "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification. "
        "Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) "
        "and determine the current sleep stage of the volunteer based on the following classification criteria:\n"
        "0: Wakefulness (W)\n"
        "1: Non-rapid eye movement sleep stage 1 (N1)\n"
        "2: Non-rapid eye movement sleep stage 2 (N2)\n"
        "3: Non-rapid eye movement sleep stage 3 (N3)\n"
        "4: Non-rapid eye movement sleep stage 4 (N4)\n"
        "5: Rapid eye movement sleep stage (R)\n"
        "The EEG data is provided in the format (time in milliseconds, Fpz-Cz voltage in μV, Pz-Oz voltage in μV). "
        "The data spans 1000ms with a sampling interval of 5ms. In your analysis, pay attention to the following "
        "characteristics of each sleep stage:\n"
        "- Wakefulness (W): High-frequency, low-amplitude waves.\n"
        "- N1: Low-amplitude, mixed-frequency waves.\n"
        "- N2: Sleep spindles and K-complexes.\n"
        "- N3: High-amplitude, low-frequency delta waves.\n"
        "- N4: Dominant delta waves.\n"
        "- REM (R): REM sleep has highly distinctive and unique characteristics. It primarily presents with rapid, irregular eye movements visible in EEG as sharp, jagged waveforms. Its core feature is low-amplitude, mixed-frequency EEG activity with prominent theta waves (4-7 Hz). While somewhat similar to N1 stage, REM has distinctive saw-tooth wave patterns, which are key diagnostic markers. Unlike N2 stage, REM lacks sleep spindles and K-complexes. The EEG in REM shows a desynchronized pattern resembling wakefulness but is accompanied by complete loss of muscle tone (muscle atonia). REM may also feature rapid, irregular transient muscle twitches, along with irregular variations in heart rate and respiration. These multiple features collectively constitute the complete picture of REM sleep, making it the most distinctive and readily identifiable among all sleep stages.\n"
        "Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. "
        "Do not include any additional text, punctuation, or explanations."
    )
    
    # 初始化指标
    metrics = {
        'total_samples': 0,
        'correct_samples': 0,
        'accuracy': 0.0,
        'avg_inference_time': 0.0,
        'confusion_matrix': np.zeros((6, 6), dtype=int)
    }
    
    # 使用tqdm创建进度条
    pbar = tqdm(test_data, desc="Test Samples", ncols=100)
    
    # 用于实时显示性能指标
    correct_count = 0
    total_count = 0
    total_time = 0.0
    
    for idx, data in enumerate(pbar):
        human_prompt = data["input"]
        true_label = int(data["output"])
        
        try:
            # 获取模型预测
            response, prediction, inference_time = get_api_prediction(
                client, human_prompt, system_prompt
            )
            
            total_time += inference_time
            
            if prediction is not None:
                correct = prediction == true_label
                if correct:
                    correct_count += 1
            else:
                correct = False
                prediction = -1  # 表示无法从回答中提取预测值
            
            total_count += 1
            
            # 更新进度条状态
            current_acc = correct_count / total_count if total_count > 0 else 0
            avg_time = total_time / total_count if total_count > 0 else 0
            
            # 设置进度条描述，实时显示准确率和平均推理时间
            pbar.set_postfix({
                'acc': f'{current_acc:.2%}',
                'avg_time': f'{avg_time:.2f}s',
                '当前预测': f'{prediction}',
                '真实标签': f'{true_label}',
                '正确': correct
            })
            
            result = {
                "id": idx,
                "true_label": true_label,
                "prediction": prediction,
                "response": response,
                "correct": correct,
                "inference_time": inference_time
            }
            
            results.append(result)
            true_labels.append(true_label)
            pred_labels.append(prediction if prediction != -1 else 0)  # 对于无法提取预测的情况，默认为0
            responses.append(response)
            
            # 更新指标
            metrics['total_samples'] += 1
            if correct:
                metrics['correct_samples'] += 1
            if prediction is not None:
                metrics['confusion_matrix'][true_label, prediction] += 1
            
            # 每隔print_interval次测试打印一次相关的测试数据
            if (idx + 1) % print_interval == 0:
                # 计算当前指标
                current_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                print("\n" + "-" * 50)
                print(f"已完成 {idx + 1} 个样本的测试:")
                print(f"总准确率: {current_metrics['accuracy']:.2%}")
                print(f"宏平均F1分数: {current_metrics['f1_macro']:.2%}")
                print(f"平均推理时间: {current_metrics['avg_inference_time']:.4f}秒/样本")
                
                # 打印每个标签的准确率
                print("\n各睡眠阶段准确率:")
                for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
                    acc = current_metrics['class_accuracies'][class_idx]
                    print(f"  {label_name}: {acc:.2%}")
                print("-" * 50 + "\n")
                
            # 每隔save_interval个样本保存一次中间结果
            if (idx + 1) % save_interval == 0:
                current_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                save_intermediate_results(idx + 1, current_metrics, results)
            
        except Exception as e:
            print(f"处理样本 {idx} 时出错: {str(e)}")
            continue  # 跳过失败样本
    
    # 计算最终指标
    final_metrics = calculate_metrics(true_labels, pred_labels, total_time)
    
    # 打印最终结果
    print("\n" + "=" * 50)
    print("睡眠分期测试结果")
    print("=" * 50)
    print(f"总样本数: {len(test_data)}")
    print(f"有效预测样本数: {len([i for i, pred in enumerate(pred_labels) if pred != -1])}")
    print(f"总准确率: {final_metrics['accuracy']:.2%}")
    print(f"宏平均F1分数: {final_metrics['f1_macro']:.2%}")
    print(f"平均推理时间: {final_metrics['avg_inference_time']:.4f}秒/样本")
    
    # 打印每个标签的准确率
    print("\n各睡眠阶段准确率:")
    for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
        acc = final_metrics['class_accuracies'][class_idx]
        print(f"  {label_name}: {acc:.2%}")
    
    # 打印混淆矩阵
    print("\n混淆矩阵:")
    cm = final_metrics['confusion_matrix']
    labels = ['W', 'N1', 'N2', 'N3', 'N4', 'REM']
    print("    " + " ".join(f"{label:^5s}" for label in labels))
    for i, label in enumerate(labels):
        print(f"{label:^5s}" + " ".join(f"{cm[i, j]:^5d}" for j in range(len(labels))))
    
    print("=" * 50)
    
    # 保存结果到指定目录
    save_results(results, final_metrics, test_data, save_dir, model_name=model_name, test_set_name=test_set_name)
    
    return final_metrics


def save_results(results, metrics, test_data, save_dir, model_name="unknown_model", test_set_name="unknown_dataset"):
    """保存测试结果到指定目录
    
    Args:
        results: 详细的测试结果
        metrics: 性能指标
        test_data: 测试数据
        save_dir: 保存目录
        model_name: 模型名称
        test_set_name: 测试集名称
    """
    import os
    import json
    import pandas as pd
    import matplotlib.pyplot as plt
    import seaborn as sns
    from datetime import datetime
    import torch
    import traceback
    import glob
    
    # 创建保存目录，包含模型名称和测试集名称
    os.makedirs(save_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    result_dir = os.path.join(save_dir, f'{model_name}_{test_set_name}_{timestamp}')
    os.makedirs(result_dir, exist_ok=True)
    
    # TensorBoard写入设置
    try:
        from torch.utils.tensorboard import SummaryWriter
        
        # 创建与训练模型相关的TensorBoard路径（通常在saves目录下）
        model_dir_parts = model_name.split('_')
        base_tb_dir = None
        
        # 检查是否指定了tensorboard_dir环境变量
        tensorboard_dir = os.environ.get("TENSORBOARD_DIR", None)
        
        if not tensorboard_dir:
            # 尝试从模型名推断TensorBoard目录
            for i in range(len(model_dir_parts)):
                possible_dir = '_'.join(model_dir_parts[:i+1])
                possible_path = f"/data/lhc/saves/{possible_dir}"
                if os.path.exists(possible_path):
                    base_tb_dir = possible_path
                    break
            
            # 如果无法推断，使用默认保存目录
            if not base_tb_dir:
                base_tb_dir = save_dir
                
            print(f"推断的TensorBoard目录: {base_tb_dir}")
        else:
            base_tb_dir = tensorboard_dir
            print(f"使用环境变量指定的TensorBoard目录: {base_tb_dir}")
        
        # 创建TensorBoard测试子目录
        test_tb_dir = os.path.join(base_tb_dir, f"test_{test_set_name}_{timestamp}")
        os.makedirs(test_tb_dir, exist_ok=True)
        print(f"创建测试集TensorBoard子目录: {test_tb_dir}")
        
        # 确保目录权限正确
        os.system(f"chmod -R 755 {test_tb_dir}")
        
        # 获取当前步骤（可以通过模型名称中的检查点编号提取，或使用默认值）
        checkpoint_step = 0
        if 'checkpoint-' in model_name:
            try:
                checkpoint_name = model_name.split('/')[-1]
                checkpoint_step = int(checkpoint_name.split('-')[-1])
            except:
                pass
        
        # 初始化TensorBoard写入器
        tb_writer = SummaryWriter(log_dir=test_tb_dir)
        print(f"成功创建TensorBoard写入器，指向目录: {test_tb_dir}")
        
        # 写入基本信息
        import socket
        host_info = f"{socket.gethostname()}_{os.getpid()}"
        tb_writer.add_text('eval_info/model', model_name, 0)
        tb_writer.add_text('eval_info/test_data', test_set_name, 0)
        tb_writer.add_text('eval_info/host', host_info, 0)
        tb_writer.add_text('eval_info/time', str(datetime.now()), 0)
        
        # 写入主要指标
        print("正在写入测试集指标到TensorBoard...")
        tb_writer.add_scalar('test/accuracy', metrics['accuracy'], checkpoint_step)
        tb_writer.add_scalar('test/f1', metrics['f1_macro'], checkpoint_step)
        tb_writer.add_scalar('test/avg_inference_time', metrics['avg_inference_time'], checkpoint_step)
        
        # 记录各个类别的指标
        for class_idx in range(6):
            stage_name = SLEEP_STAGE_LABELS[class_idx].split(' ')[0]  # 例如: 'Wake' -> 'W'
            class_acc = metrics['class_accuracies'].get(class_idx, 0.0)
            
            tb_writer.add_scalar(f'test/Class_{stage_name}/Accuracy', class_acc, checkpoint_step)
        
        # 添加混淆矩阵作为图像
        fig = plt.figure(figsize=(10, 8))
        cm = metrics['confusion_matrix']
        short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=short_labels,
                    yticklabels=short_labels)
        plt.xlabel('预测标签')
        plt.ylabel('真实标签')
        plt.title(f'混淆矩阵 - {test_set_name} - {model_name}')
        tb_writer.add_figure('test/ConfusionMatrix', fig, checkpoint_step)
        
        # 添加样本分布图
        class_counts = {}
        for r in results:
            label = r['true_label']
            if label not in class_counts:
                class_counts[label] = 0
            class_counts[label] += 1
            
        stage_samples = [class_counts.get(i, 0) for i in range(6)]
        fig = plt.figure(figsize=(10, 6))
        plt.bar([label.split(' ')[0] for label in SLEEP_STAGE_LABELS], stage_samples)
        plt.xlabel('睡眠阶段')
        plt.ylabel('样本数量')
        plt.title(f'测试集样本分布 - {test_set_name}')
        tb_writer.add_figure('test/SampleDistribution', fig, checkpoint_step)
        
        # 确保数据被写入磁盘
        tb_writer.flush()
        
        # 验证写入的事件文件
        event_files = glob.glob(os.path.join(test_tb_dir, "events.out.tfevents*"))
        if event_files:
            print(f"成功写入TensorBoard数据，事件文件: {[os.path.basename(f) for f in event_files]}")
            print(f"TensorBoard日志目录: {test_tb_dir}")
            
            # 设置事件文件权限
            for file_path in event_files:
                os.chmod(file_path, 0o644)
                print(f"设置文件权限: {file_path}")
                
            # 将此目录添加到train.py启动的TensorBoard监控中
            # 找到训练目录下的TensorBoard配置
            tb_config_path = os.path.join(base_tb_dir, ".tensorboard_logdir_spec")
            if not os.path.exists(tb_config_path):
                with open(tb_config_path, 'w') as f:
                    f.write(f"test_{test_set_name}:{test_tb_dir}")
                print(f"创建TensorBoard监控配置: {tb_config_path}")
        else:
            print(f"警告: 数据可能未被正确写入TensorBoard")
            
    except Exception as e:
        print(f"TensorBoard写入出错: {e}")
        print(f"错误详情: {traceback.format_exc()}")
        print("将继续保存其他格式的结果")
    
    # 保存详细结果为CSV
    results_df = pd.DataFrame(results)
    results_csv_path = os.path.join(result_dir, 'detailed_results.csv')
    results_df.to_csv(results_csv_path, index=False)
    print(f"Detailed results saved to: {results_csv_path}")
    
    # 保存性能指标为JSON
    metrics_json = {
        'model_name': model_name,
        'test_set_name': test_set_name,
        'total_samples': len(test_data),
        'valid_predictions': len([r for r in results if r['prediction'] != -1]),
        'accuracy': metrics['accuracy'],
        'f1_macro': metrics['f1_macro'],
        'avg_inference_time': metrics['avg_inference_time'],
        'class_accuracies': {str(k): v for k, v in metrics['class_accuracies'].items()},
        'timestamp': timestamp
    }
    metrics_path = os.path.join(result_dir, 'metrics.json')
    with open(metrics_path, 'w', encoding='utf-8') as f:
        json.dump(metrics_json, f, indent=4, ensure_ascii=False)
    print(f"Performance metrics saved to: {metrics_path}")
    
    # 保存混淆矩阵为图片
    plt.figure(figsize=(10, 8))
    cm = metrics['confusion_matrix']
    # 使用简化的标签用于混淆矩阵
    short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                xticklabels=short_labels,
                yticklabels=short_labels)
    plt.xlabel('Predicted Label')
    plt.ylabel('True Label')
    plt.title('Sleep Stage Classification Confusion Matrix')
    plt.tight_layout()
    cm_path = os.path.join(save_dir, 'confusion_matrix.png')
    plt.savefig(cm_path)
    plt.close()
    print(f"混淆矩阵已保存至: {cm_path}")
    
    # 保存测试样本分布统计
    class_counts = {}
    for r in results:
        label = r['true_label']
        if label not in class_counts:
            class_counts[label] = 0
        class_counts[label] += 1
    
    # 绘制样本分布饼图
    plt.figure(figsize=(10, 8))
    sizes = [class_counts.get(i, 0) for i in range(6)]
    plt.pie(sizes, labels=SLEEP_STAGE_LABELS, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.title('Test Sample Distribution')
    dist_path = os.path.join(result_dir, 'sample_distribution.png')
    plt.savefig(dist_path)
    plt.close()
    print(f"Sample distribution chart saved to: {dist_path}")
    
    # 保存性能指标摘要为文本文件
    summary_path = os.path.join(result_dir, 'summary.txt')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("=" * 50 + "\n")
        f.write("Sleep Stage Classification Test Results Summary\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Test Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        f.write("Dataset Statistics:\n")
        f.write(f"Total Samples: {len(test_data)}\n")
        f.write(f"Valid Prediction Samples: {len([r for r in results if r['prediction'] != -1])}\n\n")
        
        f.write("Sleep Stage Sample Distribution:\n")
        for i in range(6):
            count = class_counts.get(i, 0)
            percentage = (count / len(test_data)) * 100 if len(test_data) > 0 else 0
            label_name = SLEEP_STAGE_LABELS[i]
            f.write(f"Stage {i} ({label_name}): {count} samples ({percentage:.2f}%)\n")
        f.write("\n")
        
        f.write("Performance Metrics:\n")
        f.write(f"Overall Accuracy: {metrics['accuracy']:.4f}\n")
        f.write(f"Macro F1 Score: {metrics['f1_macro']:.4f}\n")
        f.write(f"Average Inference Time: {metrics['avg_inference_time']:.4f} seconds/sample\n\n")
        
        f.write("Accuracy by Sleep Stage:\n")
        for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
            acc = metrics['class_accuracies'][class_idx]
            f.write(f"  {label_name}: {acc:.4f}\n")
    print(f"Test summary saved to: {summary_path}")
    
    # 调用函数保存结果为Excel表格
    save_results_to_excel(results, metrics, test_data, result_dir, model_name=model_name, test_set_name=test_set_name)


def calculate_metrics(true_labels, pred_labels, total_time):
    """计算评估指标"""
    # 过滤掉无效预测
    valid_indices = [i for i, pred in enumerate(pred_labels) if pred != -1]
    if not valid_indices:
        return {
            'accuracy': 0.0,
            'f1_macro': 0.0,
            'avg_inference_time': 0.0,
            'confusion_matrix': np.zeros((6, 6), dtype=int),
            'class_accuracies': {i: 0.0 for i in range(6)}
        }
    
    valid_true = [true_labels[i] for i in valid_indices]
    valid_pred = [pred_labels[i] for i in valid_indices]
    
    # 计算总体指标
    accuracy = accuracy_score(valid_true, valid_pred)
    f1_macro = f1_score(valid_true, valid_pred, average='macro', zero_division=0)
    avg_inference_time = total_time / len(true_labels) if true_labels else 0
    confusion_matrix_result = confusion_matrix(valid_true, valid_pred, labels=range(6))
    
    # 计算每个类别的准确率
    class_accuracies = {}
    for class_idx in range(6):
        # 找出真实标签为该类别的样本索引
        class_indices = [i for i, label in enumerate(valid_true) if label == class_idx]
        if class_indices:  # 如果有该类别的样本
            # 计算该类别的准确率
            correct = sum(1 for i in class_indices if valid_pred[i] == class_idx)
            class_accuracies[class_idx] = correct / len(class_indices)
        else:
            class_accuracies[class_idx] = 0.0
    
    return {
        'accuracy': accuracy,
        'f1_macro': f1_macro,
        'avg_inference_time': avg_inference_time,
        'confusion_matrix': confusion_matrix_result,
        'class_accuracies': class_accuracies
    }


def save_results_to_excel(results, metrics, test_data, save_dir, model_name="unknown_model", test_set_name="unknown_dataset"):
    """保存测试结果为Excel表格文件
    
    Args:
        results: 详细的测试结果
        metrics: 性能指标
        test_data: 测试数据
        save_dir: 保存目录 (与混淆矩阵相同的目录)
        model_name: 模型名称
        test_set_name: 测试集名称
    """
    try:
        # 导入必要的模块
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting.rule import Rule
        from datetime import datetime

        # 直接使用传入的save_dir作为结果目录
        # 将Excel表格保存在与混淆矩阵相同的目录下，并按模型和数据集命名
        excel_path = os.path.join(save_dir, f'{model_name}_{test_set_name}_results.xlsx')
        
        # 创建一个Excel工作簿
        wb = Workbook()
        
        # 设置单元格样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # 计算各睡眠阶段的样本数
        class_counts = {}
        for r in results:
            label = r['true_label']
            if label not in class_counts:
                class_counts[label] = 0
            class_counts[label] += 1
        
        # 1. 总体指标摘要工作表
        ws_summary = wb.active
        ws_summary.title = "总体指标"
        
        # 添加标题和基本信息
        ws_summary.append(["睡眠阶段分类测试结果摘要"])
        ws_summary.append(["测试时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append(["模型名称", model_name])
        ws_summary.append(["测试集名称", test_set_name])
        ws_summary.append([])  # 空行
        
        # 数据集统计信息
        ws_summary.append(["数据集统计"])
        ws_summary.append(["总样本数", len(test_data)])
        ws_summary.append(["有效预测样本数", len([r for r in results if r['prediction'] != -1])])
        ws_summary.append([])  # 空行
        
        # 整体性能指标
        ws_summary.append(["整体性能指标"])
        ws_summary.append(["总体准确率", metrics['accuracy']])
        ws_summary.append(["宏平均F1分数", metrics['f1_macro']])
        ws_summary.append(["平均推理时间(秒/样本)", metrics['avg_inference_time']])
        ws_summary.append([])  # 空行
        
        # 各睡眠阶段准确率
        ws_summary.append(["各睡眠阶段准确率"])
        headers = ["睡眠阶段", "准确率"]
        ws_summary.append(headers)
        
        for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
            acc = metrics['class_accuracies'][class_idx]
            ws_summary.append([label_name, acc])
            
        # 应用样式
        for cell in ws_summary[1][0:1]:
            cell.font = Font(bold=True, size=14)
            
        for row in range(6, 18):
            if row == 6 or row == 10 or row == 14:
                for cell in ws_summary[row]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        
        # 2. 混淆矩阵工作表
        ws_cm = wb.create_sheet("混淆矩阵")
        
        # 添加标题
        ws_cm.append(["睡眠阶段分类混淆矩阵"])
        ws_cm.merge_cells('A1:G1')
        ws_cm['A1'].font = Font(bold=True, size=14)
        ws_cm['A1'].alignment = center_alignment
        
        # 添加列标题
        headers = ["", "预测"] + [f"{label.split(' ')[0]}" for label in SLEEP_STAGE_LABELS]
        ws_cm.append(headers)
        
        # 标记合并单元格
        ws_cm.merge_cells('B2:G2')
        ws_cm['B2'].alignment = center_alignment
        ws_cm['B2'].font = header_font
        
        # 添加行标题和数据
        ws_cm.append(["真实标签", ""])
        
        for i, label in enumerate(SLEEP_STAGE_LABELS):
            row = [label.split(' ')[0]] + [metrics['confusion_matrix'][i, j] for j in range(6)]
            ws_cm.append(row)
            
        # 应用样式
        for row in ws_cm.iter_rows(min_row=3, max_row=9, min_col=1, max_col=1):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_alignment
                
        for row in ws_cm.iter_rows(min_row=4, max_row=9, min_col=2, max_col=7):
            for cell in row:
                cell.alignment = center_alignment
                
        # 设置条件格式，高亮对角线(正确预测)
        for i in range(4, 10):
            diagonal_cell = ws_cm.cell(row=i, column=i-2)
            diagonal_cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
            
        # 3. 样本分布工作表
        ws_dist = wb.create_sheet("样本分布")
        
        # 添加标题
        ws_dist.append(["测试数据集样本分布"])
        ws_dist.merge_cells('A1:C1')
        ws_dist['A1'].font = Font(bold=True, size=14)
        ws_dist['A1'].alignment = center_alignment
        
        # 添加列标题
        ws_dist.append(["睡眠阶段", "样本数量", "百分比"])
        
        # 添加数据
        total_samples = len(test_data)
        for i in range(6):
            count = class_counts.get(i, 0)
            percentage = (count / total_samples) * 100 if total_samples > 0 else 0
            ws_dist.append([SLEEP_STAGE_LABELS[i], count, f"{percentage:.2f}%"])
            
        # 添加总计行
        ws_dist.append(["总计", total_samples, "100.00%"])
        
        # 应用样式
        for i, row in enumerate(ws_dist.iter_rows(min_row=2, max_row=9, min_col=1, max_col=3)):
            for cell in row:
                cell.border = border
                if i == 0 or i == 7:  # 标题行和总计行
                    cell.font = header_font
                    if i == 0:
                        cell.fill = header_fill
                        
                if i > 0:  # 数据行
                    cell.alignment = Alignment(horizontal='center')
            
        # 4. 详细结果工作表
        ws_detail = wb.create_sheet("详细结果")
        
        # 转换结果为DataFrame
        df_results = pd.DataFrame(results)
        
        # 添加标题
        ws_detail.append(["测试详细结果"])
        ws_detail.merge_cells(f'A1:E1')
        ws_detail['A1'].font = Font(bold=True, size=14)
        ws_detail['A1'].alignment = center_alignment
        
        # 添加表头
        headers = ["ID", "真实标签", "预测标签", "是否正确", "推理时间(秒)"]
        ws_detail.append(headers)
        
        # 添加数据
        for _, row in df_results.iterrows():
            ws_detail.append([
                row["id"], 
                row["true_label"], 
                row["prediction"], 
                "是" if row["correct"] else "否", 
                row["inference_time"]
            ])
            
        # 应用样式
        for cell in ws_detail[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            
        # 调整列宽
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = None
                
                # 获取列字母，跳过合并单元格
                for cell in col:
                    if hasattr(cell, 'column_letter'):
                        column = cell.column_letter
                        break
                
                # 如果没有找到列字母，跳过此列
                if not column:
                    continue
                    
                # 计算最大宽度
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                            
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width
        
        # 保存Excel文件
        wb.save(excel_path)
        print(f"结果摘要Excel表格已保存至: {excel_path}")
        
        # 保存混淆矩阵为图片
        try:
            plt.figure(figsize=(10, 8))
            cm = metrics['confusion_matrix']
            # 使用简化的标签用于混淆矩阵
            short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R']
            sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                       xticklabels=short_labels,
                       yticklabels=short_labels)
            plt.xlabel('Predicted Label')
            plt.ylabel('True Label')
            plt.title('Sleep Stage Classification Confusion Matrix')
            plt.tight_layout()
            cm_path = os.path.join(save_dir, 'confusion_matrix.png')
            plt.savefig(cm_path)
            plt.close()
            print(f"混淆矩阵已保存至: {cm_path}")
        except Exception as e:
            print(f"保存混淆矩阵图时出错: {e}")
            
    except Exception as e:
        print(f"保存Excel表格时出错: {e}")
        print(traceback.format_exc())


if __name__ == '__main__':
    # 设置API端口
    port = os.environ.get("API_PORT", 8000)
    
    # 初始化OpenAI客户端
    client = OpenAI(
        api_key="0",
        base_url=f"http://localhost:{port}/v1",
    )
    
    # 获取测试数据路径和模型名称
    test_data_path = os.environ.get("TEST_DATA_PATH", "/data/lhc/datasets_new/sleep/test/balanced/edf10_200hz_7500ms_tok12588_balanced_0.8_sqrt_inverse_test.json")
    test_data = load_test_data(test_data_path)
    
    # 从测试数据路径中提取测试集名称，并去掉数据量信息
    file_basename = os.path.basename(test_data_path).split('.')[0]
    # 假设数据量信息格式为"_n数字"，需要移除这部分
    import re
    test_set_name = re.sub(r'_n\d+', '', file_basename)
    
    # 尝试从环境变量或配置中获取模型名称，如果没有则使用默认值
    model_name = os.environ.get("MODEL_NAME", "llama_edf")
    
    print(f"模型: {model_name}")
    print(f"测试集: {test_set_name}")
    print(f"加载了 {len(test_data)} 个测试样本")
    
    # 创建保存结果的目录
    save_dir = os.environ.get("SAVE_DIR", "/data/lhc/results")
    os.makedirs(save_dir, exist_ok=True)
    
    # 设置打印和保存的间隔
    print_interval = int(os.environ.get("PRINT_INTERVAL", 10))
    save_interval = int(os.environ.get("SAVE_INTERVAL", 500))
    
    # 评估模型
    metrics = evaluate_model(client, test_data, print_interval=print_interval, save_interval=save_interval, save_dir=save_dir, 
                            model_name=model_name, test_set_name=test_set_name)