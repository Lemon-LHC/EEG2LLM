# 在一开始就设置环境变量禁用GPU，必须在导入TensorFlow之前
import os
os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
import sys
print("已设置CUDA_VISIBLE_DEVICES=-1，强制使用CPU")
print(f"Python版本: {sys.version}")

import mne
import numpy as np
import os
import json
from tqdm import tqdm
import shutil
import glob
import re
import multiprocessing
import warnings  # 添加warnings模块
import argparse  # 添加命令行参数解析模块
from concurrent.futures import ProcessPoolExecutor, as_completed
from transformers import AutoTokenizer  
from scipy import interpolate  # 添加用于插值的包
from collections import Counter  # 添加用于统计的包
from sklearn.metrics import recall_score, precision_score, f1_score, confusion_matrix  # 添加用于评估指标计算的包
import concurrent.futures
import pandas as pd
from sklearn.model_selection import train_test_split
from scipy.signal import resample
import pyeeg as pe
from sklearn.preprocessing import StandardScaler
import datetime
import openpyxl
import scipy.signal

# 提前导入TensorFlow相关库，并处理异常
tensorflow_available = False
tf_version = None
load_model_method = None

try:
    # 尝试导入tensorflow
    import tensorflow as tf
    tensorflow_available = True
    tf_version = tf.__version__
    print(f"成功导入TensorFlow，版本: {tf_version}")
    
    # 尝试导入keras基础模块
    try:
        from tensorflow import keras
        print(f"成功导入tensorflow.keras模块")
        
        # 自定义模型加载函数
        def custom_load_model(model_path, custom_objects=None, compile=True):
            """
            自定义模型加载函数，用于替代tensorflow.keras.models.load_model
            提供更稳健的加载机制和详细的错误处理
            
            Args:
                model_path: 模型文件路径
                custom_objects: 自定义对象字典
                compile: 是否编译模型
                
            Returns:
                加载的模型对象或None（如果加载失败）
            """
            print(f"尝试加载模型文件: {model_path}")
            
            # 检查文件是否存在
            if not os.path.exists(model_path):
                print(f"错误: 模型文件 {model_path} 不存在")
                return None
                
            # 检查文件扩展名
            if not model_path.endswith('.h5') and not model_path.endswith('.keras'):
                print(f"警告: 模型文件 {model_path} 不是标准的.h5或.keras格式")
            
            # 尝试多种方法加载模型
            errors = []
            model = None
            
            # 方法1: 使用tf.keras API
            if model is None:
                try:
                    print(f"尝试使用tf.keras.models.load_model加载模型")
                    model = tf.keras.models.load_model(model_path, custom_objects=custom_objects, compile=compile)
                    print(f"成功使用tf.keras.models.load_model加载模型")
                    return model
                except Exception as e:
                    error_msg = f"tf.keras.models.load_model错误: {str(e)}"
                    print(error_msg)
                    errors.append(error_msg)
            
            # 方法2: 尝试使用keras直接API
            if model is None:
                try:
                    print(f"尝试使用keras.models.load_model加载模型")
                    import keras
                    model = keras.models.load_model(model_path, custom_objects=custom_objects, compile=compile)
                    print(f"成功使用keras.models.load_model加载模型")
                    return model
                except Exception as e:
                    error_msg = f"keras.models.load_model错误: {str(e)}"
                    print(error_msg)
                    errors.append(error_msg)
            
            # 方法3: 尝试使用低级API (分开加载架构和权重)
            if model is None:
                try:
                    print(f"尝试使用低级API加载模型")
                    # 尝试加载模型架构
                    json_path = model_path.replace('.h5', '.json').replace('.keras', '.json')
                    if os.path.exists(json_path):
                        with open(json_path, 'r') as f:
                            model_json = f.read()
                        model = tf.keras.models.model_from_json(model_json, custom_objects=custom_objects)
                        # 加载权重
                        model.load_weights(model_path)
                        print(f"成功使用低级API加载模型")
                        return model
                    else:
                        errors.append("找不到模型架构JSON文件")
                except Exception as e:
                    error_msg = f"低级API加载错误: {str(e)}"
                    print(error_msg)
                    errors.append(error_msg)
                    
            # 如果所有方法都失败，返回None
            print(f"错误: 所有加载方法都失败，无法加载模型 {model_path}")
            print(f"错误详情:\n{', '.join(errors)}")
            return None
        
        # 使用自定义函数替代原始load_model
        load_model = custom_load_model
        load_model_method = "custom_load_model"
        print(f"已使用自定义load_model函数替代原始函数")
        
    except ImportError as ke:
        print(f"警告: 无法导入tensorflow.keras模块: {ke}")
        try:
            # 尝试直接使用自定义加载函数
            def custom_load_model(model_path, custom_objects=None, compile=True):
                """无法导入keras时的自定义加载函数"""
                print(f"错误: 无法加载模型 {model_path}，keras模块不可用")
                return None
            
            load_model = custom_load_model
            load_model_method = "fallback_custom"
            print(f"已使用fallback自定义加载函数")
        except Exception as e:
            print(f"无法创建自定义加载函数: {e}")
            def load_model(*args, **kwargs):
                print(f"错误: 模型加载功能不可用")
                return None
            load_model_method = "dummy"
except ImportError as e:
    print(f"警告: 无法导入TensorFlow: {e}")
    print("请安装TensorFlow: pip install tensorflow")
    # 创建替代函数以避免未定义错误
    def load_model(*args, **kwargs):
        print("错误: TensorFlow未正确安装，无法加载模型")
        print("请使用conda安装TensorFlow: conda install -n fine tensorflow")
        return None

import threading
import time
import inspect
import logging
import random
import pywt  # 用于小波变换
from scipy.signal import welch
from concurrent.futures import ThreadPoolExecutor

# 导入pyedflib用于处理EDF文件
try:
    import pyedflib
except ImportError:
    print("警告: pyedflib未安装，某些EDF文件可能无法正确读取")

# 导入matplotlib用于绘图（仅在开发调试时使用）
try:
    import matplotlib.pyplot as plt
except ImportError:
    print("警告: matplotlib未安装，可视化功能将不可用")

# 尝试导入情绪特征提取所需的模块
try:
    sys.path.append('/data/lhc/projects/Emotion')
    import pyeeg as pe
except ImportError:
    print("警告: 无法导入情绪特征提取模块，情绪特征提取功能将不可用")

# 在脚本开始时过滤MNE相关警告
warnings.filterwarnings("ignore", category=RuntimeWarning)
mne.set_log_level("ERROR")
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'  # 屏蔽TensorFlow的警告

# 全局变量
window_length_ms = 15000  
filter_windows = False  # 默认不过滤窗口，改为False
add_noise = False  # 默认不添加噪声，这里修改默认值

# 加载分词器
tokenizer_path = "/data/lhc/models/Llama-3.2-1B-Instruct"
tokenizer = None  

# 安全获取tokenizer的函数
def get_tokenizer():
    """安全地获取tokenizer，如果加载失败则返回None"""
    global tokenizer
    if tokenizer is None:
        try:
            tokenizer = AutoTokenizer.from_pretrained(tokenizer_path, trust_remote_code=True)
            print(f"已成功加载tokenizer: {tokenizer_path}")
        except Exception as e:
            print(f"加载tokenizer失败: {e}")
            return None
    return tokenizer

# 安全计算token长度
def safe_calculate_token_length(text):
    """安全地计算文本的token长度，失败时返回估计值"""
    tk = get_tokenizer()
    if tk is not None:
        try:
            return len(tk.encode(text))
        except Exception as e:
            print(f"计算token长度时出错: {e}")
    # 如果tokenizer不可用或出错，使用简单估算（每个字符约1.5个token）
    return int(len(text) * 1.5)

def calculate_tokens(sample):
    """计算样本的token数量"""
    try:
        # 构建完整提示
        system = sample.get('system', '')
        instruction = sample.get('instruction', '')
        input_text = sample.get('input', '')
        
        # 获取tokenizer
        tk = get_tokenizer()
        if tk is None:
            # 如果tokenizer不可用，使用估算值
            instruction_token_count = safe_calculate_token_length(instruction)
            input_token_count = safe_calculate_token_length(input_text)
            if system:
                total_token_count = safe_calculate_token_length(system + "\n" + instruction + "\n" + input_text)
            else:
                total_token_count = safe_calculate_token_length(instruction + "\n" + input_text)
        else:
            # 计算指令部分的token数量
            instruction_token_count = len(tk.encode(instruction))
            # 计算输入部分的token数量
            input_token_count = len(tk.encode(input_text))
            # 构建完整prompt
            if system:
                prompt_text = system + "\n" + instruction + "\n" + input_text
            else:
                prompt_text = instruction + "\n" + input_text
            total_token_count = len(tk.encode(prompt_text))
            
        return {
            "instruction_tokens": instruction_token_count,
            "input_tokens": input_token_count,
            "total_tokens": total_token_count
        }
    except Exception as e:
        print(f"计算token数量时出错: {str(e)}")
        # 出错时返回默认值
        return {
            "instruction_tokens": 1000,
            "input_tokens": 2000,
            "total_tokens": 3000
        }

# 添加线程锁，用于保护共享资源
print_lock = threading.Lock()

def safe_print(*args, **kwargs):
    """线程安全的打印函数"""
    with print_lock:
        print(*args, **kwargs)

# 情绪模型映射到描述性标签
EMOTION_MAPPINGS = {
    '0000': '无明显情绪',
    '0001': '低效价低唤醒(LVLA)',
    '0010': '低效价高唤醒(LVHA)',
    '0011': '低效价高+低唤醒混合',
    '0100': '高效价低唤醒(HVLA)',
    '0101': '高效价低唤醒+低效价低唤醒混合',
    '0110': '高效价低唤醒+低效价高唤醒混合',
    '0111': '高效价低唤醒+低效价高唤醒+低效价低唤醒混合',
    '1000': '高效价高唤醒(HVHA)',
    '1001': '高效价高唤醒+低效价低唤醒混合',
    '1010': '高效价高唤醒+低效价高唤醒混合',
    '1011': '高效价高唤醒+低效价高唤醒+低效价低唤醒混合',
    '1100': '高效价高唤醒+高效价低唤醒混合',
    '1101': '高效价高唤醒+高效价低唤醒+低效价低唤醒混合',
    '1110': '高效价高唤醒+高效价低唤醒+低效价高唤醒混合',
    '1111': '复杂混合情绪（全部激活）',
}
# 注：默认输出全部16种情绪组合，只有显式指定resolve_contradictions=True时才合并为典型四类。
def extract_emotion_features_sliding_window(signal, sfreq, window_length, step_size, emotion_models, normalize_features=True, resolve_contradictions=False):
    """使用滑动窗口提取情绪特征
    
    参数:
        signal: 信号数据
        sfreq: 采样频率
        window_length: 窗口长度(秒)
        step_size: 步长(秒)
        emotion_models: 情绪模型字典
        normalize_features: 是否标准化特征
        resolve_contradictions: 是否解决矛盾的情绪预测
        
    返回:
        features_list: 特征列表
        emotion_codes: 情绪编码列表
    """
    if not isinstance(emotion_models, dict) or len(emotion_models) < 4:
        safe_print("警告: 情绪模型无效，无法提取情绪特征")
        return [], []
    
    # 计算窗口大小和步长(样本数)
    window_samples = int(window_length * sfreq)
    step_samples = int(step_size * sfreq)
    
    # 如果信号长度小于窗口大小，则无法进行处理
    if signal.shape[1] < window_samples:
        safe_print(f"信号长度({signal.shape[1]})小于窗口大小({window_samples})，无法提取情绪特征")
        return [], []
    
    features_list = []
    emotion_codes = []
    
    # 滑动窗口提取特征
    start = 0
    while start + window_samples <= signal.shape[1]:
        # 提取当前窗口的信号
        window = signal[:, start:start + window_samples]
        
        # 提取特征
        features = extract_emotion_features(window, normalize=normalize_features)
        
        if features.size > 0:
            # 使用模型预测情绪
            emotion_code, _, _ = predict_emotions_multi_model(emotion_models, features, resolve_contradictions=resolve_contradictions)
            
            # 添加到结果列表
            features_list.append(features)
            emotion_codes.append(emotion_code)
        
        # 移动窗口
        start += step_samples
    
    return features_list, emotion_codes

def format_signal_data(signal, channel_names, sfreq=None):
    """将信号数据转换为文本格式，使用微伏为单位但不显示单位符号，使用>分隔数据点
    
    参数:
        signal: 信号数据数组
        channel_names: 通道名称列表
        sfreq: 采样频率（Hz），默认为None，则使用100Hz
    """
    formatted_data = []
    # 如果没有提供采样频率，默认使用100Hz
    if sfreq is None:
        sfreq = 100  
    
    # 动态计算时间点，根据采样频率将采样点索引转换为毫秒
    # 1000/sfreq 表示每个采样点之间的毫秒数
    time_points = np.arange(signal.shape[1]) * (1000/sfreq)  
    
    # 放大信号至微伏级别，典型脑电信号在微伏级别
    scale_to_microvolt = 1000000  
    
    # 构建数据点字符串列表
    data_points = []
    for t in range(signal.shape[1]):
        # 将值转换为微伏并确保有足够的精度
        ch1_val = float(signal[0, t]) * scale_to_microvolt
        ch2_val = float(signal[1, t]) * scale_to_microvolt if signal.shape[0] > 1 else 0.0
        
        # 只保留电压值，不包含单位和括号，使用逗号分隔通道值
        point_data = f"{ch1_val:.2f},{ch2_val:.2f}"
        data_points.append(point_data)
    
    # 使用 > 分隔数据点，并在开头添加通道名称
    return f"{', '.join(channel_names)}>{'>'.join(data_points)}>"

def create_llm_sample(signal, stage, channel_names, sfreq, emotion_models, include_emotion, emotion_window_length, emotion_step_size, emotion_codes_sequence=None, eeg_window_sec=15.0):
    """创建用于大模型训练的样本，接收预计算的情绪序列"""
    # -- 移除全局 window_length_ms 的依赖 --
    # global window_length_ms
    global EMOTION_MAPPINGS

    # 格式化信号数据
    formatted_signal = format_signal_data(signal, channel_names, sfreq)

    # 计算采样间隔（毫秒）
    interval = int(1000/sfreq)

    # -- 使用传入的 eeg_window_sec 计算窗口长度（毫秒） --
    window_length_ms = int(eeg_window_sec * 1000)
    window_length_sec_actual = eeg_window_sec # 使用实际秒数
    # safe_print(f"[Debug create_llm_sample] Using EEG window: {window_length_sec_actual} sec ({window_length_ms} ms)")

    # 获取情绪信息
    emotion_info = ""
    emotion_input_prefix = ""
    emotion_codes = emotion_codes_sequence if emotion_codes_sequence is not None else [] # 使用传入的序列

    # --- 修改：直接使用传入的 emotion_codes 序列 ---
    if include_emotion and emotion_codes: # 检查序列是否存在且非空
        try:
            # 直接使用传入的序列计算统计信息
            total_codes = len(emotion_codes)
            
            # 过滤无效编码（以防万一）
            valid_emotion_codes = [code for code in emotion_codes if isinstance(code, str) and len(code) == 4]
            if len(valid_emotion_codes) != total_codes:
                safe_print(f"警告: 传入的情绪序列包含无效编码，已过滤。原始长度: {total_codes}, 有效长度: {len(valid_emotion_codes)}")
                emotion_codes = valid_emotion_codes
                total_codes = len(emotion_codes)

            # 如果过滤后为空，则不添加情绪信息
            if total_codes == 0:
                 safe_print("警告: 传入的有效情绪序列为空，不添加情绪信息")
                 include_emotion = False # 禁用情绪信息
            else:
                # 计算各情绪的占比 (基于传入的序列)
                hvha_count = sum(1 for code in emotion_codes if code[0] == '1')
                hvla_count = sum(1 for code in emotion_codes if code[1] == '1')
                lvha_count = sum(1 for code in emotion_codes if code[2] == '1')
                lvla_count = sum(1 for code in emotion_codes if code[3] == '1')

                hvha_ratio = int(100 * hvha_count / total_codes)
                hvla_ratio = int(100 * hvla_count / total_codes)
                lvha_ratio = int(100 * lvha_count / total_codes)
                lvla_ratio = int(100 * lvla_count / total_codes)

                # 统一统计16类情绪占比
                all_emotion_codes = [
                    '0000', '1000', '0100', '0010', '0001', '1100', '0011', '1010',
                    '0101', '1001', '0110', '1110', '1101', '1011', '0111', '1111'
                ]
                code_counter = Counter(emotion_codes)
                ratio_str = 'Emotion ratio: ' + ', '.join([
                    f"{code}:{int(100 * code_counter.get(code, 0) / total_codes)}%" for code in all_emotion_codes
                ]) + '.'
                
                emotion_sequence = ">".join(emotion_codes)
                sequence_str = f"Emotion sequence: {emotion_sequence}."
                
                emotion_input_prefix = f"{ratio_str} {sequence_str} "

                # 生成 emotion_info (描述信息)
                # 使用完整的 EMOTION_MAPPINGS 定义
                emotion_code_map = {
                    '0000': 'No significant emotion',
                    '0001': 'Low valence low arousal (LVLA)',
                    '0010': 'Low valence high arousal (LVHA)',
                    '0011': 'Low valence high+low arousal mixed',
                    '0100': 'High valence low arousal (HVLA)',
                    '0101': 'High valence low arousal + low valence low arousal mixed',
                    '0110': 'High valence low arousal + low valence high arousal mixed',
                    '0111': 'High valence low arousal + low valence high arousal + low valence low arousal mixed',
                    '1000': 'High valence high arousal (HVHA)',
                    '1001': 'High valence high arousal + low valence low arousal mixed',
                    '1010': 'High valence high arousal + low valence high arousal mixed',
                    '1011': 'High valence high arousal + low valence high arousal + low valence low arousal mixed',
                    '1100': 'High valence high arousal + high valence low arousal mixed',
                    '1101': 'High valence high arousal + high valence low arousal + low valence low arousal mixed',
                    '1110': 'High valence high arousal + high valence low arousal + low valence high arousal mixed',
                    '1111': 'Complex mixed emotions (all activated)',
                } # 结束字典定义
                emotion_code_desc = '\n'.join([f"'{k}': '{v}'" for k, v in emotion_code_map.items()])
                emotion_info = (
                    f"Emotion Code Explanation:\n{emotion_code_desc}\nEach code corresponds to a 4-bit binary emotion label. The statistics show the proportion of each code within the window, and the emotion sequence represents the sliding window emotion codes.\n"
                    f"This sample contains {len(emotion_codes)} emotion code points, derived from internal {emotion_window_length}s windows with a {emotion_step_size}s step." # 修改描述
                )

                # 可选：构建 emotion_data_internal (如果需要在别处使用)
                # emotion_data_internal = { ... }

        except Exception as e:
            safe_print(f"处理传入的情绪序列时出错: {str(e)}")
            include_emotion = False # 出错则禁用情绪信息

    # --- 移除内部的情绪预测调用 ---
    # (删除或注释掉原来调用 extract_emotion_features_sliding_window 的代码块)

    # ... (构建 instruction 的逻辑不变，它会根据 include_emotion 和 emotion_info 是否存在来添加信息) ...
    # Create the instruction section (all English)
    instruction = f"""You are a neurobiological expert specializing in EEG data analysis and sleep stage classification. Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) and determine the current sleep stage of the volunteer based on the following classification criteria:
0: W (Wakefulness)
1: N1 (Light Sleep)
2: N2 (Intermediate Sleep)
3: N3 (Deep Sleep)
4: N4 (Very Deep Sleep)
5: REM (Rapid Eye Movement)

The EEG data is provided in the format: 'channel names > data points', where each data point is formatted as 'Fpz-Cz voltage in μV, Pz-Oz voltage in μV' and separated by '>' symbols. The data spans {window_length_ms}ms ({window_length_sec_actual} seconds) with a sampling interval of {interval}ms, meaning each data point is {interval}ms apart.

Emotion code sequence length: {len(emotion_codes) if include_emotion else 'N/A'}, derived from internal {emotion_window_length}s windows with {emotion_step_size}s step.

EEG frequency bands and their meanings:
- δ (Delta) wave: 0.5-4 Hz, mainly appears in deep sleep
- θ (Theta) wave: 4-8 Hz, common in light sleep and REM sleep
- α (Alpha) wave: 8-13 Hz, prominent in relaxed wakefulness
- β (Beta) wave: 13-30 Hz, prominent during focus and alertness
- γ (Gamma) wave: >30 Hz, associated with high-level cognitive processing

In your analysis, pay attention to the following characteristics of each sleep stage:
- W (Wakefulness): Mainly alpha and beta waves, high frequency, low amplitude, obvious eye movements.
- N1 (Light Sleep): Reduced alpha activity, increased theta waves, low amplitude, mixed frequency.
- N2 (Intermediate Sleep): Sleep spindles (12-14Hz bursts) and K-complexes (high amplitude biphasic waves), background theta activity.
- N3/N4 (Deep/Very Deep Sleep): Dominated by high amplitude, low frequency delta waves (>20%), significantly slowed brain activity.
- REM (Rapid Eye Movement): REM sleep has highly unique features. Characterized by rapid, irregular eye movements visible in EEG as sawtooth waves. Key features include low amplitude, mixed frequency EEG activity with prominent theta waves (4-7 Hz). While similar to N1, REM has unique sawtooth patterns. Unlike N2, REM lacks sleep spindles and K-complexes. EEG in REM resembles wakefulness but with complete muscle atonia.

Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. Do not include any additional text, punctuation, or explanations."""

    # Add emotion information to instruction (if any, all English)
    if include_emotion and emotion_info:
        instruction = f"{instruction}\n{emotion_info}"
    
    # 根据不同的睡眠阶段设置输出标签
    sleep_stages = {
        0: "0", # W (清醒)
        1: "1", # N1 (浅睡眠)
        2: "2", # N2 (轻睡眠)
        3: "3", # N3 (深睡眠)
        4: "4", # N4 (深睡眠)
        5: "5"  # REM (快速眼动)
    }
    output_text = sleep_stages.get(stage, "Unknown")
    
    # 在input前添加情绪信息前缀
    if include_emotion and emotion_input_prefix:
        formatted_signal = f"{emotion_input_prefix}{formatted_signal}"
    
    # 创建新的格式样本
    sample = {
        "instruction": instruction,
        "input": formatted_signal,
        "output": output_text,
        "system": "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification."
    }
    
    # (移除原来处理内部 emotion_data 的部分)
    
    return sample

def clean_output_directories(base_dir):
    """清理输出目录中的所有json文件"""
    processed_dir = os.path.join(base_dir)
    if os.path.exists(processed_dir):
        print("清理之前生成的json文件...")
        # 只删除json文件，而不是整个目录
        for file_name in os.listdir(processed_dir):
            if file_name.endswith('.json'):
                file_path = os.path.join(processed_dir, file_name)
                try:
                    os.remove(file_path)
                    print(f"已删除: {file_path}")
                except Exception as e:
                    print(f"删除文件 {file_path} 时出错: {str(e)}")
    else:
        # 如果目录不存在，创建它
        os.makedirs(processed_dir, exist_ok=True)

def prepare_output_directories(base_dir):
    """准备输出目录，返回三个主要数据集目录列表
    
    Args:
        base_dir: 基础输出目录
        
    Returns:
        包含train、test、all三个目录路径的列表
    """
    # 创建主输出目录
    os.makedirs(base_dir, exist_ok=True)
    
    # 清理目录
    clean_output_directories(base_dir)
    
    # 创建train、test和all子目录
    train_dir = os.path.join(base_dir, 'train')
    test_dir = os.path.join(base_dir, 'test')
    all_dir = os.path.join(base_dir, 'all')
    
    # 只创建主目录和三大子目录
    os.makedirs(train_dir, exist_ok=True)
    os.makedirs(test_dir, exist_ok=True)
    os.makedirs(all_dir, exist_ok=True)
    
    # 返回所有目录
    return [train_dir, test_dir, all_dir]

def interpolate_signal(signal, original_sfreq, target_sfreq):
    """对信号进行插值以增加采样率
    
    参数:
        signal: 原始信号数据数组，形状为 [n_channels, n_points]
        original_sfreq: 原始采样频率 (Hz)
        target_sfreq: 目标采样频率 (Hz)
        
    返回:
        插值后的信号数据数组
    """
    if original_sfreq == target_sfreq:
        return signal
    
    # 创建原始时间点（以秒为单位）
    n_points = signal.shape[1]
    original_times = np.arange(n_points) / original_sfreq
    
    # 创建新的时间点（以秒为单位）
    ratio = target_sfreq / original_sfreq
    n_new_points = int(n_points * ratio)
    new_times = np.linspace(0, original_times[-1], n_new_points)
    
    # 对每个通道进行插值
    interpolated_signal = np.zeros((signal.shape[0], n_new_points))
    
    for ch_idx in range(signal.shape[0]):
        # 使用三次样条插值
        f = interpolate.interp1d(original_times, signal[ch_idx, :], kind='cubic', bounds_error=False, fill_value="extrapolate")
        interpolated_signal[ch_idx, :] = f(new_times)
    
    return interpolated_signal

def read_edf_file(edf_path, target_sfreq=100):
    """读取EDF文件并进行预处理
    
    Args:
        edf_path: EDF文件路径
        target_sfreq: 目标采样频率 (Hz)
        
    Returns:
        raw: MNE Raw对象
        available_channels: 可用通道列表
    """
    # 使用verbose='error'只显示错误，忽略警告
    raw = mne.io.read_raw_edf(edf_path, preload=True, verbose='error')
    original_sfreq = raw.info['sfreq']
    
    # 检查并保留指定通道
    channels_to_keep = ['EEG Fpz-Cz', 'EEG Pz-Oz']
    available_channels = []
    for ch in channels_to_keep:
        if ch in raw.ch_names:
            available_channels.append(ch)

    if not available_channels:
        raise ValueError("No required channels found in the data")

    # 使用新的API方式避免警告
    raw.pick(available_channels, verbose='error')
    
    # 检查数据是否有效
    data = raw.get_data()
    
    # 检查数据是否为零或接近零
    signal_max = np.max(np.abs(data))
    
    # 数据信号太弱时进行放大（但不在这里转换为微伏，仅放大信号）
    if signal_max < 1e-5:
        # 信号弱则放大到适当范围
        scale_factor = 1.0 / signal_max if signal_max > 0 else 1e6
        data = data * scale_factor
        # 创建新的RawArray对象
        info = raw.info
        raw = mne.io.RawArray(data, info, verbose='error')
    
    # 如果目标采样率与原始采样率不同，则进行重采样
    if target_sfreq != original_sfreq:
        # 使用MNE的resample方法而不是直接修改info['sfreq']
        raw.resample(target_sfreq, npad='auto', verbose='error')
    
    return raw, available_channels

def extract_stage_windows(raw_data, annotations, channel_names, emotion_models, include_emotion, emotion_window_length, emotion_step_size, add_noise, max_windows, normalize_features, eeg_window_sec=15.0, eeg_step_sec=15.0):
    """
    提取不同睡眠阶段的窗口。
    修改后的逻辑：在每个睡眠阶段注释的持续时间内，以 eeg_step_sec 为步长滑动提取窗口，
    然后应用 max_windows 限制每个阶段最终提取的窗口总数。
    """
    safe_print(f"[Debug extract_stage_windows] Entering function. Received max_windows={max_windows}, eeg_window_sec={eeg_window_sec}, eeg_step_sec={eeg_step_sec}. Logic: Sliding window within annotations.")
    max_processing_time = 300
    start_time = time.time()

    data = raw_data.get_data()
    sfreq = raw_data.info['sfreq']
    eeg_window_samples = int(eeg_window_sec * sfreq)
    # --- 新增：计算 EEG 步长对应的采样点数 ---
    eeg_step_samples = max(1, int(eeg_step_sec * sfreq)) # 步长至少为 1 个采样点
    safe_print(f"[Debug extract_stage_windows] EEG Window: {eeg_window_sec} sec => {eeg_window_samples} samples.")
    safe_print(f"[Debug extract_stage_windows] EEG Step: {eeg_step_sec} sec => {eeg_step_samples} samples (sfreq={sfreq}Hz)")
    # --- 结束新增 ---

    ch_indices = []
    for ch_name in channel_names:
        if ch_name in raw_data.ch_names:
            ch_idx = raw_data.ch_names.index(ch_name)
            ch_indices.append(ch_idx)
    if not ch_indices:
        safe_print("错误: 找不到所需的通道")
        return {}
    ch_data = data[ch_indices, :]

    stages = []
    for annot in annotations:
        onset = int(annot['onset'] * sfreq)
        duration = int(annot['duration'] * sfreq)
        stage_desc = annot['description']
        stage_num = -1 # Default to invalid
        if stage_desc.endswith('W'): stage_num = 0
        elif stage_desc.endswith('1'): stage_num = 1
        elif stage_desc.endswith('2'): stage_num = 2
        elif stage_desc.endswith('3'): stage_num = 3
        elif stage_desc.endswith('4'): stage_num = 4
        elif stage_desc.endswith('R'): stage_num = 5
        else:
            safe_print(f"[Debug extract_stage_windows] Skipping unknown annotation stage: '{stage_desc}' at onset {annot['onset']}")
            continue
        stages.append((onset, duration, stage_num))

    safe_print(f"[Debug extract_stage_windows] Total valid annotation blocks parsed: {len(stages)}")
    stage_counts_raw = Counter([s[2] for s in stages])
    safe_print(f"[Debug extract_stage_windows] Raw annotation counts per stage: {stage_counts_raw}")

    windows_by_stage = {}
    total_windows_extracted_file = 0
    total_potential_windows_file = 0
    total_boundary_skips_file = 0

    for stage_num in range(6):
        if time.time() - start_time > max_processing_time:
            safe_print(f"错误: 文件处理超时 (>{max_processing_time}s)")
            break # 超时则停止处理此文件

        stage_annotations = [s for s in stages if s[2] == stage_num]
        if not stage_annotations: continue

        all_stage_windows_info = []
        stage_potential_windows = 0
        stage_boundary_skips = 0

        with tqdm(total=len(stage_annotations), desc=f"Scanning Stage {stage_num} Annotations") as pbar_annot:
            for onset, duration, _ in stage_annotations:
                if time.time() - start_time > max_processing_time: # 内部超时检查
                   safe_print(f"错误: 文件处理超时 (>{max_processing_time}s) during stage {stage_num} annotation scan.")
                   break

                for win_start in range(onset, onset + duration - eeg_window_samples + 1, eeg_step_samples):
                    # 边界检查
                    if win_start + eeg_window_samples > ch_data.shape[1]:
                        stage_boundary_skips += 1
                        continue

                    stage_potential_windows += 1

                    window_data = ch_data[:, win_start:win_start + eeg_window_samples].copy()
                    if add_noise:
                         noise_level = 0.05 * np.std(window_data)
                         noise = np.random.normal(0, noise_level, window_data.shape)
                         window_data += noise
                    window_info_dict = {
                         'data': window_data.astype(np.float64),
                         'stage': stage_num,
                         'start': win_start / sfreq,
                         'end': (win_start + eeg_window_samples) / sfreq
                    }
                    all_stage_windows_info.append(window_info_dict)
                pbar_annot.update(1)
            if time.time() - start_time > max_processing_time: break # Break outer loop if timed out

        num_before_limit = len(all_stage_windows_info)
        limit_applied = False
        final_stage_windows = all_stage_windows_info # 默认使用所有窗口

        if max_windows is not None and max_windows > 0 and num_before_limit > max_windows:
            safe_print(f"[Debug extract_stage_windows] Stage {stage_num}: Applying max_windows={max_windows}. Limiting from {num_before_limit} potential windows.")
            np.random.shuffle(all_stage_windows_info) # 打乱以随机选取
            final_stage_windows = all_stage_windows_info[:max_windows]
            limit_applied = True
        elif max_windows is not None and max_windows <= 0:
            safe_print(f"[Debug extract_stage_windows] Stage {stage_num}: max_windows is {max_windows}. No limit applied to {num_before_limit} potential windows.")
        elif max_windows is None:
             safe_print(f"[Debug extract_stage_windows] Stage {stage_num}: max_windows is None. No limit applied to {num_before_limit} potential windows.")


        stage_windows_extracted = len(final_stage_windows)

        safe_print(f"[Debug extract_stage_windows] Stage {stage_num}: Found {stage_potential_windows} potential windows, skipped {stage_boundary_skips} at boundary.")
        safe_print(f"[Debug extract_stage_windows] Stage {stage_num}: Limit applied: {limit_applied}. Final extracted windows for stage: {stage_windows_extracted} (from {num_before_limit} potential).")

        if final_stage_windows:
            windows_by_stage[stage_num] = final_stage_windows

        total_potential_windows_file += stage_potential_windows
        total_windows_extracted_file += stage_windows_extracted
        total_boundary_skips_file += stage_boundary_skips

    window_counts = {stage: len(windows) for stage, windows in windows_by_stage.items()}
    safe_print(f"[Debug extract_stage_windows] Final extracted window counts for file: {window_counts}")
    safe_print(f"[Debug extract_stage_windows] Summary for file: Found {total_potential_windows_file} potential windows, skipped {total_boundary_skips_file} at boundary, extracted final {total_windows_extracted_file} windows (after max_windows limit).")
    safe_print("[Debug extract_stage_windows] Exiting function.")
    return windows_by_stage

def find_annotation_file(edf_path):
    """根据信号文件路径找到对应的注释文件路径"""
    base_name = os.path.basename(edf_path)
    dir_name = os.path.dirname(edf_path)
    
    # 提取基本ID，例如从'SC4001E0-PSG.edf'提取'SC4001'或从'ST7022J0-PSG.edf'提取'ST7022'
    match = re.match(r'((?:SC|ST)\d+)[A-Z]\d+-PSG\.edf', base_name)
    if not match:
        return None
    
    base_id = match.group(1)
    
    # 查找对应的注释文件
    annotation_pattern = os.path.join(dir_name, f"{base_id}*-Hypnogram.edf")
    annotation_files = glob.glob(annotation_pattern)
    
    if not annotation_files:
        return None
    
    return annotation_files[0]

def process_file(edf_path):
    """处理单个EDF文件
    
    Args:
        edf_path: EDF文件路径
        
    Returns:
        处理结果
    """
    try:
        # 查找对应的注释文件
        annotation_path = find_annotation_file(edf_path)
        if not annotation_path:
            safe_print(f"未找到文件 {os.path.basename(edf_path)} 对应的注释文件，跳过")
            # 提取文件ID以便返回统一格式的错误信息
            file_id = os.path.basename(edf_path).replace('.edf', '').replace('-PSG', '').replace('-Hypnogram', '')
            return {
                'success': False,
                'file_id': file_id,
                'stage_counts': {},
                'file_path': ''
            }
        
        safe_print(f"处理文件: {os.path.basename(edf_path)} -> {os.path.basename(annotation_path)}")
        
        # 从当前函数获取全局参数
        frame = inspect.currentframe()
        try:
            outer_frame = frame.f_back
            for var_name in ['output_base_dir', 'target_sfreq', 'emotion_models', 'include_emotion', 
                            'emotion_window_length', 'emotion_step_size', 'device', 'add_noise', 
                            'max_windows', 'timeout', 'normalize_features']:
                if var_name in outer_frame.f_locals:
                    globals()[var_name] = outer_frame.f_locals[var_name]
        finally:
            del frame
        
        # 获取输出目录，不使用硬编码的默认值
        output_base_dir = globals().get('output_base_dir')
        target_sfreq = globals().get('target_sfreq')
        emotion_models = globals().get('emotion_models')
        include_emotion = globals().get('include_emotion')
        emotion_window_length = globals().get('emotion_window_length')
        emotion_step_size = globals().get('emotion_step_size')
        device = globals().get('device')
        add_noise = globals().get('add_noise')
        max_windows = globals().get('max_windows')
        timeout = globals().get('timeout')
        normalize_features = globals().get('normalize_features')
        
        # 处理文件
        result = process_and_save_direct(
            edf_path,
            annotation_path,
            output_base_dir,
            target_sfreq,
            emotion_models,
            include_emotion,
            emotion_window_length,
            emotion_step_size,
            device,
            add_noise,
            max_windows,
            timeout,
            normalize_features,
            globals().get('resolve_emotion_conflict', False),
            eeg_window_sec=15.0,
            eeg_step_sec=15.0
        )
        
        return result
    except Exception as e:
        safe_print(f"处理文件 {os.path.basename(edf_path)} 时出错: {e}")
        import traceback
        traceback.print_exc()
        # 提取文件ID以便返回统一格式的错误信息
        file_id = os.path.basename(edf_path).replace('.edf', '').replace('-PSG', '').replace('-Hypnogram', '')
        return {
            'success': False,
            'file_id': file_id,
            'stage_counts': {},
            'file_path': ''
        }

def process_and_save_direct(edf_path, annotation_path, target_sfreq, emotion_models, include_emotion, emotion_window_length, emotion_step_size, device, add_noise, max_windows, timeout, normalize_features, resolve_emotion_conflict=False, eeg_window_sec=15.0, eeg_step_sec=15.0):
    """处理单个EDF文件，提取窗口和特征，生成LLM样本。

    Args:
        # ... (参数不变) ...
        eeg_step_sec (float): EEG 窗口滑动的步长 (秒).

    Returns:
        # ... (返回值不变) ...
    """
    safe_print(f"[PID {os.getpid()}] process_and_save_direct received: edf='{os.path.basename(edf_path)}', max_windows={max_windows}, eeg_window_sec={eeg_window_sec}, eeg_step_sec={eeg_step_sec}, timeout={timeout}")

    safe_print(f"[PID {os.getpid()}] 开始处理文件: {os.path.basename(edf_path)}")
    file_id_match = re.match(r'([A-Z0-9]+)', os.path.basename(edf_path))
    file_id = file_id_match.group(1) if file_id_match else f"unknown_{hash(edf_path)}"

    all_llm_samples = [] # 用于存储此文件生成的所有LLM样本
    stage_counts_file = {} # 用于存储此文件的阶段统计

    try:
        # 读取EDF文件
        safe_print(f"[PID {os.getpid()}] 读取EDF: {os.path.basename(edf_path)}")
        raw_data, channel_names = read_edf_file(edf_path, target_sfreq)
        if raw_data is None:
            safe_print(f"[PID {os.getpid()}] 无法读取EDF文件: {edf_path}")
            return { 'success': False, 'file_id': file_id, 'stage_counts': {}, 'samples': [], 'error': '无法读取EDF' }
        safe_print(f"[PID {os.getpid()}] 完成读取EDF: {os.path.basename(edf_path)}")

        # 读取标注文件
        safe_print(f"[PID {os.getpid()}] 读取注释: {os.path.basename(annotation_path)}")
        try:
            annotations = mne.read_annotations(annotation_path)
        except Exception as e:
            safe_print(f"[PID {os.getpid()}] 无法读取注释文件 {annotation_path}: {e}")
            return { 'success': False, 'file_id': file_id, 'stage_counts': {}, 'samples': [], 'error': f'无法读取注释: {e}' }
        safe_print(f"[PID {os.getpid()}] 完成读取注释: {os.path.basename(annotation_path)}")

        # 设置注释
        raw_data.set_annotations(annotations)

        # 提取不同睡眠阶段的窗口
        safe_print(f"[PID {os.getpid()}] 提取阶段窗口: {os.path.basename(edf_path)}")
        try:
            windows_by_stage = extract_stage_windows(
                raw_data,
                annotations,
                channel_names,
                emotion_models=emotion_models,
                include_emotion=include_emotion,
                emotion_window_length=emotion_window_length,
                emotion_step_size=emotion_step_size,
                add_noise=add_noise,
                max_windows=max_windows,
                normalize_features=normalize_features,
                eeg_window_sec=eeg_window_sec,
                eeg_step_sec=eeg_step_sec
            )
        except Exception as e:
            safe_print(f"[PID {os.getpid()}] 提取窗口时出错 {edf_path}: {e}")
            import traceback
            traceback.print_exc()
            return { 'success': False, 'file_id': file_id, 'stage_counts': {}, 'samples': [], 'error': f'提取窗口错误: {e}' }
        safe_print(f"[PID {os.getpid()}] 完成提取阶段窗口: {os.path.basename(edf_path)}")


        total_extracted_windows = sum(len(wl) for wl in windows_by_stage.values())
        stage_counts_file = {stage: len(windows) for stage, windows in windows_by_stage.items()}
        safe_print(f"[PID {os.getpid()}] 文件 {os.path.basename(edf_path)} 各阶段窗口数量: {stage_counts_file}")

        if total_extracted_windows == 0:
             safe_print(f"[PID {os.getpid()}] 文件 {os.path.basename(edf_path)} 没有提取到有效窗口，跳过")
             return { 'success': True, 'file_id': file_id, 'stage_counts': stage_counts_file, 'samples': [] }

        processed_window_count = 0
        # --- 在循环开始前添加总计日志 ---
        safe_print(f"[PID {os.getpid()}] {file_id}: 开始处理提取到的 {total_extracted_windows} 个窗口...")
        # --- 结束添加 ---

        for stage, windows_info_list in windows_by_stage.items():
            if not windows_info_list:
                continue

            # --- 添加阶段处理开始的日志 ---
            safe_print(f"[PID {os.getpid()}] {file_id}: 开始处理 Stage {stage} 的 {len(windows_info_list)} 个窗口...")
            # --- 结束添加 ---

            for i, window_info in enumerate(windows_info_list):
                current_window_label = f"FileID {file_id}, Stage {stage}, Window {i+1}/{len(windows_info_list)}" # 索引从1开始，更自然

                # --- 添加单个窗口处理开始的日志 ---
                safe_print(f"[PID {os.getpid()}] {current_window_label}: 开始处理...")
                # --- 结束添加 ---

                data = window_info['data']
                emotion_code_sequence = []
                dominant_emotion_code = "0000"

                # --- 情绪处理 (添加日志) ---
                if include_emotion and emotion_models:
                    try:
                        # --- 添加情绪特征提取开始日志 ---
                        safe_print(f"[PID {os.getpid()}] {current_window_label}: 开始提取情绪特征...")
                        # --- 结束添加 ---
                        features = extract_emotion_features(
                            signal=data,
                            sfreq=target_sfreq,
                            window_length_sec=emotion_window_length,
                            step_size_sec=emotion_step_size,
                            normalize=normalize_features,
                            print_details=False # 除非需要否则关闭详细特征日志
                        )
                        # --- 添加情绪特征提取完成日志 ---
                        features_shape_str = features.shape if features.size > 0 else 'Empty'
                        safe_print(f"[PID {os.getpid()}] {current_window_label}: 情绪特征提取完成. Shape: {features_shape_str}")
                        # --- 结束添加 ---


                        if features.size > 0:
                            # --- 添加情绪预测开始日志 ---
                            safe_print(f"[PID {os.getpid()}] {current_window_label}: 开始进行情绪预测...")
                            # --- 结束添加 ---
                            pred_sequence, dominant_label, detailed_info = predict_emotions_multi_model(
                                emotion_models, features,
                                resolve_contradictions=resolve_emotion_conflict,
                                device=device
                            )
                            # --- 添加情绪预测完成日志 ---
                            sequence_len = len(pred_sequence)
                            safe_print(f"[PID {os.getpid()}] {current_window_label}: 情绪预测完成. Sequence len: {sequence_len}, Dominant: {dominant_label}({detailed_info.get('dominant_code','?')})")
                            # --- 结束添加 ---

                            emotion_code_sequence = pred_sequence
                            if emotion_code_sequence:
                                 dominant_emotion_code = Counter(emotion_code_sequence).most_common(1)[0][0]

                    except Exception as e:
                        safe_print(f"[PID {os.getpid()}] 处理窗口 {current_window_label} 的情绪时出错: {e}")

                # --- 生成LLM样本 (添加日志) ---
                # --- 添加LLM样本创建开始日志 ---
                safe_print(f"[PID {os.getpid()}] {current_window_label}: 开始创建LLM样本...")
                # --- 结束添加 ---
                try:
                    sample = create_llm_sample(
                        data,
                        stage,
                        channel_names,
                        target_sfreq,
                        emotion_models=None,
                        include_emotion=include_emotion,
                        emotion_window_length=emotion_window_length,
                        emotion_step_size=emotion_step_size,
                        emotion_codes_sequence=emotion_code_sequence,
                        eeg_window_sec=eeg_window_sec
                    )
                    # --- 添加LLM样本创建完成日志 ---
                    safe_print(f"[PID {os.getpid()}] {current_window_label}: LLM样本创建完成.")
                    # --- 结束添加 ---

                    all_llm_samples.append(sample)
                except Exception as e:
                     safe_print(f"[PID {os.getpid()}] 创建LLM样本 {current_window_label} 时出错: {e}")
                     continue


                processed_window_count += 1
                # --- 添加单个窗口处理完成和周期性进度日志 ---
                safe_print(f"[PID {os.getpid()}] {current_window_label}: 处理完成.")
                if processed_window_count % 100 == 0 or processed_window_count == total_extracted_windows: # 每处理100个窗口或最后一个窗口时打印
                    safe_print(f"[PID {os.getpid()}] {file_id}: 已处理 {processed_window_count}/{total_extracted_windows} 个窗口...")
                # --- 结束添加 ---

        # --- 移除不再需要的 pbar 相关代码 ---

        safe_print(f"[PID {os.getpid()}] 文件 {os.path.basename(edf_path)} 处理完成，生成 {len(all_llm_samples)} 个样本。")
        return {
            'success': True,
            'file_id': file_id,
            'stage_counts': stage_counts_file,
            'samples': all_llm_samples
        }

    except Exception as e:
        safe_print(f"[PID {os.getpid()}] 处理文件 {edf_path} 时发生未预料的错误: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'file_id': file_id,
            'stage_counts': {},
            'samples': [],
            'error': f'未知错误: {e}'
        }

def extract_stages_from_annotations(annotations):
    """从注释中提取睡眠阶段序列
    
    Args:
        annotations: MNE注释对象
    
    Returns:
        stages: 包含onset、duration和stage的字典
    """
    # 提取描述、起始时间和持续时间
    descriptions = [annot['description'] for annot in annotations]
    onsets = annotations.onset
    durations = annotations.duration
    
    safe_print(f"注释描述列表: {descriptions[:10]}... (共{len(descriptions)}个)")
    
    # 定义阶段映射字典 - 扩展支持更多标签格式
    stage_map = {
        'Sleep stage W': 0,  # W (清醒)
        'Sleep stage 1': 1,  # N1 (浅睡眠)
        'Sleep stage 2': 2,  # N2 (轻睡眠)
        'Sleep stage 3': 3,  # N3 (深睡眠)
        'Sleep stage 4': 4,  # N4 (深睡眠)
        'Sleep stage R': 5   # REM (快速眼动)
    }
    
    # 创建结果字典
    result = {
        'onset': [],
        'duration': [],
        'stage': []
    }
    
    # 处理每个标注
    for desc, onset, duration in zip(descriptions, onsets, durations):
        if desc in stage_map:
            result['onset'].append(onset)
            result['duration'].append(duration)
            result['stage'].append(stage_map[desc])
    
    # 如果没有有效的睡眠阶段，生成均匀分布的假阶段
    if not result['stage']:
        safe_print("警告: 未找到任何有效的睡眠阶段注释，生成均匀分布的假阶段")
        # 生成30秒一个阶段，交替使用W和N2阶段
        duration_sec = int(annotations[-1]['onset'] - annotations[0]['onset']) if len(annotations) > 1 else 3600
        num_stages = max(1, int(duration_sec / 30))
        
        # 生成假数据
        start_time = annotations[0]['onset'] if len(annotations) > 0 else 0
        for i in range(num_stages):
            result['onset'].append(start_time + i * 30)
            result['duration'].append(30.0)  # 30秒
            result['stage'].append(0 if i % 2 == 0 else 2)  # 交替W和N2
    
    # 统计各类阶段数量
    stage_counts = Counter(result['stage'])
    safe_print(f"提取了 {len(result['stage'])} 个睡眠阶段，各阶段分布: W:{stage_counts.get(0, 0)}, N1:{stage_counts.get(1, 0)}, N2:{stage_counts.get(2, 0)}, N3:{stage_counts.get(3, 0)}, N4:{stage_counts.get(4, 0)}, REM:{stage_counts.get(5, 0)}，其他:0")
    
    return result

def process_directory(
    input_dir,
    output_dirs,
    max_files,
    n_jobs,
    target_sfreq,
    balance_strategy,
    balance_alpha,
    weight_method,
    file_pattern,
    file_type,
    include_emotion,
    emotion_model_dir,
    emotion_window_length,
    emotion_step_size,
    device,
    add_noise,
    max_windows,
    timeout,
    normalize_features,
    eeg_window_sec=15.0,
    eeg_step_sec=15.0, # <<<--- 新增参数
    resolve_emotion_conflict=False
    ):
    """处理目录中的EDF文件，聚合数据，分割，平衡，并按指定格式保存。

    Args:
        # ... 其他参数 ...
        eeg_step_sec (float): EEG 窗口滑动的步长 (秒).
        # ... 其他参数 ...
    """
    # --- 增加日志：记录 eeg_step_sec ---
    safe_print("\n--- Entering process_directory ---")
    safe_print(f"Parameters received:")
    # ... (打印其他参数) ...
    safe_print(f"  eeg_window_sec: {eeg_window_sec}")
    safe_print(f"  eeg_step_sec: {eeg_step_sec}") # <<<--- 打印新参数
    safe_print(f"  include_emotion: {include_emotion}")
    # ... (打印其他参数) ...
    safe_print("-" * 30)
    # --- 结束增加日志 ---

    # ... (目录创建和文件查找逻辑不变) ...
    train_output_dir = output_dirs['train']
    test_output_dir = output_dirs['test']
    all_output_dir = output_dirs['all'] # 用于保存统计信息

    # 确保输出目录存在
    os.makedirs(train_output_dir, exist_ok=True)
    os.makedirs(test_output_dir, exist_ok=True)
    os.makedirs(all_output_dir, exist_ok=True)

    # --- 文件查找逻辑 (再次修订，确保过滤严格执行) ---
    # 1. 构建基础搜索模式
    search_pattern = os.path.join(input_dir, file_pattern if file_pattern else "*-PSG.edf")
    safe_print(f"基础文件搜索模式: {search_pattern}")

    # 2. 获取所有匹配基础模式的文件
    all_potential_files = glob.glob(search_pattern)
    safe_print(f"找到 {len(all_potential_files)} 个可能匹配的文件。")

    # 3. 根据 file_type 参数进行严格过滤
    file_type_upper = file_type.upper() # 转换为大写以便比较
    filtered_edf_files = []
    if file_type_upper == "SC":
        filtered_edf_files = [f for f in all_potential_files if os.path.basename(f).upper().startswith("SC")]
        safe_print(f"筛选 SC 类型文件，找到 {len(filtered_edf_files)} 个。")
    elif file_type_upper == "ST":
        filtered_edf_files = [f for f in all_potential_files if os.path.basename(f).upper().startswith("ST")]
        safe_print(f"筛选 ST 类型文件，找到 {len(filtered_edf_files)} 个。")
    elif file_type_upper == "ALL":
        filtered_edf_files = all_potential_files # 保留所有文件
        safe_print(f"处理所有文件类型 ('ALL')，共 {len(filtered_edf_files)} 个。")
    else:
        safe_print(f"警告: 未知的 file_type '{file_type}'. 将处理所有找到的文件 ({len(all_potential_files)}).")
        filtered_edf_files = all_potential_files # 默认处理所有

    # 检查过滤后是否还有文件
    if not filtered_edf_files:
        safe_print(f"错误: 在目录 '{input_dir}' 中未找到类型为 '{file_type}' 且匹配模式 '{search_pattern}' 的 PSG EDF 文件。")
        return

    # 4. 应用 max_files 限制 (对过滤后的列表)
    if max_files > 0 and len(filtered_edf_files) > max_files:
        safe_print(f"文件数量超过 max_files ({max_files})，将从 {len(filtered_edf_files)} 个文件中选取前 {max_files} 个。")
        # 可以考虑在此处添加随机打乱逻辑，如果需要随机选取子集
        # random.shuffle(filtered_edf_files)
        edf_files_to_process = filtered_edf_files[:max_files]
    else:
        edf_files_to_process = filtered_edf_files

    actual_max_files = len(edf_files_to_process) # 最终实际处理的文件数
    safe_print(f"最终将处理 {actual_max_files} 个文件。")

    # --- 加载情绪模型 (如果需要) ---
    emotion_models = None
    if include_emotion:
        safe_print("加载情绪模型...")
        emotion_models = load_emotion_models(emotion_model_dir)
        if not emotion_models:
            safe_print("错误：无法加载情绪模型，请检查路径和模型文件。")
            return # 或者根据需要决定是否继续（不带情绪信息）
        safe_print(f"成功加载 {len(emotion_models)} 个情绪模型。")


    # --- 并行处理文件 ---
    all_aggregated_samples = []
    total_stage_counts = Counter()
    processed_file_count = 0
    failed_file_count = 0
    processed_file_types = {"SC": 0, "ST": 0} # 统计实际处理的文件类型

    # 准备传递给子进程的参数
    process_args = []
    safe_print("准备子进程参数...")
    for edf_path in edf_files_to_process:
        annotation_path = find_annotation_file(edf_path)
        if not annotation_path:
            safe_print(f"警告: 未找到 {os.path.basename(edf_path)} 的注释文件，跳过。")
            failed_file_count += 1
            continue

        temp_include_emotion = include_emotion

        process_args.append((
            edf_path, annotation_path, target_sfreq, emotion_models,
            temp_include_emotion,
            emotion_window_length, emotion_step_size, device,
            add_noise, max_windows,
            timeout, normalize_features,
            resolve_emotion_conflict,
            eeg_window_sec,
            eeg_step_sec # <<<--- 传递新参数给子进程
        ))
    # --- 结束修改 ---
    num_files_to_process = len(process_args) # 获取将要处理的文件总数
    safe_print(f"共准备了 {num_files_to_process} 个文件的参数。") # 添加日志

    # 使用多进程处理
    safe_print(f"开始使用 {n_jobs} 个进程并行处理 {num_files_to_process} 个文件...")
    results = []
    try:
        safe_print("进入 ProcessPoolExecutor 上下文管理器...") # 添加日志
        with ProcessPoolExecutor(max_workers=n_jobs) as executor:
            safe_print("ProcessPoolExecutor 已启动。开始提交任务...") # 添加日志
            # 将 future 映射到文件路径，以便在结果处理中识别
            futures = {executor.submit(process_and_save_direct, *p_args): p_args[0] for p_args in process_args}
            safe_print(f"已提交 {len(futures)} 个任务。开始使用 as_completed 等待结果...") # 添加日志

            # 使用 as_completed 迭代完成的任务，并添加 tqdm 进度条
            for future in tqdm(as_completed(futures), total=num_files_to_process, desc="处理文件"):
                file_path_for_future = futures[future] # 获取与此 future 关联的文件路径
                # safe_print(f"一个任务已完成 (文件: {os.path.basename(file_path_for_future)})。尝试获取结果...") # 减少冗余日志
                try:
                    # 获取结果，应用超时
                    result = future.result(timeout=timeout if timeout > 0 else None)
                    # safe_print(f"成功获取文件 {os.path.basename(file_path_for_future)} 的结果。 Success: {result.get('success', 'N/A')}") # 减少冗余日志
                    results.append(result)
                except TimeoutError:
                    safe_print(f"错误: 处理文件 {os.path.basename(file_path_for_future)} 超时 (>{timeout}s)") # 添加日志
                    failed_file_count += 1
                    # 可以考虑取消任务，但通常意义不大
                except Exception as exc:
                    safe_print(f'处理文件 {os.path.basename(file_path_for_future)} 时子进程生成了一个异常: {exc}') # 添加日志
                    # import traceback # 如果需要详细堆栈
                    # traceback.print_exc()
                    failed_file_count += 1 # 算作失败
        safe_print("已退出 ProcessPoolExecutor 上下文管理器。") # 添加日志

    except Exception as e:
        safe_print(f"多进程处理过程中发生严重错误: {e}")
        import traceback
        traceback.print_exc() # 打印详细错误
        return

    # --- 收集处理结果 ---
    for result in results:
        if result and result['success']:
            all_aggregated_samples.extend(result['samples'])
            total_stage_counts.update(result['stage_counts'])
            processed_file_count += 1
            # 统计文件类型
            file_id_str = result.get('file_id', 'unknown')
            if file_id_str.upper().startswith("SC"): # Use uppercase comparison
                processed_file_types["SC"] += 1
            elif file_id_str.upper().startswith("ST"): # Use uppercase comparison
                processed_file_types["ST"] += 1
        elif result:
            failed_file_count += 1
            safe_print(f"文件 {result.get('file_id', 'unknown')} 处理失败: {result.get('error', '未知错误')}")

    safe_print(f"文件处理完成。成功: {processed_file_count}, 失败: {failed_file_count}")
    safe_print(f"共聚合 {len(all_aggregated_samples)} 个样本。")
    safe_print(f"聚合数据各阶段原始分布: {dict(total_stage_counts)}")

    if not all_aggregated_samples:
        safe_print("没有成功聚合任何样本，无法继续。")
        return

    # --- 训练/测试集分割 (9:1) ---
    safe_print("进行训练/测试集分割 (90% 训练, 10% 测试)...")
    np.random.seed(42) # 保证可复现
    np.random.shuffle(all_aggregated_samples)
    split_index = int(len(all_aggregated_samples) * 0.9)
    raw_train_data = all_aggregated_samples[:split_index]
    raw_test_data = all_aggregated_samples[split_index:]
    safe_print(f"原始训练集样本数: {len(raw_train_data)}, 原始测试集样本数: {len(raw_test_data)}")

    # --- 平衡训练集 ---
    safe_print(f"开始平衡训练集 (策略={balance_strategy}, Alpha={balance_alpha}, 权重={weight_method})...")
    balanced_train_data = balance_dataset(raw_train_data, strategy=balance_strategy,
                                         balance_alpha=balance_alpha, weight_method=weight_method)
    safe_print(f"平衡后训练集样本数: {len(balanced_train_data)}")

    # --- 计算平均 Token 数 ---
    safe_print("计算平均 Token 数 (基于前10个训练样本)...")
    avg_tokens = 0
    num_samples_for_token_calc = min(10, len(balanced_train_data))
    if num_samples_for_token_calc > 0:
        total_tokens_for_avg = 0
        for i in range(num_samples_for_token_calc):
            try:
                # 确保 calculate_tokens 输入的是单个样本字典
                token_info = calculate_tokens(balanced_train_data[i])
                total_tokens_for_avg += token_info.get("total_tokens", 0)
            except Exception as e:
                safe_print(f"计算第 {i+1} 个样本的 token 时出错: {e}")
        avg_tokens = int(total_tokens_for_avg / num_samples_for_token_calc) if num_samples_for_token_calc > 0 else 0
    safe_print(f"计算得到的平均 Token 数: {avg_tokens}")

    # --- 生成文件名 (添加 eeg_step_sec 信息) ---
    timestamp = time.strftime("%Y%m%d%H%M")
    emotion_step_str = str(int(emotion_step_size)) if include_emotion and emotion_step_size.is_integer() else str(emotion_step_size) if include_emotion else "na"
    eeg_window_str = str(int(eeg_window_sec)) if eeg_window_sec.is_integer() else str(eeg_window_sec)
    # --- 新增：格式化 EEG 步长字符串 ---
    eeg_step_str = str(int(eeg_step_sec)) if eeg_step_sec.is_integer() else str(eeg_step_sec)
    # --- 结束新增 ---

    max_windows_str = f"win{max_windows}" if max_windows is not None and max_windows > 0 else "win_all"

    # --- 修改：在 filename_prefix 中添加 eeg_step_str ---
    filename_prefix = f"sleep_{file_type.lower()}_{actual_max_files}_{target_sfreq}hz"
    filename_prefix += f"_eeg{eeg_window_str}s-step{eeg_step_str}s" # <<<--- 修改此处格式
    if include_emotion:
        filename_prefix += f"_emo{emotion_window_length}s-step{emotion_step_str}s" # 可选：更明确情绪参数
    filename_prefix += f"_{max_windows_str}_tok{avg_tokens}"
    # --- 结束修改 ---

    if balance_strategy != "original" and balance_strategy != "none":
        filename_prefix += f"_bal{balance_alpha}_{weight_method}"

    base_filename_part = filename_prefix
    final_filename_prefix = f"{base_filename_part}_{timestamp}"

    train_filename = f"{final_filename_prefix}_train.json"
    test_filename = f"{final_filename_prefix}_test.json"
    stats_excel_filename = f"{final_filename_prefix}_stats.xlsx"
    stats_json_filename = f"{final_filename_prefix}_stats.json"
    # --- 结束文件名生成修改 ---

    # --- 保存训练集和测试集 ---
    def save_dataset(data, filename, output_directory):
        output_path = os.path.join(output_directory, filename)
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            safe_print(f"成功保存 {len(data)} 个样本到: {output_path}")
        except Exception as e:
            safe_print(f"保存文件 {output_path} 时出错: {e}")

    safe_print("保存处理后的数据集...")
    save_dataset(balanced_train_data, train_filename, train_output_dir)
    save_dataset(raw_test_data, test_filename, test_output_dir)

    # --- 统计最终阶段分布 ---
    safe_print("计算最终阶段统计...")
    final_train_counts = count_stages(balanced_train_data)
    final_test_counts = count_stages(raw_test_data)
    # 原始总数统计在前面聚合时已经完成 (total_stage_counts)
    original_total_samples = sum(total_stage_counts.values())

    # 打印统计信息
    print_stage_statistics(all_aggregated_samples, balanced_train_data, raw_test_data, processed_file_types)

    # --- 准备并保存统计信息 (在 params 中添加 eeg_step_sec) ---
    dataset_info = {
        "dataset_name": final_filename_prefix,
        "original_total_samples": original_total_samples,
        "train_samples": len(balanced_train_data),
        "test_samples": len(raw_test_data),
        "original_stage_counts": dict(total_stage_counts),
        "train_stage_counts": final_train_counts,
        "test_stage_counts": final_test_counts,
        "processed_files_count": processed_file_count,
        "failed_files_count": failed_file_count,
        "processed_file_types": processed_file_types,
        "input_dir": input_dir,
        "params": {
            "max_files_requested": max_files,
            "max_files_processed": actual_max_files,
            "target_sfreq": target_sfreq,
            "eeg_window_sec": eeg_window_sec,
            "eeg_step_sec": eeg_step_sec, # <<<--- 添加参数到统计信息
            "include_emotion": include_emotion,
            "emotion_window_length": emotion_window_length if include_emotion else None,
            "emotion_step_size": emotion_step_size if include_emotion else None,
            "resolve_emotion_conflict": resolve_emotion_conflict if include_emotion else None,
            "add_noise": add_noise,
            "normalize_features": normalize_features,
            "balance_strategy": balance_strategy,
            "balance_alpha": balance_alpha,
            "weight_method": weight_method,
            "avg_tokens": avg_tokens,
            "timestamp": timestamp,
            "file_type_filter": file_type,
            "file_pattern_filter": file_pattern,
            "timeout": timeout # 添加 timeout 到 params
        }
    }

    # 添加调试信息：打印即将保存到Excel的数据
    safe_print("\n--- DEBUG: 即将保存到Excel的数据 ---")
    safe_print(f"原始样本总数: {dataset_info.get('original_total_samples', 0)}")
    safe_print(f"训练集样本数: {dataset_info.get('train_samples', 0)}")
    safe_print(f"测试集样本数: {dataset_info.get('test_samples', 0)}")
    safe_print(f"原始阶段计数: {dataset_info.get('original_stage_counts', {})}")
    safe_print(f"训练集阶段计数: {dataset_info.get('train_stage_counts', {})}")
    safe_print(f"测试集阶段计数: {dataset_info.get('test_stage_counts', {})}")
    safe_print(f"处理成功文件数: {dataset_info.get('processed_files_count', 0)}")
    safe_print(f"处理失败文件数: {dataset_info.get('failed_files_count', 0)}")
    safe_print("--- END DEBUG ---\n")

    # 保存统计到 Excel
    excel_path = os.path.join(all_output_dir, stats_excel_filename)
    try:
        save_metrics_to_excel(dataset_info, excel_path)
        safe_print(f"数据集统计信息已保存到: {excel_path}")
    except Exception as e:
        safe_print(f"保存 Excel 统计文件时出错: {e}")

    # 保存统计到 JSON
    json_path = os.path.join(all_output_dir, stats_json_filename)
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(dataset_info, f, ensure_ascii=False, indent=2)
        safe_print(f"数据集统计信息已保存为JSON: {json_path}")
    except Exception as e:
        safe_print(f"保存 JSON 统计文件时出错: {e}")

    # 更新根目录下的 dataset_info.json (如果需要的话)
    # update_dataset_info(train_filename, train_output_dir) # 这个函数可能需要调整或移除

    safe_print("\n--- 处理流程结束 ---")

def balance_dataset(data, strategy="balanced", balance_alpha=0.7, weight_method="sqrt_inverse"):
    """对数据集进行平衡处理
    
    参数:
        data: 要平衡的数据列表，每个元素应该是包含"output"字段的字典
        strategy: 平衡策略，可以是"balanced"(平衡采样)或"original"(保持原始分布)
        balance_alpha: 平衡系数，0表示完全均衡，1表示保持原始分布
        weight_method: 权重计算方法，可以是"inverse"(反比)、"sqrt_inverse"(平方根反比)或"log_inverse"(对数反比)
        
    返回:
        平衡后的数据列表
    """
    if strategy == "original":
        return data
    
    # 统计每个标签的样本数量
    label_counts = Counter([int(item["output"]) for item in data])
    total_samples = len(data)
    
    # 计算目标分布
    if strategy == "balanced":
        # 计算均匀分布
        n_classes = len(label_counts)
        uniform_dist = {label: 1.0/n_classes for label in label_counts.keys()}
        
        # 计算原始分布
        original_dist = {label: count/total_samples for label, count in label_counts.items()}
        
        # 使用balance_alpha融合均匀分布和原始分布
        target_dist = {}
        for label in label_counts.keys():
            target_dist[label] = balance_alpha * original_dist[label] + (1 - balance_alpha) * uniform_dist[label]
        
        # 归一化目标分布
        dist_sum = sum(target_dist.values())
        target_dist = {label: prob/dist_sum for label, prob in target_dist.items()}
    else:
        raise ValueError(f"不支持的平衡策略: {strategy}")
    
    # 根据权重方法计算每个样本的权重
    label_weights = {}
    if weight_method == "inverse":
        # 使用频率的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / count if count > 0 else 0
    elif weight_method == "sqrt_inverse":
        # 使用频率平方根的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / np.sqrt(count) if count > 0 else 0
    elif weight_method == "log_inverse":
        # 使用频率对数的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / np.log(count + 1)
    else:
        raise ValueError(f"不支持的权重方法: {weight_method}")
    
    # 归一化权重
    weight_sum = sum(label_weights.values())
    label_weights = {label: weight/weight_sum for label, weight in label_weights.items()}
    
    # 按标签分组
    samples_by_label = {label: [] for label in label_counts.keys()}
    for item in data:
        label = int(item["output"])
        samples_by_label[label].append(item)
    
    # 计算每个标签需要的样本数
    target_counts = {}
    target_total = total_samples  # 保持总样本数不变
    for label, prob in target_dist.items():
        target_counts[label] = int(target_total * prob)
    
    # 调整目标计数以确保总和等于目标总样本数
    count_diff = target_total - sum(target_counts.values())
    if count_diff != 0:
        # 将差额分配给样本最多的类别
        max_label = max(label_counts, key=label_counts.get)
        target_counts[max_label] += count_diff
    
    # 构建平衡后的数据集
    balanced_data = []
    for label, target_count in target_counts.items():
        available_samples = samples_by_label[label]
        
        if not available_samples:
            print(f"警告: 标签 {label} 没有可用样本")
            continue
        
        # 如果需要的样本比可用的多，进行过采样（随机重复采样）
        if target_count > len(available_samples):
            # 先添加所有可用样本
            sampled = available_samples.copy()
            # 然后随机采样剩余需要的样本
            additional_needed = target_count - len(available_samples)
            additional_samples = np.random.choice(available_samples, size=additional_needed, replace=True)
            sampled.extend(additional_samples)
        # 如果需要的样本比可用的少，进行欠采样（随机选择一部分）
        elif target_count < len(available_samples):
            sampled = list(np.random.choice(available_samples, size=target_count, replace=False))
        # 如果刚好相等，直接使用全部样本
        else:
            sampled = available_samples
        
        balanced_data.extend(sampled)
    
    # 打乱平衡后的数据
    np.random.shuffle(balanced_data)
    
    return balanced_data

def save_metrics_to_excel(dataset_info, output_path):
    """将数据集统计信息保存到Excel文件

    Args:
        dataset_info: 包含数据集统计信息的字典
        output_path: 输出文件路径
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 创建一个新的Excel工作簿
        wb = Workbook()

        # 移除默认的Sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 创建"概览"工作表
        overview_sheet = wb.create_sheet("概览")

        # 添加标题
        overview_sheet["A1"] = "数据集统计信息"
        overview_sheet["A1"].font = Font(size=14, bold=True)
        overview_sheet.merge_cells("A1:E1")
        overview_sheet["A1"].alignment = Alignment(horizontal="center")

        # 添加基本信息
        overview_sheet["A3"] = "数据集名称:"
        overview_sheet["B3"] = dataset_info.get("dataset_name", "未命名")

        overview_sheet["A4"] = "原始样本总数:"
        overview_sheet["B4"] = dataset_info.get("original_total_samples", 0) # 修正键名

        overview_sheet["A5"] = "训练集样本数:"
        overview_sheet["B5"] = dataset_info.get("train_samples", 0)

        overview_sheet["A6"] = "测试集样本数:"
        overview_sheet["B6"] = dataset_info.get("test_samples", 0)

        overview_sheet["A7"] = "平衡策略:"
        overview_sheet["B7"] = dataset_info.get("params", {}).get("balance_strategy", "未知")

        overview_sheet["A8"] = "平衡因子:"
        overview_sheet["B8"] = dataset_info.get("params", {}).get("balance_alpha", "未知")

        overview_sheet["A9"] = "创建日期:"
        # --- 修改：从 params 获取 timestamp ---
        overview_sheet["B9"] = dataset_info.get("params", {}).get("timestamp", "未知")
        # --- 结束修改 ---

        # 设置列宽
        overview_sheet.column_dimensions["A"].width = 20
        overview_sheet.column_dimensions["B"].width = 30

        # 创建"分布统计"工作表
        dist_sheet = wb.create_sheet("分布统计")

        # 添加标题
        dist_sheet["A1"] = "数据集分布统计"
        dist_sheet["A1"].font = Font(size=14, bold=True)
        dist_sheet.merge_cells("A1:G1") # 合并7列以适应表头
        dist_sheet["A1"].alignment = Alignment(horizontal="center")

        # 添加表头
        headers = ["睡眠阶段", "原始总数", "原始占比", "训练集数量", "训练集占比", "测试集数量", "测试集占比"]
        for i, header in enumerate(headers):
            dist_sheet.cell(row=3, column=i+1, value=header) # 从第3行开始
            dist_sheet.cell(row=3, column=i+1).font = Font(bold=True)
            dist_sheet.cell(row=3, column=i+1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # 获取分布数据
        original_counts = dataset_info.get('original_stage_counts', {})
        train_counts = dataset_info.get('train_stage_counts', {})
        test_counts = dataset_info.get('test_stage_counts', {})

        # 计算总量
        original_total = sum(original_counts.values())
        train_total = sum(train_counts.values())
        test_total = sum(test_counts.values())

        # 添加数据
        row = 4
        # 使用 stage_names 字典确保所有阶段都被列出
        stage_names = {
            0: "W (清醒)", 1: "N1 (浅睡眠)", 2: "N2 (轻睡眠)",
            3: "N3 (深睡眠)", 4: "N4 (深睡眠)", 5: "REM (快速眼动)"
        }
        for stage in sorted(stage_names.keys()):
            orig_count = original_counts.get(stage, 0)
            train_count = train_counts.get(stage, 0)
            test_count = test_counts.get(stage, 0)

            orig_percent = orig_count / original_total * 100 if original_total > 0 else 0
            train_percent = train_count / train_total * 100 if train_total > 0 else 0
            test_percent = test_count / test_total * 100 if test_total > 0 else 0

            # 写入行数据
            dist_sheet.cell(row=row, column=1, value=f"{stage}: {stage_names[stage]}")
            dist_sheet.cell(row=row, column=2, value=orig_count)
            dist_sheet.cell(row=row, column=3, value=f"{orig_percent:.2f}%")
            dist_sheet.cell(row=row, column=4, value=train_count)
            dist_sheet.cell(row=row, column=5, value=f"{train_percent:.2f}%")
            dist_sheet.cell(row=row, column=6, value=test_count)
            dist_sheet.cell(row=row, column=7, value=f"{test_percent:.2f}%")

            row += 1

        # 添加总计行
        dist_sheet.cell(row=row, column=1, value="总计")
        dist_sheet.cell(row=row, column=1).font = Font(bold=True)
        dist_sheet.cell(row=row, column=2, value=original_total)
        dist_sheet.cell(row=row, column=3, value="100.00%")
        dist_sheet.cell(row=row, column=4, value=train_total)
        dist_sheet.cell(row=row, column=5, value="100.00%")
        dist_sheet.cell(row=row, column=6, value=test_total)
        dist_sheet.cell(row=row, column=7, value="100.00%")


        # 设置列宽
        for col_idx, width in enumerate([25, 12, 12, 12, 12, 12, 12], 1):
             dist_sheet.column_dimensions[chr(64 + col_idx)].width = width


        # 创建"参数"工作表
        params_sheet = wb.create_sheet("参数")

        # 添加标题
        params_sheet["A1"] = "处理参数"
        params_sheet["A1"].font = Font(size=14, bold=True)
        params_sheet.merge_cells("A1:B1")
        params_sheet["A1"].alignment = Alignment(horizontal="center")

        # 添加参数
        row = 3
        for key, value in dataset_info.get("params", {}).items():
            params_sheet.cell(row=row, column=1, value=key)
            # --- 修改：确保值是字符串 ---
            params_sheet.cell(row=row, column=2, value=str(value))
            # --- 结束修改 ---
            row += 1

        # 设置列宽
        params_sheet.column_dimensions["A"].width = 25
        params_sheet.column_dimensions["B"].width = 35

        # 创建"文件信息"工作表
        files_sheet = wb.create_sheet("文件信息")

        # 添加标题
        files_sheet["A1"] = "文件处理统计"
        files_sheet["A1"].font = Font(size=14, bold=True)
        files_sheet.merge_cells("A1:B1")
        files_sheet["A1"].alignment = Alignment(horizontal="center")

        # 添加文件统计
        row = 3
        files_sheet.cell(row=row, column=1, value="处理成功文件数:")
        files_sheet.cell(row=row, column=2, value=dataset_info.get("processed_files_count", 0))
        row += 1
        files_sheet.cell(row=row, column=1, value="处理失败文件数:")
        files_sheet.cell(row=row, column=2, value=dataset_info.get("failed_files_count", 0))
        row += 1
        files_sheet.cell(row=row, column=1, value="处理的文件类型:")
        files_sheet.cell(row=row, column=2, value=str(dataset_info.get("processed_file_types", {})))

        # 设置列宽
        files_sheet.column_dimensions["A"].width = 25
        files_sheet.column_dimensions["B"].width = 35


        # 保存工作簿
        wb.save(output_path)
        # safe_print(f"统计信息已成功保存到Excel文件: {output_path}") # 已在主流程打印

    except ImportError:
         safe_print("错误: 需要 'pandas' 和 'openpyxl' 库来保存Excel文件。请运行 'pip install pandas openpyxl'")
    except Exception as e:
        import traceback
        safe_print(f"保存Excel统计信息时出错: {e}")
        traceback.print_exc()
        # ... (CSV fallback logic remains the same) ...
        try:
            import csv
            with open(output_path.replace('.xlsx', '.csv'), 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["key", "value"])
                # --- 修改：更结构化地写入CSV ---
                for section, data in dataset_info.items():
                    if isinstance(data, dict):
                        writer.writerow([section, ""]) # Section header
                        for k, v in data.items():
                            if isinstance(v, dict):
                                writer.writerow([f"  {k}", ""])
                                for sk, sv in v.items():
                                     writer.writerow([f"    {sk}", sv])
                            else:
                                writer.writerow([f"  {k}", v])
                    else:
                        writer.writerow([section, data])
                # --- 结束修改 ---
            safe_print(f"统计信息已保存为CSV文件 (Excel保存失败)")
        except ImportError:
             safe_print("无法保存为CSV，请确保 'csv' 模块可用。")
        except Exception as csv_e:
             safe_print(f"保存CSV统计信息时也出错: {csv_e}")
             safe_print("无法保存统计信息，请检查权限或磁盘空间")


def load_emotion_models(model_dir):
    """加载情绪模型

    参数:
        model_dir: 模型目录路径

    返回:
        加载的四个情绪模型字典 {'HVHA': model1, 'HVLA': model2, 'LVHA': model3, 'LVLA': model4}
    """
    # 模型名称映射
    models = {}
    model_names = ['HVHA', 'HVLA', 'LVHA', 'LVLA']
    model_files = {
        'HVHA': 'emotion_2classes_2Bipolar_HVHA.h5',
        'HVLA': 'emotion_2classes_2Bipolar_HVLA.h5',
        'LVHA': 'emotion_2classes_2Bipolar_LVHA.h5',
        'LVLA': 'emotion_2classes_2Bipolar_LVLA.h5'
    }

    safe_print("开始加载四象限情绪模型，目录: " + model_dir)
    # --- 不再需要打印加载方法，因为我们直接修改了 load_model 调用 ---
    # safe_print("使用加载方法: custom_load_model")

    # 尝试加载每个模型
    for name in model_names:
        model_path = os.path.join(model_dir, model_files[name])
        safe_print(f"开始加载模型 {name} 从 {model_path}")

        try:
            # --- 修改：直接调用 tf.keras.models.load_model 并设置 compile=False ---
            model_file = model_path
            safe_print(f"尝试加载模型文件: {model_file} (用于推理，不编译)")

            if not os.path.exists(model_file):
                 safe_print(f"错误: 模型文件 {model_file} 不存在")
                 continue

            # 使用 tf.keras.models.load_model 加载，明确指定不编译
            safe_print("尝试使用tf.keras.models.load_model(..., compile=False)加载模型")
            try:
                with tf.device('/cpu:0'):
                    # 添加 compile=False 参数
                    model = tf.keras.models.load_model(model_file, compile=False)
                safe_print("成功使用tf.keras.models.load_model(compile=False)加载模型")
            except Exception as e:
                safe_print(f"使用tf.keras.models.load_model(compile=False)加载失败: {e}")
                # 如果 compile=False 失败，可以尝试不带 compile 参数加载（可能某些旧模型需要）
                safe_print("尝试不带 compile 参数再次加载...")
                try:
                    with tf.device('/cpu:0'):
                        model = tf.keras.models.load_model(model_file)
                    safe_print("成功使用tf.keras.models.load_model加载模型 (无 compile 参数)")
                except Exception as e_fallback:
                     safe_print(f"不带 compile 参数加载也失败: {e_fallback}")
                     continue # 跳过此模型
            # --- 结束修改 ---

            # 获取模型的输入和输出形状
            input_shape = model.input_shape
            output_shape = model.output_shape
            safe_print(f"模型 {name} 输入形状: {input_shape}")
            safe_print(f"模型 {name} 输出形状: {output_shape}")

            # 检查输出形状是否合法（预期为(None, 2)，表示二分类）
            if len(output_shape) < 2 or output_shape[-1] != 2: # 更灵活的检查
                safe_print(f"警告: 模型 {name} 输出形状 {output_shape} 可能不符合预期的二分类(..., 2)")

            # 将模型添加到字典
            models[name] = model
            safe_print(f"成功加载模型 {name}")

        except Exception as e:
            safe_print(f"加载模型 {name} 失败: {e}")
            # 记录详细错误信息
            import traceback
            traceback.print_exc()

    # 验证是否成功加载所有4个模型
    if len(models) != 4:
        safe_print(f"警告: 未能加载全部四个情绪模型，只加载了 {len(models)}/4 个模型")
    else:
         safe_print(f"已成功加载全部 4 个情绪模型。") # 确认成功

    # ... (模型兼容性测试代码保持不变) ...
    if models:
        # 创建一个随机测试输入 - 10个特征点
        test_input = np.random.random((1, 10, 1))
        safe_print(f"测试模型兼容性，输入形状: {test_input.shape}")

        # 测试每个模型
        models_to_remove = []
        for name, model in models.items():
            try:
                with tf.device('/cpu:0'):
                    safe_print(f"测试模型 {name}...")
                    output = model.predict(test_input, verbose=0)
                safe_print(f"模型 {name} 测试成功，输出形状: {output.shape}")
            except Exception as e:
                safe_print(f"模型 {name} 测试失败: {e}")
                import traceback
                traceback.print_exc()
                # 标记要删除的模型
                models_to_remove.append(name)

        # 移除不兼容的模型
        for name in models_to_remove:
            safe_print(f"移除不兼容模型: {name}")
            models.pop(name, None)

        # 再次检查模型数量
        if len(models) == 0:
            safe_print("错误: 所有模型测试都失败")
            return None

    return models

def extract_emotion_features(signal, sfreq, window_length_sec, step_size_sec, normalize=False, print_details=False):
    """ 提取情绪特征 - 使用指定的窗口长度和步长滑动。
        每个步长生成一个特征向量。
        内部特征计算使用固定的2秒子窗口和pyeeg.bin_power。

    Args:
        signal: 信号数据，形状 [n_channels, n_points]
        sfreq: 采样频率 (Hz)
        window_length_sec: 用于提取特征的窗口长度 (秒, e.g., 2.0)
        step_size_sec: 滑动窗口的步长 (秒, e.g., 5.0)
        normalize: 是否标准化特征
        print_details: 是否打印详细信息

    Returns:
        features: 提取的特征矩阵，形状 (num_steps, 10, 1) 或空数组
    """
    # 检查信号有效性
    if signal is None or signal.size == 0 or len(signal.shape) != 2:
        if print_details: safe_print("错误: 输入信号无效或形状不正确")
        return np.array([])

    channel_num = signal.shape[0]
    n_points = signal.shape[1]

    # 确保至少有2个通道
    if channel_num < 2:
        if print_details: safe_print(f"错误: 需要至少2个通道，但只接收到 {channel_num} 个")
        return np.array([])

    # 外层窗口和步长 (来自参数)
    outer_window_samples = int(window_length_sec * sfreq)
    outer_step_samples = int(step_size_sec * sfreq)
    if outer_step_samples <= 0: # 保证步长至少为1
        outer_step_samples = 1
        if print_details: safe_print(f"警告: 情绪步长计算为0或负数，强制设为1采样点")


    # 内层特征提取参数 (fixed from 4dataloader.py logic for bin_power)
    internal_band = [4, 8, 12, 16, 25, 45]
    # internal_window_size 理论上应等于 outer_window_samples，因为我们对整个2s窗口算一次特征
    internal_window_size = outer_window_samples # 使用外层窗口长度作为内部计算长度


    # 检查信号长度是否足够容纳至少一个外层窗口
    if n_points < outer_window_samples:
        if print_details: safe_print(f"信号太短({n_points}点)，无法提取外层窗口特征(需要{outer_window_samples}点)")
        return np.array([])

    # 存储每个外层窗口提取出的最终特征向量 (每个向量长度为10)
    final_feature_vectors = []

    # --- 外层循环：按 step_size_sec (e.g., 5s) 滑动 ---
    outer_start = 0
    while outer_start + outer_window_samples <= n_points:
        # 提取当前外层窗口数据 (e.g., 2s long)
        current_outer_window_data = signal[:, outer_start : outer_start + outer_window_samples]

        # --- 特征计算：为这个外层窗口计算一个10维特征向量 ---
        feature_vector_for_outer_window = []
        try:
            # 检查提取的数据段是否满足计算要求 (长度应等于 outer_window_samples)
            if current_outer_window_data.shape[1] == outer_window_samples:
                for ch_idx in range(channel_num):
                    # 直接对这个窗口数据计算 bin_power
                    power_values, _ = pe.bin_power(current_outer_window_data[ch_idx], internal_band, sfreq)
                    if len(power_values) == 5:
                        feature_vector_for_outer_window.extend(power_values)
                    else:
                        # Pad if bin_power fails or returns unexpected length
                        feature_vector_for_outer_window.extend([0.0] * 5)
                        if print_details: safe_print(f"警告: 在外层窗口 {outer_start}, 通道 {ch_idx} 的 pe.bin_power 返回长度不为5 (得到 {len(power_values)})")

                if len(feature_vector_for_outer_window) == 10:
                    final_feature_vectors.append(feature_vector_for_outer_window)
                else:
                    # Append zeros if channel concatenation failed
                    final_feature_vectors.append([0.0] * 10)
                    if print_details: safe_print(f"警告: 在外层窗口 {outer_start} 特征拼接后长度不为10")
            else:
                 # Should not happen with the loop condition, but as a safeguard
                 final_feature_vectors.append([0.0] * 10)
                 if print_details: safe_print(f"警告: 在外层窗口 {outer_start} 提取的数据长度 ({current_outer_window_data.shape[1]}) 与预期 ({outer_window_samples}) 不符")

        except Exception as e:
            final_feature_vectors.append([0.0] * 10) # Append zeros on error
            if print_details: safe_print(f"在外层窗口 {outer_start} 计算特征时出错: {e}")

        # 移动到下一个外层窗口的起始点
        outer_start += outer_step_samples

    # --- 后处理 ---
    if not final_feature_vectors:
        if print_details: safe_print("未能提取任何外层窗口特征")
        return np.array([])

    features_matrix = np.array(final_feature_vectors) # Shape: (num_outer_windows, 10)
    num_steps_generated = features_matrix.shape[0]
    if print_details: safe_print(f"使用 {window_length_sec}s 窗口, {step_size_sec}s 步长，生成了 {num_steps_generated} 个特征向量")


    if normalize:
        try:
            scaler = StandardScaler()
            features_matrix = scaler.fit_transform(features_matrix)
        except Exception as e:
             if print_details: safe_print(f"特征标准化时出错: {e}")
             # Return unnormalized if error

    # --- 重塑以匹配模型输入 (N, 10, 1) ---
    # N 现在是 num_outer_windows
    try:
        final_features = features_matrix.reshape(features_matrix.shape[0], features_matrix.shape[1], 1)
        if print_details: safe_print(f"特征提取完成，最终形状: {final_features.shape}") # 应为 (num_outer_windows, 10, 1)
        return final_features
    except Exception as e:
        if print_details: safe_print(f"特征重塑时出错: {e}")
        return np.array([])

def predict_emotions_multi_model(emotion_models, features, resolve_contradictions=False, device='cpu'):
    """使用多个模型预测情绪 - 为每个内部窗口生成预测，返回序列"""
    # ... (参数和特征有效性检查不变) ...
    if len(features.shape) != 3 or features.shape[1] != 10 or features.shape[2] != 1:
        safe_print(f"警告: 输入特征形状 {features.shape} 与预期 (N, 10, 1) 不符")
        return [], "错误", {"error": f"特征形状错误: {features.shape}"} # 返回空列表

    num_internal_windows = features.shape[0]
    
    try:
        # 用于存储每个内部窗口的四模型预测结果和置信度
        window_predictions = {name: [] for name in emotion_models.keys()}
        window_confidences = {name: [] for name in emotion_models.keys()}

        # 强制在CPU上预测
        with tf.device('/cpu:0'):
            # 遍历每个情绪模型
            for model_name, model in emotion_models.items():
                # 预测所有内部窗口
                probs = model.predict(features, batch_size=num_internal_windows, verbose=0)
                # probs 的形状是 (num_internal_windows, 2)

                # 获取每个内部窗口的预测类别 (0 或 1) 和置信度
                preds_class_1 = (np.argmax(probs, axis=1) == 1).astype(int)
                confs = np.max(probs, axis=1) # 取每个窗口预测类别的概率作为置信度

                window_predictions[model_name] = preds_class_1
                window_confidences[model_name] = confs

        # --- 组合预测结果 ---
        # 为每个内部窗口生成一个4位编码
        combined_codes_list = []
        for i in range(num_internal_windows):
            code = (
                f"{window_predictions['HVHA'][i]}"
                f"{window_predictions['HVLA'][i]}"
                f"{window_predictions['LVHA'][i]}"
                f"{window_predictions['LVLA'][i]}"
            )
            combined_codes_list.append(code)

        # --- 解决矛盾（如果需要）---
        final_codes_list = combined_codes_list
        if resolve_contradictions:
             # 准备置信度列表，形状为 (num_internal_windows, 4)
             confidences_list_per_window = []
             for i in range(num_internal_windows):
                 confidences_list_per_window.append([
                     window_confidences['HVHA'][i],
                     window_confidences['HVLA'][i],
                     window_confidences['LVHA'][i],
                     window_confidences['LVLA'][i]
                 ])
             # 解决每个窗口的矛盾
             final_codes_list = resolve_prediction_contradiction(combined_codes_list, confidences=confidences_list_per_window)

        # --- 计算主导情绪（作为单一代表值，可选）---
        if final_codes_list:
            dominant_code = Counter(final_codes_list).most_common(1)[0][0]
            dominant_label = EMOTION_MAPPINGS.get(dominant_code, "未知情绪")
        else:
            dominant_code = "0000" # 或其他默认值
            dominant_label = EMOTION_MAPPINGS.get(dominant_code)

        # 返回完整的情绪编码序列、主导标签和详细信息
        detailed_info = {
            "predictions_per_window": window_predictions, # 每个模型对每个窗口的预测
            "confidences_per_window": window_confidences, # 每个模型对每个窗口的置信度
            "raw_codes_sequence": combined_codes_list,    # 原始组合编码序列
            "final_codes_sequence": final_codes_list      # 最终（可能解决矛盾后）的编码序列
        }

        # 返回最终的序列，以及主导标签和信息
        return final_codes_list, dominant_label, detailed_info

    except Exception as e:
        safe_print(f"情绪预测时出错: {e}")
        import traceback
        traceback.print_exc()
        # 返回空列表或错误指示
        return [], "错误", {"error": f"预测异常: {str(e)}"}

def resolve_prediction_contradiction(binary_codes_list, confidences=None):
    """解决一系列情绪预测中的矛盾
    
    Args:
        binary_codes_list: 四位二进制字符串的列表
        confidences: 置信度列表的列表，每个内部列表包含 [HVHA, HVLA, LVHA, LVLA] 的置信度
                     如果提供，长度应与 binary_codes_list 相同
                     
    Returns:
        resolved_codes_list: 解决矛盾后的四位二进制字符串列表
    """
    resolved_codes_list = []
    num_codes = len(binary_codes_list)
    
    # 检查置信度列表是否有效
    use_confidences = confidences is not None and len(confidences) == num_codes

    for i in range(num_codes):
        binary_code = binary_codes_list[i]
        current_confidences = confidences[i] if use_confidences else None
        
        # --- 单个代码的矛盾解决逻辑 (与之前类似，但作用于单个代码) ---
        
        # 如果代码非法，添加默认值并继续
        if not binary_code or not isinstance(binary_code, str) or len(binary_code) != 4 or not all(c in '01' for c in binary_code):
            resolved_codes_list.append("0000")
            continue

        original_code = binary_code
        resolved = list(binary_code)

        # 判断是否存在矛盾
        h1, h2, h3, h4 = (bit == '1' for bit in binary_code) # HVHA, HVLA, LVHA, LVLA
        arousal_conflict = (h1 and h3) or (h2 and h4) # 高唤醒 vs 低唤醒
        valence_conflict = (h1 and h2) or (h3 and h4) # 高价 vs 低价
        
        if not arousal_conflict and not valence_conflict:
            resolved_codes_list.append(original_code) # 没有矛盾
            continue

        # --- 开始解决矛盾 ---
        positions = ['HVHA', 'HVLA', 'LVHA', 'LVLA']

        if current_confidences and len(current_confidences) == 4:
            # 按置信度解决
            confidence_dict = {pos: conf for pos, conf in zip(positions, current_confidences)}
            sorted_emotions = sorted(confidence_dict.items(), key=lambda x: x[1], reverse=True)
            highest_emotion = sorted_emotions[0][0]
            highest_idx = positions.index(highest_emotion)

            if sum(int(bit) for bit in binary_code) > 1: # 如果有多个激活
                resolved = ['0'] * 4
                resolved[highest_idx] = '1'
                # safe_print(f"窗口 {i}: 根据置信度选择情绪 {highest_emotion}") # 可以取消注释用于调试
        else:
            # 按规则解决 (没有置信度信息)
            if binary_code == "1111":
                resolved = ['1', '0', '0', '0'] # 只保留HVHA
            elif sum(int(bit) for bit in binary_code) > 1:
                priority_order = [0, 2, 1, 3]  # HVHA, LVHA, HVLA, LVLA 优先级
                for idx in priority_order:
                    if binary_code[idx] == '1':
                        resolved = ['0'] * 4
                        resolved[idx] = '1'
                        break
        
        resolved_code = ''.join(resolved)

        # 避免全零输出
        if resolved_code == "0000" and '1' in original_code:
            for j in range(4):
                if original_code[j] == '1':
                    resolved = ['0'] * 4
                    resolved[j] = '1'
                    resolved_code = ''.join(resolved)
                    break
        
        # if original_code != resolved_code: # 可以取消注释用于调试
        #     safe_print(f"窗口 {i}: 情绪矛盾解决: 从 {original_code} 修改为 {resolved_code}")

        resolved_codes_list.append(resolved_code)
        
    return resolved_codes_list

def update_dataset_info(train_filename, output_dir):
    """更新dataset_info.json文件中的信息，增强错误处理能力"""
    dataset_info_path = "/data/lhc/projects/LLaMA-Factory/data/dataset_info.json"
    backup_path = dataset_info_path + ".bak"
    
    # 备份原始文件
    try:
        shutil.copy2(dataset_info_path, backup_path)
        print(f"已创建备份文件: {backup_path}")
    except Exception as e:
        print(f"创建备份文件失败: {str(e)}")
    
    # 创建新的dataset_info数据
    train_key = train_filename.replace('.json', '')
    
    # 提取token数量和文件类型
    token_count = 0
    match = re.search(r'tok(\d+)_', train_filename)
    if match:
        token_count = int(match.group(1))
    
    # 提取文件类型前缀
    file_type_prefix = ""
    if train_filename.startswith('SC_'):
        file_type_prefix = "SC"
    elif train_filename.startswith('ST_'):
        file_type_prefix = "ST"
    elif train_filename.startswith('SC_ST_'):
        file_type_prefix = "SC_ST"
    
    # 读取训练文件获取样本数量和各阶段分布
    train_file_path = os.path.join(output_dir, train_filename)
    sample_count = 0
    stage_distribution = {}

    
    try:
        if os.path.exists(train_file_path):
            with open(train_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                # 现在文件格式是直接的数据列表，不需要检查格式
                sample_count = len(data)
                
                # 手动计算各阶段分布
                stage_counts = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0}
                
                for item in data:
                    if "output" in item:
                        try:
                            stage = int(item["output"])
                            stage_counts[stage] = stage_counts.get(stage, 0) + 1
                        except (ValueError, KeyError):
                            pass
                
                # 计算百分比
                stage_percentages = {}
                for stage, count in stage_counts.items():
                    percentage = (count / sample_count * 100) if sample_count > 0 else 0
                    stage_percentages[stage] = round(percentage, 2)

                
                # 生成阶段分布信息
                stage_names = {
                    0: "N3 (深睡眠)",
                    1: "N2 (轻睡眠)",
                    2: "N1 (浅睡眠)",
                    3: "REM (快速眼动)",
                    4: "W (清醒)",
                    5: "其他 (未定义或运动)"
                }
                
                stage_distribution = {
                    "counts": stage_counts,
                    "percentages": stage_percentages,
                    "names": stage_names
                }
        else:
            print(f"警告：找不到训练文件 {train_file_path} 来计算样本数量")
    except Exception as e:
        print(f"读取训练文件获取样本数量时出错: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # 确定相对路径
    relative_path = os.path.relpath(output_dir, "/data/lhc/projects/LLaMA-Factory/data")
    
    # 创建数据集信息
    dataset_entry = {
        "script_url": relative_path,
        "file_name": train_filename,
        "columns": {
            "prompt": "instruction",
            "query": "input",
            "response": "output",
            "system": "system"
        },
        "token_count": token_count,
        "sample_count": sample_count,
        "stage_distribution": stage_distribution
    }
    
    # 添加文件类型信息
    if file_type_prefix:
        dataset_entry["file_type"] = file_type_prefix
    
    # 读取现有的dataset_info数据
    current_data = {}
    try:
        with open(dataset_info_path, 'r', encoding='utf-8') as f:
            current_data = json.load(f)
    except Exception as e:
        print(f"读取dataset_info.json失败: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # 更新数据并保存
    current_data[train_key] = dataset_entry
    
    try:
        with open(dataset_info_path, 'w', encoding='utf-8') as f:
            json.dump(current_data, f, ensure_ascii=False, indent=2)
        print(f"更新dataset_info.json成功: 添加了{train_key}")
        return True
    except Exception as e:
        print(f"更新dataset_info.json失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def count_stages(data):
    """统计数据集中各睡眠阶段的样本数"""
    stage_counts = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for item in data:
        if "output" in item:
            try:
                stage = int(item["output"])
                if stage in stage_counts:
                    stage_counts[stage] += 1
                else:
                    safe_print(f"警告: 在 count_stages 中发现无效的 stage 值: {stage}，跳过样本。")
            except (ValueError, TypeError):
                 safe_print(f"警告: 在 count_stages 中无法将 output '{item.get('output')}' 转换为整数，跳过样本。")
        else:
            safe_print(f"警告: 在 count_stages 中发现缺少 'output' 字段的样本，跳过。")
    return stage_counts

def print_stage_statistics(all_data, train_data, test_data, file_types=None):
    """打印各个睡眠阶段的详细样本统计信息

    参数:
        all_data: 所有收集到的数据
        train_data: 平衡训练集
        test_data: 测试集
        file_types: 文件类型统计字典，格式为 {'SC': n, 'ST': m}

    返回:
        元组: (all_counts, train_counts, test_counts) 包含各数据集中不同睡眠阶段的样本计数
    """
    # 定义睡眠阶段名称映射
    stage_names = {
        0: "W (清醒)",      # 修正: 映射到正确标签
        1: "N1 (浅睡眠)",    # 修正: 映射到正确标签
        2: "N2 (轻睡眠)",    # 修正: 映射到正确标签
        3: "N3 (深睡眠)",    # 修正: 映射到正确标签
        4: "N4 (深睡眠)",    # 修正: 映射到正确标签 (N3/N4 通常合并或按需处理)
        5: "REM (快速眼动)" # 修正: 映射到正确标签
    }


    # 计算各数据集的阶段分布 (现在调用的是外部定义的 count_stages)
    all_counts = count_stages(all_data)
    train_counts = count_stages(train_data)
    test_counts = count_stages(test_data)

    # 打印表头
    print("\n" + "="*100)
    print("各睡眠阶段样本数量统计报告")
    print("="*100)

    # 如果提供了文件类型统计，显示文件类型分布
    if file_types:
        print("\n【文件类型分布】")
        print(f"{'文件类型':^15} | {'文件数量':^10} | {'占比':^10}")
        print("-"*40)
        total_files = sum(file_types.values())
        for file_type, count in file_types.items():
            percentage = (count / total_files * 100) if total_files > 0 else 0
            print(f"{file_type:^15} | {count:^10} | {percentage:^8.2f}%")
        print("-"*40)
        print(f"{'总计':^15} | {total_files:^10} | {'100.00':^8}%")

    # 打印原始数据分布
    print("\n【原始数据】")
    print(f"{'睡眠阶段':^25} | {'样本数':^10} | {'占比':^10}")
    print("-"*50)
    total_samples = sum(all_counts.values())
    # --- 修改：使用 stage_names 字典迭代并获取 count ---
    for stage in sorted(stage_names.keys()):
        count = all_counts.get(stage, 0) # 使用 get 获取 count
        percentage = (count / total_samples * 100) if total_samples > 0 else 0
        print(f"{stage}: {stage_names[stage]:20} | {count:^10} | {percentage:^8.2f}%")
    # --- 结束修改 ---
    print("-"*50)
    print(f"{'总计':^25} | {total_samples:^10} | {'100.00':^8}%")

    # 打印平衡后的分布
    print("\n【平衡数据集分布】")
    print(f"{'睡眠阶段':^25} | {'训练集':^10} | {'训练集占比':^10} | {'测试集':^10} | {'测试集占比':^10} | {'总计':^10}")
    print("-"*85)
    # --- 修改：使用 stage_names 字典迭代并获取 count ---
    for stage in sorted(stage_names.keys()):
        train_count = train_counts.get(stage, 0) # 使用 get 获取 count
        test_count = test_counts.get(stage, 0)   # 使用 get 获取 count
        total_count = train_count + test_count

        train_total = sum(train_counts.values())
        test_total = sum(test_counts.values())

        train_percentage = (train_count / train_total * 100) if train_total > 0 else 0
        test_percentage = (test_count / test_total * 100) if test_total > 0 else 0

        print(f"{stage}: {stage_names[stage]:20} | {train_count:^10} | {train_percentage:^10.2f}% | {test_count:^10} | {test_percentage:^10.2f}% | {total_count:^10}")
    # --- 结束修改 ---

    print("-"*85)
    train_total = sum(train_counts.values())
    test_total = sum(test_counts.values())
    print(f"{'总计':^25} | {train_total:^10} | {'100.00':^10}% | {test_total:^10} | {'100.00':^10}% | {train_total+test_total:^10}")

    # 阶段间样本比例分析
    print("\n【数据集平衡性分析】")

    # 原始数据 (添加检查 all_counts 是否为空)
    if all_counts:
        max_stage_orig = max(all_counts.items(), key=lambda x: x[1])
        # 过滤掉 count 为 0 的阶段再找最小值
        min_stage_orig = min([item for item in all_counts.items() if item[1] > 0], key=lambda x: x[1]) if any(c > 0 for c in all_counts.values()) else (None, 0)
        imbalance_ratio_orig = max_stage_orig[1] / min_stage_orig[1] if min_stage_orig[1] > 0 else float('inf')
        print(f"原始数据最多/最少阶段比例: {imbalance_ratio_orig:.2f} ({max_stage_orig[0]}阶段:{max_stage_orig[1]}样本 vs {min_stage_orig[0] if min_stage_orig[0] is not None else 'N/A'}阶段:{min_stage_orig[1]}样本)")
    else:
        print("原始数据为空，无法分析平衡性")


    # 平衡训练集 (添加检查 train_counts 是否为空)
    if train_counts:
        max_stage_bal = max(train_counts.items(), key=lambda x: x[1])
        # 过滤掉 count 为 0 的阶段再找最小值
        min_stage_bal = min([item for item in train_counts.items() if item[1] > 0], key=lambda x: x[1]) if any(c > 0 for c in train_counts.values()) else (None, 0)
        imbalance_ratio_bal = max_stage_bal[1] / min_stage_bal[1] if min_stage_bal[1] > 0 else float('inf')
        print(f"平衡后训练集最多/最少阶段比例: {imbalance_ratio_bal:.2f} ({max_stage_bal[0]}阶段:{max_stage_bal[1]}样本 vs {min_stage_bal[0] if min_stage_bal[0] is not None else 'N/A'}阶段:{min_stage_bal[1]}样本)")
    else:
         print("平衡后训练集为空，无法分析平衡性")


    print("="*100)

    # 返回计数数据供main函数使用
    return all_counts, train_counts, test_counts

def process_multi_emotion_predictions(signals, stages, file_id, emotion_models, output_dir, window_length=2, step_size=0.1, sfreq=100, device='cpu', max_windows=50, timeout_seconds=0, resolve_contradictions=False):
    """处理信号数据，预测多模型情绪，并保存为JSON文件
    
    Args:
        signals: 信号数据 [channels, samples]
        stages: 睡眠阶段数据
        file_id: 文件ID
        emotion_models: 情绪模型字典
        output_dir: 输出目录
        window_length: 窗口长度(秒)
        step_size: 步长(秒)，默认0.1秒
        sfreq: 采样频率
        device: 运行设备（cpu/gpu/auto）
        max_windows: 最大处理窗口数，默认50
        timeout_seconds: 处理超时时间(秒)，0表示不限制时间，默认0秒
        resolve_contradictions: 是否解决矛盾的情绪预测，默认为False
        
    Returns:
        output_file: 输出文件路径
    """
    try:
        # 开始时间
        start_time = time.time()
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 获取信号总时长(秒)
        total_duration_sec = signals.shape[1] / sfreq
        
        # 窗口样本数和步长样本数
        window_samples = int(window_length * sfreq)
        step_samples = max(1, int(step_size * sfreq))  # 至少是1个样本点
        
        # 分段处理信号
        n_samples = signals.shape[1]
        n_channels = signals.shape[0]
        
        # 计算总窗口数
        total_windows = (n_samples - window_samples) // step_samples + 1
        safe_print(f"开始处理情绪预测, 总采样点数: {n_samples}, 窗口大小: {window_samples}, 步长: {step_samples}")
        safe_print(f"信号总时长: {total_duration_sec:.2f}秒, 理论序列长度: {int((total_duration_sec - window_length) / step_size) + 1}个点")
        
        # 限制窗口数量
        if total_windows > max_windows:
            safe_print(f"警告: 窗口数量({total_windows})超过限制({max_windows})，将进行均匀采样")
            # 计算采样间隔，均匀采样max_windows个窗口
            sample_interval = max(1, total_windows // max_windows)
            total_windows = (total_windows + sample_interval - 1) // sample_interval
        else:
            sample_interval = 1
        
        safe_print(f"处理窗口数: {total_windows}, 采样间隔: {sample_interval}")
        
        # 创建批处理大小
        batch_size = 64  # 批处理大小
        
        # 用于存储段索引和预测结果
        segment_indices = []
        combined_predictions = []
        
        # 当前段索引
        segment_idx = 0
        
        # 用于批处理的缓冲区
        features_batch = []
        segment_indices_batch = []
        
        # 进度条
        pbar = tqdm(total=total_windows, desc="处理情绪窗口")
        
        # 遍历信号，分批提取特征和预测
        sample_count = 0
        window_count = 0
        for start in range(0, n_samples - window_samples + 1, step_samples * sample_interval):
            # 检查超时（当timeout_seconds为0时表示不限制时间，跳过超时检查）
            if timeout_seconds > 0 and time.time() - start_time > timeout_seconds:
                safe_print(f"警告: 处理时间超过{timeout_seconds}秒，提前结束处理")
                break
                
            end = start + window_samples
            # 提取当前窗口
            window = signals[:, start:end]
            
            # 提取特征
            features = extract_emotion_features(window, sfreq)
            
            # 添加到批处理缓冲区
            features_batch.append(features)
            segment_indices_batch.append(segment_idx)
            
            # 如果到达下一个完整段的边界，增加段索引
            if (start + step_samples * sample_interval) % (window_samples) == 0 and start > 0:
                segment_idx += 1
            
            # 当积累了一批或处理到最后时进行批量预测
            if len(features_batch) >= batch_size or start + step_samples * sample_interval >= n_samples - window_samples:
                if features_batch:
                    # 批量提取特征
                    batch_features = np.vstack(features_batch)
                    
                    # 批量预测
                    for i in range(0, len(batch_features), batch_size):
                        end_idx = min(i + batch_size, len(batch_features))
                        mini_batch = batch_features[i:end_idx]
                        
                        # 预测情绪
                        emotion_code, _, _ = predict_emotions_multi_model(
                            mini_batch,
                            emotion_models,
                            device=device,
                            resolve_contradictions=resolve_contradictions
                        )
                        combined_predictions.extend(emotion_code)
                        segment_indices.extend(segment_indices_batch[i:end_idx])
                    
                    # 更新进度条
                    pbar.update(len(features_batch))
                    
                    # 清空批处理缓冲区
                    features_batch = []
                    segment_indices_batch = []
            
            window_count += 1
            
            # 定期更新进度信息，减少更新频率
            if window_count % 500 == 0:
                elapsed = time.time() - start_time
                windows_per_sec = window_count / max(1, elapsed)
                eta = (total_windows - window_count) / max(1, windows_per_sec)
                safe_print(f"已处理 {window_count}/{total_windows} 窗口，速度: {windows_per_sec:.2f}窗口/秒，预计剩余时间: {eta:.2f}秒")
        
        # 关闭进度条
        pbar.close()
        
        # 转换为numpy数组
        segment_indices = np.array(segment_indices)
        
        # 处理预测结果
        if combined_predictions:
            safe_print(f"完成预测，总共处理了{len(combined_predictions)}个窗口")
            # 处理预测结果
            results = process_predictions_multi_model(combined_predictions, segment_indices, stages)
            
            # 保存结果到JSON文件
            output_file = os.path.join(output_dir, f"{file_id}_multi_emotion.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                output_data = {
                    "file_id": file_id,
                    "window_length_sec": window_length,
                    "step_size_sec": step_size,
                    "sampling_rate": sfreq,
                    "processing_time_sec": time.time() - start_time,
                    "total_segments": len(np.unique(segment_indices)),
                    "total_windows_processed": len(combined_predictions),
                    "model_types": list(emotion_models.keys()) if emotion_models else [],
                    "estimated_sequence_length": int((total_duration_sec - window_length) / step_size) + 1,
                    "emotion_predictions": results
                }
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            safe_print(f"多模型情绪预测结果已保存到: {output_file}")
            safe_print(f"总处理时间: {time.time() - start_time:.2f}秒")
            return output_file
        else:
            safe_print(f"警告: 未能提取任何情绪特征")
            return None
    except Exception as e:
        safe_print(f"处理多模型情绪预测时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def process_predictions_multi_model(combined_predictions, segment_ids, sleep_stages, output_dir):
    """
    处理多模型组合预测结果
    
    Args:
        combined_predictions: 组合的情绪预测结果，包含(预测结果, 置信度)元组
        segment_ids: 片段ID列表
        sleep_stages: 睡眠阶段列表
        output_dir: 输出目录
    
    Returns:
        result_filename: 输出文件名
    """
    # 提取预测结果和置信度
    predictions = []
    confidences = []
    if isinstance(combined_predictions, tuple) and len(combined_predictions) == 2:
        predictions, confidences = combined_predictions
    else:
        predictions = combined_predictions
        confidences = [[0.0] * 4] * len(combined_predictions)  # 默认置信度
    
    # 按段分组处理预测
    segment_predictions = {}
    for i, (pred, segment_id, sleep_stage) in enumerate(zip(predictions, segment_ids, sleep_stages)):
        # 确保预测是字符串格式
        if not isinstance(pred, str):
            # 如果是数字列表，转换为二进制字符串
            if isinstance(pred, list) and len(pred) == 4:
                pred = ''.join(str(int(p)) for p in pred)
            else:
                # 跳过无效预测
                continue
        
        if segment_id not in segment_predictions:
            segment_predictions[segment_id] = {
                'segment_id': segment_id,
                'sleep_stage': sleep_stage,
                'predictions': [],
                'confidences': []
            }
        
        segment_predictions[segment_id]['predictions'].append(pred)
        if i < len(confidences):
            segment_predictions[segment_id]['confidences'].append(confidences[i])
    
    # 分析每个段的预测
    results = []
    for segment_id, data in segment_predictions.items():
        predictions = data['predictions']
        confs = data['confidences']
        
        # 计算各情绪组合的频率
        counts = Counter(predictions)
        total = len(predictions)
        
        # 计算情绪比例
        emotion_ratios = {code: count/total for code, count in counts.items()}
        
        # 找出最主要的情绪
        dominant_emotion = max(emotion_ratios.items(), key=lambda x: x[1])
        dominant_code = dominant_emotion[0]
        dominant_ratio = dominant_emotion[1]
        
        # 获取情绪标签
        emotion_label = EMOTION_MAPPINGS.get(dominant_code, '未知情绪')
        
        # 构建结果
        result = {
            'segment_id': segment_id,
            'sleep_stage': data['sleep_stage'],
            'dominant_emotion_code': dominant_code,
            'dominant_emotion_label': emotion_label,
            'dominant_ratio': dominant_ratio,
            'emotion_distribution': {code: {'ratio': ratio, 'label': EMOTION_MAPPINGS.get(code, '未知情绪')} 
                                     for code, ratio in emotion_ratios.items()}
        }
        
        results.append(result)
    
    # 保存结果到JSON文件
    result_filename = os.path.join(output_dir, f"emotion_predictions_{time.strftime('%Y%m%d_%H%M%S')}.json")
    with open(result_filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    safe_print(f"情绪预测结果已保存到: {result_filename}")
    return result_filename

def main():
    # --- 将设置启动方式的调用移到 main 函数内部的开头 ---
    # 确保这个调用在任何 ProcessPoolExecutor 实例化之前执行
    try:
        current_method = multiprocessing.get_start_method(allow_none=True)
        # 只有在未设置或为 fork 时才设置 'spawn'
        if current_method is None or current_method == 'fork':
             # 在Linux/macOS上强制使用spawn，有助于避免TensorFlow等库与fork的冲突
             multiprocessing.set_start_method('spawn', force=True)
             safe_print("已在 main 函数内部将多进程启动方式设置为 'spawn'")
        else:
             safe_print(f"多进程启动方式已设置为 '{current_method}'，不再更改。")
    except ValueError as e:
        safe_print(f"警告: 设置多进程启动方式 'spawn' 时出错: {e}. 可能已被设置或不支持。")
    except Exception as e:
         safe_print(f"设置多进程启动方式为 'spawn' 时发生未知错误: {e}")
    # --- 结束移动 ---

    safe_print("[Debug] Entering main function...")
    start_time = time.time()

    parser = argparse.ArgumentParser(description='处理EEG数据生成包含情绪特征的LLM训练样本')

    # --- 使用参数映射简化参数定义 ---
    param_mapping = {
        'input_dir': '/data/lhc/datasets/sleep-edfx',
        # 'output_dir': '/data/lhc/datasets_new/emotion', # 不再直接使用
        'max_files': 44,
        'n_jobs': 44,    # 确认调试时使用 n_jobs=1
        'target_sfreq': 100,
        'balance_strategy': 'balanced', #'none',
        'balance_alpha': 0.5,
        'weight_method': 'sqrt_inverse',
        'file_pattern': None,
        'file_type': 'st',
        'include_emotion': True,
        'emotion_model_dir': '/data/lhc/models/emotion',
        'emotion_window_length': 2.0,
        'emotion_step_size': 0.5,
        'resolve_emotion_conflict': False,
        'device': 'cpu',
        'add_noise': False,
        # --- 修改：更新 max_windows 的注释 ---
        'max_windows': 0, # 每个睡眠阶段最多处理的窗口数。0 或 None 表示不限制。使用正整数（如 1 或 10）进行快速调试。
        # --- 结束修改 ---
        'timeout': 300000,
        'normalize_features': True,
        'eeg_window_sec': 10.0,
        # --- 新增：EEG 窗口滑动步长参数 ---
        'eeg_step_sec': 10.0, # 默认步长等于窗口长度 (无重叠)
        # --- 结束新增 ---
    }

    # --- 定义固定的输出目录 ---
    output_base_dir = '/data/lhc/datasets_new/emotion'
    output_dirs = {
        'train': os.path.join(output_base_dir, 'train'),
        'test': os.path.join(output_base_dir, 'test'),
        'all': os.path.join(output_base_dir, 'all') # 存放统计信息
    }


    # --- 添加命令行参数 (会自动包含新增的 eeg_step_sec) ---
    for param, default in param_mapping.items():
        arg_name = f'--{param.replace("_", "-")}'
        if isinstance(default, bool):
             parser.add_argument(arg_name, action=argparse.BooleanOptionalAction, default=default, help=f'{param} (default: {default})')
        elif isinstance(default, int):
            parser.add_argument(arg_name, type=int, default=default, help=f'{param} (default: {default})')
        elif isinstance(default, float):
             parser.add_argument(arg_name, type=float, default=default, help=f'{param} (default: {default})')
        else:
            parser.add_argument(arg_name, type=str, default=default, help=f'{param} (default: {default})')


    args = parser.parse_args()

    # --- 将解析后的参数更新回字典 (会自动包含新增的 eeg_step_sec) ---
    parsed_params = {}
    for param in param_mapping.keys():
        value = getattr(args, param, param_mapping[param])
        if isinstance(param_mapping[param], bool):
             parsed_params[param] = value if value is not None else param_mapping[param]
        else:
            parsed_params[param] = value


    # --- 强制 n_jobs=1 进行调试 (可选) ---
    # parsed_params['n_jobs'] = 1
    # safe_print(f"注意：已强制 n_jobs = 1 进行调试。")
    # --- 结束强制 ---

    # --- 设置CPU核心数 (如果 n_jobs <= 0) ---
    if parsed_params['n_jobs'] <= 0:
        num_cores = multiprocessing.cpu_count()
        parsed_params['n_jobs'] = max(1, int(num_cores * 0.5))
        safe_print(f"n_jobs <= 0, 自动设置为可用核心的50%: {parsed_params['n_jobs']}")
    else:
        safe_print(f"使用指定的 n_jobs: {parsed_params['n_jobs']}")


    # --- 调用核心处理函数 (使用字典解包传递所有参数，包括新增的 eeg_step_sec) ---
    safe_print("开始处理目录...")
    process_directory(
        output_dirs=output_dirs,
        **parsed_params # 使用字典解包传递所有参数
    )

    end_time = time.time()
    safe_print(f"总处理时间: {end_time - start_time:.2f} 秒")
    safe_print("[Debug] Exiting main function.")

if __name__ == "__main__":
    # 确保在Windows或macOS上使用多进程时，这是入口点
    # multiprocessing.freeze_support() # 如果在非Unix系统上可能需要
    main()