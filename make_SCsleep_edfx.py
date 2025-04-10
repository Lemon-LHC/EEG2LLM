import mne
import numpy as np
import os
import json
from tqdm import tqdm
import shutil
import glob
import re
import multiprocessing
import warnings  # 添加warnings模块
from concurrent.futures import ProcessPoolExecutor, as_completed
from transformers import AutoTokenizer  
from scipy import interpolate  # 添加用于插值的包
from collections import Counter  # 添加用于统计的包
from sklearn.metrics import recall_score, precision_score, f1_score, confusion_matrix  # 添加用于评估指标计算的包

# 在脚本开始时过滤MNE相关警告
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=UserWarning, module="mne")

# 加载分词器
tokenizer_path = "/data/lhc/models/Llama-3.2-1B-Instruct"
tokenizer = None  

def format_signal_data(signal, channel_names, sfreq=None):
    """将信号数据转换为文本格式，使用微伏为单位但不显示单位符号，使用>分隔数据点
    
    参数:
        signal: 信号数据数组
        channel_names: 通道名称列表
        sfreq: 采样频率（Hz），默认为None，则使用100Hz
    """
    formatted_data = []
    # 如果没有提供采样频率，默认使用100Hz
    if sfreq is None:
        sfreq = 100  
    
    # 动态计算时间点，根据采样频率将采样点索引转换为毫秒
    # 1000/sfreq 表示每个采样点之间的毫秒数
    time_points = np.arange(signal.shape[1]) * (1000/sfreq)  
    
    # 放大信号至微伏级别，典型脑电信号在微伏级别
    scale_to_microvolt = 1000000  
    
    # 构建数据点字符串列表
    data_points = []
    for t in range(signal.shape[1]):
        # 将值转换为微伏并确保有足够的精度
        ch1_val = float(signal[0, t]) * scale_to_microvolt
        ch2_val = float(signal[1, t]) * scale_to_microvolt if signal.shape[0] > 1 else 0.0
        
        # 只保留电压值，不包含单位和括号，使用逗号分隔通道值
        point_data = f"{ch1_val:.2f},{ch2_val:.2f}"
        data_points.append(point_data)
    
    # 使用 > 分隔数据点，并在开头添加通道名称
    return f"{', '.join(channel_names)}>{'>'.join(data_points)}>"

def create_llm_sample(signal, stage, channel_names, sfreq):
    """创建用于大模型训练的样本，使用新的格式：instruction/input/output/system
    
    参数:
        signal: 信号数据数组
        stage: 睡眠阶段编号
        channel_names: 通道名称列表
        sfreq: 采样频率（Hz）
    """
    # 获取全局变量window_length_ms
    global window_length_ms
    
    # 将采样频率传递给format_signal_data函数
    formatted_signal = format_signal_data(signal, channel_names, sfreq)
    
    # 计算采样间隔（毫秒）
    interval = int(1000/sfreq)
    
    # 计算窗口长度（秒）
    window_length_sec = window_length_ms / 1000
    
    # 分离指令和输入数据
    instruction = f"""You are a neurobiological expert specializing in EEG data analysis and sleep stage classification. Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) and determine the current sleep stage of the volunteer based on the following classification criteria:
0: Wakefulness (W)
1: Non-rapid eye movement sleep stage 1 (N1)
2: Non-rapid eye movement sleep stage 2 (N2)
3: Non-rapid eye movement sleep stage 3 (N3)
4: Non-rapid eye movement sleep stage 4 (N4)
5: Rapid eye movement sleep stage (R)
The EEG data is provided in the format: 'channel names > data points', where each data point is formatted as 'Fpz-Cz voltage in μV, Pz-Oz voltage in μV' and separated by '>' symbols. For example: 'EEG Fpz-Cz, EEG Pz-Oz>30.12,1.11>-65.46,11.92>-13.17,33.13>'. The data spans {window_length_ms}ms ({window_length_sec} seconds) with a sampling interval of {interval}ms, meaning each data point is {interval}ms apart. In your analysis, pay attention to the following characteristics of each sleep stage:
- Wakefulness (W): High-frequency, low-amplitude waves.
- N1: Low-amplitude, mixed-frequency waves.
- N2: Sleep spindles and K-complexes.
- N3: High-amplitude, low-frequency delta waves.
- N4: Dominant delta waves.
- REM (R): REM sleep has highly distinctive and unique characteristics. It primarily presents with rapid, irregular eye movements visible in EEG as sharp, jagged waveforms. Its core feature is low-amplitude, mixed-frequency EEG activity with prominent theta waves (4-7 Hz). While somewhat similar to N1 stage, REM has distinctive saw-tooth wave patterns, which are key diagnostic markers. Unlike N2 stage, REM lacks sleep spindles and K-complexes. The EEG in REM shows a desynchronized pattern resembling wakefulness but is accompanied by complete loss of muscle tone (muscle atonia). REM may also feature rapid, irregular transient muscle twitches, along with irregular variations in heart rate and respiration. These multiple features collectively constitute the complete picture of REM sleep, making it the most distinctive and readily identifiable among all sleep stages.
Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. Do not include any additional text, punctuation, or explanations. """
    
    # 创建新的格式样本
    sample = {
        "instruction": instruction,
        "input": formatted_signal,
        "output": f"{stage}",
        "system": "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification."
    }
    return sample

def clean_output_directories(base_dir):
    """清理输出目录中的所有json文件"""
    processed_dir = os.path.join(base_dir)
    if os.path.exists(processed_dir):
        print("清理之前生成的json文件...")
        # 只删除json文件，而不是整个目录
        for file_name in os.listdir(processed_dir):
            if file_name.endswith('.json'):
                file_path = os.path.join(processed_dir, file_name)
                try:
                    os.remove(file_path)
                    print(f"已删除: {file_path}")
                except Exception as e:
                    print(f"删除文件 {file_path} 时出错: {str(e)}")
    else:
        # 如果目录不存在，创建它
        os.makedirs(processed_dir, exist_ok=True)

def interpolate_signal(signal, original_sfreq, target_sfreq):
    """对信号进行插值以增加采样率
    
    参数:
        signal: 原始信号数据数组，形状为 [n_channels, n_points]
        original_sfreq: 原始采样频率 (Hz)
        target_sfreq: 目标采样频率 (Hz)
        
    返回:
        插值后的信号数据数组
    """
    if original_sfreq == target_sfreq:
        return signal
    
    # 创建原始时间点（以秒为单位）
    n_points = signal.shape[1]
    original_times = np.arange(n_points) / original_sfreq
    
    # 创建新的时间点（以秒为单位）
    ratio = target_sfreq / original_sfreq
    n_new_points = int(n_points * ratio)
    new_times = np.linspace(0, original_times[-1], n_new_points)
    
    # 对每个通道进行插值
    interpolated_signal = np.zeros((signal.shape[0], n_new_points))
    
    for ch_idx in range(signal.shape[0]):
        # 使用三次样条插值
        f = interpolate.interp1d(original_times, signal[ch_idx, :], kind='cubic', bounds_error=False, fill_value="extrapolate")
        interpolated_signal[ch_idx, :] = f(new_times)
    
    return interpolated_signal

def read_edf_file(edf_path, target_sfreq=100):
    """读取EDF文件并进行预处理"""
    # 使用verbose='error'只显示错误，忽略警告
    raw = mne.io.read_raw_edf(edf_path, preload=True, verbose='error')
    original_sfreq = raw.info['sfreq']
    
    # 检查并保留指定通道
    channels_to_keep = ['EEG Fpz-Cz', 'EEG Pz-Oz']
    available_channels = []
    for ch in channels_to_keep:
        if ch in raw.ch_names:
            available_channels.append(ch)

    if not available_channels:
        raise ValueError("No required channels found in the data")

    # 使用新的API方式避免警告
    raw.pick(available_channels, verbose='error')
    
    # 检查数据是否有效
    data = raw.get_data()
    
    # 检查数据是否为零或接近零
    signal_max = np.max(np.abs(data))
    
    # 数据信号太弱时进行放大（但不在这里转换为微伏，仅放大信号）
    if signal_max < 1e-5:
        # 信号弱则放大到适当范围
        scale_factor = 1.0 / signal_max if signal_max > 0 else 1e6
        data = data * scale_factor
        # 创建新的RawArray对象
        info = raw.info
        raw = mne.io.RawArray(data, info, verbose='error')
    
    # 如果目标采样率与原始采样率不同，则进行重采样
    if target_sfreq != original_sfreq:
        # 使用MNE的resample方法而不是直接修改info['sfreq']
        raw.resample(target_sfreq, npad='auto', verbose='error')
    
    return raw, available_channels

def extract_stage_windows(raw_data, annotations, channel_names):
    """提取固定时长的窗口"""
    stage_mapping = {
        'Sleep stage W': 0,
        'Sleep stage 1': 1,
        'Sleep stage 2': 2,
        'Sleep stage 3': 3,
        'Sleep stage 4': 4,
        'Sleep stage R': 5
    }

    sfreq = raw_data.info['sfreq']  
    # 使用全局变量window_length_ms来确定窗口长度
    global window_length_ms
    # 计算窗口长度对应的秒数
    window_length_sec = window_length_ms / 1000
    # 计算窗口在当前采样率下对应的采样点数
    window_samples = int(sfreq * window_length_sec)  
    signals = raw_data.get_data()  
    
    # 检查信号数据的特征
    signal_max = np.max(np.abs(signals))
    signal_mean = np.mean(np.abs(signals))
    signal_std = np.std(signals)
    
    # 确保信号有足够的变化，便于大模型学习
    if signal_std < 0.001 and signal_max > 0:
        # 添加一些高频变化使信号更具特征
        np.random.seed(42)  
        signal_range = signal_max * 0.1  
        noise = np.random.normal(0, signal_range, signals.shape)
        signals = signals + noise
    
    features = []
    window_count = 0
    accepted_windows = 0
    
    # 检测不同的窗口模式，以确保样本多样性
    processed_patterns = set()

    for annot in annotations:
        stage = annot['description']
        if stage not in stage_mapping:
            continue

        start_idx = int(annot['onset'] * sfreq)
        end_idx = int((annot['onset'] + annot['duration']) * sfreq)

        for win_start in range(start_idx, end_idx - window_samples + 1, window_samples):
            window_count += 1
            win_end = win_start + window_samples
            
            if win_end <= signals.shape[1]:
                window = signals[:, win_start:win_end]  
                
                # 检查窗口是否有意义
                window_max = np.max(np.abs(window))
                window_std = np.std(window)
                
                # 跳过没有变化的窗口
                if window_std < 0.0001 * signal_max:
                    continue
                
                # 创建窗口特征指纹，使用统计特征而不是原始值
                pattern_key = f"{np.mean(window[0]):.3f}_{np.std(window[0]):.3f}_{np.mean(window[1]):.3f}_{np.std(window[1]):.3f}"
                
                # 跳过高度相似的窗口，但保留各个阶段的代表性样本
                stage_pattern_key = f"{stage}_{pattern_key}"
                if stage_pattern_key in processed_patterns and np.random.random() > 0.1:  
                    continue
                processed_patterns.add(stage_pattern_key)
                
                sample = create_llm_sample(window, stage_mapping[stage], channel_names, sfreq)
                features.append(sample)
                accepted_windows += 1

    # 打印窗口提取统计信息
    acceptance_rate = accepted_windows / window_count * 100 if window_count > 0 else 0
    print(f"窗口提取: 总计检查 {window_count} 个窗口，接受 {accepted_windows} 个窗口 ({acceptance_rate:.1f}%)")
    
    # 检查每个阶段是否有足够的样本
    stage_counts = {}
    for feature in features:
        stage = int(feature["output"])
        if stage not in stage_counts:
            stage_counts[stage] = 0
        stage_counts[stage] += 1
    
    # 打印各阶段样本数量
    print("各睡眠阶段样本数量:")
    for stage in range(6):
        count = stage_counts.get(stage, 0)
        print(f"阶段 {stage}: {count} 个样本" + (" - 警告: 样本不足" if count < 10 else ""))
    
    return features

def find_annotation_file(edf_path):
    """根据信号文件路径找到对应的注释文件路径"""
    base_name = os.path.basename(edf_path)
    dir_name = os.path.dirname(edf_path)
    
    # 提取基本ID，例如从'SC4001E0-PSG.edf'提取'SC4001'
    match = re.match(r'(SC\d+)[A-Z]\d+-PSG\.edf', base_name)
    if not match:
        return None
    
    base_id = match.group(1)
    
    # 查找对应的注释文件
    annotation_pattern = os.path.join(dir_name, f"{base_id}*-Hypnogram.edf")
    annotation_files = glob.glob(annotation_pattern)
    
    if not annotation_files:
        return None
    
    return annotation_files[0]

def process_and_save_direct(edf_path, annotation_path, output_base_dir, target_sfreq=100):
    """直接处理并保存到最终目录结构，确保各数据集中标签分布平衡
    
    参数:
        edf_path: EDF文件路径
        annotation_path: 注释文件路径
        output_base_dir: 输出基础目录
        target_sfreq: 目标采样频率 (Hz)
    """
    try:
        print(f"处理文件 {os.path.basename(edf_path)}...")
        raw, channel_names = read_edf_file(edf_path, target_sfreq)
        annotations = mne.read_annotations(annotation_path)
        features = extract_stage_windows(raw, annotations, channel_names)

        if not features:
            print("未提取到有效特征!")
            return 0

        # 创建输出目录
        processed_dir = os.path.join(output_base_dir)
        os.makedirs(processed_dir, exist_ok=True)

        # 按标签分组
        stage_features = {0: [], 1: [], 2: [], 3: [], 4: [], 5: []}
        for feature in features:
            stage = int(feature["output"])
            stage_features[stage].append(feature)
        
        # 转换格式为Alpaca格式
        alpaca_data = []
        system_prompt = "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification."
        # 计算窗口长度（秒）
        window_length_sec = window_length_ms / 1000
        
        task_instruction = f"""Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) and determine the current sleep stage of the volunteer based on the following classification criteria:
0: Wakefulness (W)
1: Non-rapid eye movement sleep stage 1 (N1)
2: Non-rapid eye movement sleep stage 2 (N2)
3: Non-rapid eye movement sleep stage 3 (N3)
4: Non-rapid eye movement sleep stage 4 (N4)
5: Rapid eye movement sleep stage (R)
The EEG data is provided in the format: 'channel names > data points', where each data point is formatted as 'Fpz-Cz voltage in μV, Pz-Oz voltage in μV' and separated by '>' symbols. For example: 'EEG Fpz-Cz, EEG Pz-Oz>30.12,1.11>-65.46,11.92>-13.17,33.13>'. The data spans {window_length_ms}ms ({window_length_sec} seconds) with a sampling interval that depends on the original sampling rate. In your analysis, pay attention to the following characteristics of each sleep stage:
- Wakefulness (W): High-frequency, low-amplitude waves.
- N1: Low-amplitude, mixed-frequency waves.
- N2: Sleep spindles and K-complexes.
- N3: High-amplitude, low-frequency delta waves.
- N4: Dominant delta waves.
- REM (R): REM sleep has highly distinctive and unique characteristics. It primarily presents with rapid, irregular eye movements visible in EEG as sharp, jagged waveforms. Its core feature is low-amplitude, mixed-frequency EEG activity with prominent theta waves (4-7 Hz). While somewhat similar to N1 stage, REM has distinctive saw-tooth wave patterns, which are key diagnostic markers. Unlike N2 stage, REM lacks sleep spindles and K-complexes. The EEG in REM shows a desynchronized pattern resembling wakefulness but is accompanied by complete loss of muscle tone (muscle atonia). REM may also feature rapid, irregular transient muscle twitches, along with irregular variations in heart rate and respiration. These multiple features collectively constitute the complete picture of REM sleep, making it the most distinctive and readily identifiable among all sleep stages.
Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. Do not include any additional text, punctuation, or explanations. If the data is insufficient to determine the sleep stage, return -1."""

        for feature in features:
            # 直接使用新格式的数据
            alpaca_sample = {
                "instruction": task_instruction,
                "input": feature["input"],
                "output": feature["output"],
                "system": system_prompt
            }
            alpaca_data.append(alpaca_sample)

        # 分层抽样：为每个标签分配训练/测试集
        train_data = []
        test_data = []
        
        np.random.seed(42)  
        
        for stage, stage_list in stage_features.items():
            if not stage_list:
                print(f"警告: 标签 {stage} 没有样本")
                continue
                
            # 打乱每个标签的样本
            indices = np.random.permutation(len(stage_list))
            stage_list = [stage_list[i] for i in indices]
            
            # 按比例分配 (90%训练，10%测试)
            split_idx = int(len(stage_list) * 0.9)
            
            train_data.extend(stage_list[:split_idx])
            test_data.extend(stage_list[split_idx:])
        
        # 转换格式为Alpaca格式
        def convert_to_alpaca(data):
            return [{
                "instruction": item["instruction"],
                "input": item["input"],
                "output": item["output"],
                "system": item["system"]
            } for item in data]

        train_alpaca = convert_to_alpaca(train_data)
        test_alpaca = convert_to_alpaca(test_data)

        # 保存到单个文件（新增保存完整数据集）
        def save_single_file(data, filename):
            output_path = os.path.join(processed_dir, filename)
            try:
                # 检查文件是否存在
                file_exists = os.path.exists(output_path)
                
                if file_exists:
                    # 如果文件存在，先读取现有内容
                    with open(output_path, 'r', encoding='utf-8') as f:
                        try:
                            existing_data = json.load(f)
                        except json.JSONDecodeError:
                            existing_data = []
                    # 合并数据
                    combined_data = existing_data + data
                else:
                    combined_data = data
                
                # 写入合并后的数据
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(combined_data, f, ensure_ascii=False, indent=2)
                
                print(f"成功{'追加' if file_exists else '保存'} {len(data)} 个样本到 {filename}，文件中现有样本总数：{len(combined_data)}")
                return len(combined_data)  
            except Exception as e:
                print(f"保存文件错误 {output_path}: {str(e)}")
                return 0  

        # 保存完整数据集
        save_single_file(alpaca_data, "all_data.json")
        save_single_file(train_alpaca, "train.json")
        save_single_file(test_alpaca, "test.json")

        return len(train_data) + len(test_data)

    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0

def process_single_file(edf_file, target_sfreq=100):
    """处理单个EDF文件并返回提取的特征
    
    参数:
        edf_file: EDF文件路径
        target_sfreq: 目标采样频率 (Hz)
    """
    # 查找对应的注释文件
    annotation_file = find_annotation_file(edf_file)
    
    if annotation_file:
        try:
            # 处理文件并获取特征
            raw, channel_names = read_edf_file(edf_file, target_sfreq)
            annotations = mne.read_annotations(annotation_file)
            features = extract_stage_windows(raw, annotations, channel_names)
            
            if features:
                print(f"从 {os.path.basename(edf_file)} 提取了 {len(features)} 个样本")
                return features
            else:
                print(f"{os.path.basename(edf_file)}: 未提取到有效特征!")
                return []
        except Exception as e:
            print(f"处理文件 {os.path.basename(edf_file)} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    else:
        print(f"找不到 {os.path.basename(edf_file)} 的注释文件")
        return []

def balance_dataset(data, strategy="balanced", balance_alpha=0.7, weight_method="sqrt_inverse"):
    """对数据集进行平衡处理
    
    参数:
        data: 要平衡的数据列表，每个元素应该是包含"output"字段的字典
        strategy: 平衡策略，可以是"balanced"(平衡采样)或"original"(保持原始分布)
        balance_alpha: 平衡系数，0表示完全均衡，1表示保持原始分布
        weight_method: 权重计算方法，可以是"inverse"(反比)、"sqrt_inverse"(平方根反比)或"log_inverse"(对数反比)
        
    返回:
        平衡后的数据列表
    """
    if strategy == "original":
        return data
    
    # 统计每个标签的样本数量
    label_counts = Counter([int(item["output"]) for item in data])
    total_samples = len(data)
    
    # 计算目标分布
    if strategy == "balanced":
        # 计算均匀分布
        n_classes = len(label_counts)
        uniform_dist = {label: 1.0/n_classes for label in label_counts.keys()}
        
        # 计算原始分布
        original_dist = {label: count/total_samples for label, count in label_counts.items()}
        
        # 使用balance_alpha融合均匀分布和原始分布
        target_dist = {}
        for label in label_counts.keys():
            target_dist[label] = balance_alpha * original_dist[label] + (1 - balance_alpha) * uniform_dist[label]
        
        # 归一化目标分布
        dist_sum = sum(target_dist.values())
        target_dist = {label: prob/dist_sum for label, prob in target_dist.items()}
    else:
        raise ValueError(f"不支持的平衡策略: {strategy}")
    
    # 根据权重方法计算每个样本的权重
    label_weights = {}
    if weight_method == "inverse":
        # 使用频率的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / count if count > 0 else 0
    elif weight_method == "sqrt_inverse":
        # 使用频率平方根的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / np.sqrt(count) if count > 0 else 0
    elif weight_method == "log_inverse":
        # 使用频率对数的倒数作为权重
        for label, count in label_counts.items():
            label_weights[label] = 1.0 / np.log(count + 1)
    else:
        raise ValueError(f"不支持的权重方法: {weight_method}")
    
    # 归一化权重
    weight_sum = sum(label_weights.values())
    label_weights = {label: weight/weight_sum for label, weight in label_weights.items()}
    
    # 按标签分组
    samples_by_label = {label: [] for label in label_counts.keys()}
    for item in data:
        label = int(item["output"])
        samples_by_label[label].append(item)
    
    # 计算每个标签需要的样本数
    target_counts = {}
    target_total = total_samples  # 保持总样本数不变
    for label, prob in target_dist.items():
        target_counts[label] = int(target_total * prob)
    
    # 调整目标计数以确保总和等于目标总样本数
    count_diff = target_total - sum(target_counts.values())
    if count_diff != 0:
        # 将差额分配给样本最多的类别
        max_label = max(label_counts, key=label_counts.get)
        target_counts[max_label] += count_diff
    
    # 构建平衡后的数据集
    balanced_data = []
    for label, target_count in target_counts.items():
        available_samples = samples_by_label[label]
        
        if not available_samples:
            print(f"警告: 标签 {label} 没有可用样本")
            continue
        
        # 如果需要的样本比可用的多，进行过采样（随机重复采样）
        if target_count > len(available_samples):
            # 先添加所有可用样本
            sampled = available_samples.copy()
            # 然后随机采样剩余需要的样本
            additional_needed = target_count - len(available_samples)
            additional_samples = np.random.choice(available_samples, size=additional_needed, replace=True)
            sampled.extend(additional_samples)
        # 如果需要的样本比可用的少，进行欠采样（随机选择一部分）
        elif target_count < len(available_samples):
            sampled = list(np.random.choice(available_samples, size=target_count, replace=False))
        # 如果刚好相等，直接使用全部样本
        else:
            sampled = available_samples
        
        balanced_data.extend(sampled)
    
    # 打乱平衡后的数据
    np.random.shuffle(balanced_data)
    
    return balanced_data

def calculate_stage_metrics(true_labels, pred_labels):
    """计算各个睡眠阶段的评估指标
    
    参数:
        true_labels: 真实标签列表
        pred_labels: 预测标签列表
        
    返回:
        包含各项指标的字典
    """
    # 计算各阶段的召回率
    stage_recalls = {}
    stage_precisions = {}
    stage_f1_scores = {}
    
    # 计算混淆矩阵
    cm = confusion_matrix(true_labels, pred_labels, labels=range(6))
    
    # 打印召回率计算总结
    print("\n===== 各睡眠阶段召回率计算结果 =====")
    print("睡眠阶段\t召回率\t精确率\tF1分数\t样本数")
    print("-" * 50)
    
    # 对每个睡眠阶段计算指标
    for stage in range(6):
        # 二分类处理 - 当前阶段为正类，其他为负类
        true_binary = [1 if label == stage else 0 for label in true_labels]
        pred_binary = [1 if label == stage else 0 for label in pred_labels]
        
        # 计算该阶段的召回率(查全率)
        recall = recall_score(true_binary, pred_binary, zero_division=0)
        stage_recalls[stage] = recall
        
        # 计算该阶段的精确率(查准率)
        precision = precision_score(true_binary, pred_binary, zero_division=0)
        stage_precisions[stage] = precision
        
        # 计算该阶段的F1分数
        f1 = f1_score(true_binary, pred_binary, zero_division=0)
        stage_f1_scores[stage] = f1
        
        # 计算该阶段的样本数量
        stage_samples = true_binary.count(1)
        
        # 打印该阶段的指标
        stage_name = ["Wake (W)", "NREM Stage 1 (N1)", "NREM Stage 2 (N2)", 
                      "NREM Stage 3 (N3)", "NREM Stage 4 (N4)", "REM Sleep (R)"][stage]
        print(f"{stage} ({stage_name}):\t{recall:.4f}\t{precision:.4f}\t{f1:.4f}\t{stage_samples}")
    
    # 计算宏平均指标
    macro_recall = np.mean(list(stage_recalls.values()))
    macro_precision = np.mean(list(stage_precisions.values()))
    macro_f1 = np.mean(list(stage_f1_scores.values()))
    
    print("-" * 50)
    print(f"宏平均指标:\t{macro_recall:.4f}\t{macro_precision:.4f}\t{macro_f1:.4f}\t{len(true_labels)}")
    print("=" * 50)
    
    return {
        "stage_recalls": stage_recalls,
        "stage_precisions": stage_precisions,
        "stage_f1_scores": stage_f1_scores,
        "macro_recall": macro_recall,
        "macro_precision": macro_precision,
        "macro_f1": macro_f1,
        "confusion_matrix": cm
    }

def save_metrics_to_excel(metrics, dataset_info, output_path):
    """将评估指标保存到Excel文件
    
    参数:
        metrics: 评估指标字典
        dataset_info: 数据集信息字典
        output_path: 输出Excel文件路径
    """
    try:
        # 导入必要的模块
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "睡眠阶段指标"
        
        # 设置样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        recall_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 添加标题
        ws.append(["睡眠阶段分类数据集指标评估"])
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=14)
        
        # 添加基本信息
        ws.append(["评估时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append(["数据集名称", dataset_info.get("dataset_name", "未命名")])
        ws.append(["训练集样本数", dataset_info.get("train_samples", 0)])
        ws.append(["测试集样本数", dataset_info.get("test_samples", 0)])
        ws.append([])  # 空行
        
        # 添加各睡眠阶段分布
        ws.append(["各睡眠阶段样本分布"])
        ws['A7'].font = header_font
        ws['A7'].fill = header_fill
        
        ws.append(["睡眠阶段", "样本数", "比例"])
        for i in range(3):
            ws.cell(row=8, column=i+1).font = header_font
        
        stage_names = ["Wake (W)", "NREM Stage 1 (N1)", "NREM Stage 2 (N2)", 
                       "NREM Stage 3 (N3)", "NREM Stage 4 (N4)", "REM Sleep (R)"]
        
        for stage, name in enumerate(stage_names):
            count = dataset_info.get("stage_counts", {}).get(stage, 0)
            total = sum(dataset_info.get("stage_counts", {}).values())
            percentage = count / total * 100 if total > 0 else 0
            ws.append([name, count, f"{percentage:.2f}%"])
        
        # 添加总样本数
        ws.append(["总计", total, "100.00%"])
        
        # 添加各阶段评估指标
        ws.append([])  # 空行
        ws.append(["各睡眠阶段评估指标"])
        ws['A16'].font = header_font
        ws['A16'].fill = header_fill
        
        # 创建更详细的标题行，突出显示召回率
        ws.append(["睡眠阶段", "召回率 (Recall)", "精确率 (Precision)", "F1分数"])
        for i in range(4):
            ws.cell(row=17, column=i+1).font = header_font
        
        # 特别标记召回率列
        ws.cell(row=17, column=2).fill = recall_fill
        
        # 计算每个阶段的真实样本数
        true_counts = {}
        for label in dataset_info.get("test_true_labels", []):
            if label not in true_counts:
                true_counts[label] = 0
            true_counts[label] += 1
        
        for stage, name in enumerate(stage_names):
            recall = metrics["stage_recalls"].get(stage, 0)
            precision = metrics["stage_precisions"].get(stage, 0)
            f1 = metrics["stage_f1_scores"].get(stage, 0)
            
            # 添加一行数据
            row = ws.append([name, f"{recall:.4f}", f"{precision:.4f}", f"{f1:.4f}"])
            
            # 为召回率单元格添加高亮
            ws.cell(row=ws.max_row, column=2).fill = recall_fill
        
        # 添加宏平均指标
        ws.append(["宏平均", 
                   f"{metrics.get('macro_recall', 0):.4f}", 
                   f"{metrics.get('macro_precision', 0):.4f}", 
                   f"{metrics.get('macro_f1', 0):.4f}"])
        for i in range(4):
            ws.cell(row=24, column=i+1).font = header_font
        
        # 为宏平均召回率添加高亮
        ws.cell(row=24, column=2).fill = recall_fill
            
        # 添加单独的召回率报告部分
        ws.append([])  # 空行
        ws.append(["睡眠阶段召回率详细报告"])
        ws['A26'].font = header_font
        ws['A26'].fill = header_fill
        
        # 添加召回率报告标题
        ws.append(["睡眠阶段", "真实样本数", "召回率", "错误分类数", "注释"])
        for i in range(5):
            ws.cell(row=27, column=i+1).font = header_font
        
        # 添加每个阶段的召回率详情
        for stage, name in enumerate(stage_names):
            recall = metrics["stage_recalls"].get(stage, 0)
            true_count = true_counts.get(stage, 0)
            error_count = int(true_count * (1 - recall)) if true_count > 0 else 0
            
            # 添加注释
            note = ""
            if recall < 0.6 and true_count > 0:
                note = "需要提高"
            elif recall >= 0.8:
                note = "表现良好"
            
            ws.append([name, true_count, f"{recall:.4f}", error_count, note])
        
        # 添加混淆矩阵
        ws.append([])
        ws.append(["混淆矩阵"])
        ws['A33'].font = header_font
        ws['A33'].fill = header_fill
        
        # 添加标题行
        headers = ["真实\预测"] + [f"{name.split(' ')[0]}" for name in stage_names]
        ws.append(headers)
        for i in range(len(headers)):
            ws.cell(row=34, column=i+1).font = header_font
        
        # 添加混淆矩阵数据
        cm = metrics.get("confusion_matrix", [])
        for i, name in enumerate(stage_names):
            row_data = [name.split(' ')[0]] + [cm[i, j] if len(cm) > 0 else 0 for j in range(6)]
            ws.append(row_data)
        
        # 保存文件
        wb.save(output_path)
        print(f"指标已保存到Excel文件: {output_path}")
        return True
    except Exception as e:
        print(f"保存Excel文件出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def process_directory(input_dir, output_dirs, max_files=10, n_jobs=1, target_sfreq=100, balance_strategy="balanced", balance_alpha=0.7, weight_method="sqrt_inverse"):
    """处理整个目录中的所有EDF文件，并确保标签分布平衡
    
    参数:
        input_dir: 输入目录路径
        output_dirs: 包含训练集、测试集和全部数据集输出目录的字典，格式为 {'train': train_dir, 'test': test_dir, 'all': all_dir}
        max_files: 最大处理文件数
        n_jobs: 并行处理的进程数
        target_sfreq: 目标采样频率 (Hz)
        balance_strategy: 平衡策略
        balance_alpha: 平衡系数
        weight_method: 权重计算方法
        
    返回:
        元组: (all_data, balanced_train_data, balanced_test_data)
    """
    # 提取各个输出目录
    train_output_dir = output_dirs['train']
    test_output_dir = output_dirs['test']
    all_output_dir = output_dirs['all']
    
    # 确保各个输出目录存在
    os.makedirs(train_output_dir, exist_ok=True)
    os.makedirs(test_output_dir, exist_ok=True)
    os.makedirs(all_output_dir, exist_ok=True)
    
    # 查找所有PSG信号文件
    edf_pattern = os.path.join(input_dir, "*-PSG.edf")
    edf_files = glob.glob(edf_pattern)
    
    if not edf_files:
        print(f"在 {input_dir} 中未找到PSG EDF文件")
        return [], [], []
    
    # 取前N个文件
    edf_files = edf_files[:max_files]
    print(f"将处理 {len(edf_files)} 个PSG EDF文件")
    
    # 收集所有数据，稍后再分割
    all_data = []
    
    # 使用多进程并行处理文件
    if n_jobs > 1:
        print(f"启动并行处理，使用 {n_jobs} 个进程...")
        with ProcessPoolExecutor(max_workers=n_jobs) as executor:
            # 提交所有任务，传递目标采样率
            futures = [executor.submit(process_single_file, edf_file, target_sfreq) for edf_file in edf_files]
            
            # 收集结果
            for future in tqdm(as_completed(futures), total=len(futures), desc="处理EDF文件"):
                try:
                    result = future.result()
                    if result:
                        all_data.extend(result)
                except Exception as e:
                    print(f"处理任务时出错: {str(e)}")
    else:
        # 单进程处理每个文件
        for edf_file in tqdm(edf_files, desc="处理EDF文件"):
            # 查找对应的注释文件
            annotation_file = find_annotation_file(edf_file)
            
            if annotation_file:
                print(f"处理: {os.path.basename(edf_file)}")
                try:
                    # 处理文件并获取特征
                    raw, channel_names = read_edf_file(edf_file, target_sfreq)
                    annotations = mne.read_annotations(annotation_file)
                    features = extract_stage_windows(raw, annotations, channel_names)
                    
                    if features:
                        all_data.extend(features)
                        print(f"提取了 {len(features)} 个样本")
                    else:
                        print("未提取到有效特征")
                except Exception as e:
                    print(f"处理文件时出错: {str(e)}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"跳过 {os.path.basename(edf_file)}: 未找到匹配的注释文件")
    
    if not all_data:
        print("未收集到任何有效数据!")
        return [], [], []
    
    # 按标签分组
    stage_features = {0: [], 1: [], 2: [], 3: [], 4: [], 5: []}
    for feature in all_data:
        stage = int(feature["output"])
        stage_features[stage].append(feature)
    
    # 打印各阶段原始样本数量
    print("\n原始数据各阶段样本分布:")
    for stage in sorted(stage_features.keys()):
        print(f"阶段 {stage}: {len(stage_features[stage])} 个样本")
    
    # 分割原始数据为训练集和测试集（按9:1比例）
    np.random.seed(42)
    np.random.shuffle(all_data)
    train_split_idx = int(len(all_data) * 0.9)
    raw_train_data = all_data[:train_split_idx]
    raw_test_data = all_data[train_split_idx:]
    
    # 对训练集和测试集分别进行平衡处理
    print(f"\n使用平衡策略处理数据 (策略={balance_strategy}, 平衡系数={balance_alpha})")
    balanced_train_data = balance_dataset(raw_train_data, strategy=balance_strategy, 
                                         balance_alpha=balance_alpha, weight_method=weight_method)
    # 测试集保持原始分布
    balanced_test_data = raw_test_data
    
    # 打印平衡后的数据集信息
    print(f"平衡后 - 训练集: {len(balanced_train_data)} 个样本, 测试集: {len(balanced_test_data)} 个样本")
    
    # 转换格式为Alpaca格式
    system_prompt = "You are a neurobiological expert specializing in EEG data analysis and sleep stage classification."
    # 计算窗口长度（秒）
    window_length_sec = window_length_ms / 1000
    
    task_instruction = f"""Your task is to analyze the provided EEG data (including voltage values from the Fpz-Cz and Pz-Oz channels) and determine the current sleep stage of the volunteer based on the following classification criteria:
0: Wakefulness (W)
1: Non-rapid eye movement sleep stage 1 (N1)
2: Non-rapid eye movement sleep stage 2 (N2)
3: Non-rapid eye movement sleep stage 3 (N3)
4: Non-rapid eye movement sleep stage 4 (N4)
5: Rapid eye movement sleep stage (R)
The EEG data is provided in the format: 'channel names > data points', where each data point is formatted as 'Fpz-Cz voltage in μV, Pz-Oz voltage in μV' and separated by '>' symbols. For example: 'EEG Fpz-Cz, EEG Pz-Oz>30.12,1.11>-65.46,11.92>-13.17,33.13>'. The data spans {window_length_ms}ms ({window_length_sec} seconds) with a sampling interval that depends on the original sampling rate. In your analysis, pay attention to the following characteristics of each sleep stage:
- Wakefulness (W): High-frequency, low-amplitude waves.
- N1: Low-amplitude, mixed-frequency waves.
- N2: Sleep spindles and K-complexes.
- N3: High-amplitude, low-frequency delta waves.
- N4: Dominant delta waves.
- REM (R): REM sleep has highly distinctive and unique characteristics. It primarily presents with rapid, irregular eye movements visible in EEG as sharp, jagged waveforms. Its core feature is low-amplitude, mixed-frequency EEG activity with prominent theta waves (4-7 Hz). While somewhat similar to N1 stage, REM has distinctive saw-tooth wave patterns, which are key diagnostic markers. Unlike N2 stage, REM lacks sleep spindles and K-complexes. The EEG in REM shows a desynchronized pattern resembling wakefulness but is accompanied by complete loss of muscle tone (muscle atonia). REM may also feature rapid, irregular transient muscle twitches, along with irregular variations in heart rate and respiration. These multiple features collectively constitute the complete picture of REM sleep, making it the most distinctive and readily identifiable among all sleep stages.
Your response must be a single number (0, 1, 2, 3, 4, or 5) corresponding to the sleep stage. Do not include any additional text, punctuation, or explanations. """
    
    # 对训练集和测试集进行随机打乱
    np.random.seed(42)  
    np.random.shuffle(balanced_train_data)
    np.random.shuffle(balanced_test_data)
    
    print("\n已对数据集进行随机打乱")
    
    # 转换为Alpaca格式
    train_alpaca = [{
        "instruction": task_instruction,
        "input": item["input"],
        "output": item["output"],
        "system": item["system"]
    } for item in balanced_train_data]
    test_alpaca = [{
        "instruction": task_instruction,
        "input": item["input"],
        "output": item["output"],
        "system": item["system"]
    } for item in balanced_test_data]
    all_alpaca = train_alpaca + test_alpaca
    
    # 保存数据集
    def save_dataset(data, filename, output_directory):
        output_path = os.path.join(output_directory, filename)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)  
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"已保存 {len(data)} 个样本到 {output_path}")
    
    # 计算token数量
    print("计算token数量...")
    # 只计算一个样本的token数量作为参考
    if train_alpaca:
        token_info = calculate_tokens(train_alpaca[0])
        avg_total_tokens = token_info["total_tokens"]
        print(f"样本token统计 - 指令: {token_info['instruction_tokens']}, 输入: {token_info['input_tokens']}, 总计: {token_info['total_tokens']}")
    else:
        avg_total_tokens = 0
        print("没有训练样本可供计算token")
    
    # 使用全局变量window_length_ms
    sampling_rate = target_sfreq      
    
    # 构建文件名前缀
    base_prefix = f"edf{max_files}_{sampling_rate}hz_{window_length_ms}ms_tok{avg_total_tokens}_"
    
    # 保存平衡方式的数据集
    print("\n保存数据集...")
    # 构建平衡后缀，避免重复"balanced"单词
    # 只有当策略不是"balanced"时才包含策略名称
    strategy_part = f"{balance_strategy}_" if balance_strategy != "balanced" else ""
    balanced_suffix = f"balanced_{strategy_part}{balance_alpha}_{weight_method}"
    train_filename = f"{base_prefix}{balanced_suffix}_train.json"
    test_filename = f"{base_prefix}{balanced_suffix}_test.json"
    all_filename = f"{base_prefix}{balanced_suffix}_all_data.json"
    
    # 保存到各目录
    save_dataset(train_alpaca, train_filename, train_output_dir)
    save_dataset(test_alpaca, test_filename, test_output_dir)
    save_dataset(all_alpaca, all_filename, all_output_dir)
    
    # 更新dataset_info.json文件
    update_dataset_info(train_filename, train_output_dir)
    
    # 计算并保存评估指标
    print("\n计算数据集各睡眠阶段指标...")
    
    # 准备测试集的真实标签和预测标签（模拟预测，用于计算基线召回率）
    true_labels = [int(item["output"]) for item in balanced_test_data]
    
    # 创建一个模拟的预测标签（以最频繁的类别作为预测）
    counter = Counter(true_labels)
    most_common = counter.most_common(1)[0][0]
    baseline_pred_labels = [most_common] * len(true_labels)
    
    # 计算各睡眠阶段召回率等指标
    print("\n===== 测试集召回率基线评估 =====")
    print(f"使用多数类别 ({most_common}) 作为基线预测进行评估")
    metrics = calculate_stage_metrics(true_labels, baseline_pred_labels)
    
    # 准备数据集信息，并添加真实标签以便在Excel中计算召回率详情
    dataset_info = {
        "dataset_name": f"Sleep-EDF (max_files={max_files}, rate={target_sfreq}Hz)",
        "train_samples": len(balanced_train_data),
        "test_samples": len(balanced_test_data),
        "stage_counts": Counter([int(item["output"]) for item in all_alpaca]),
        "test_true_labels": true_labels
    }
    
    # 保存指标到Excel
    excel_path = os.path.join(all_output_dir, f"sleep_stage_metrics_{max_files}_{target_sfreq}hz.xlsx")
    save_metrics_to_excel(metrics, dataset_info, excel_path)
    
    print(f"\n处理完成。")
    print(f"总样本: {len(all_alpaca)}，训练集: {len(train_alpaca)}，测试集: {len(test_alpaca)}")
    print(f"平均token数: {avg_total_tokens}")
    print(f"阶段指标（包括召回率）已保存到: {excel_path}")
    
    # 返回所需的数据集
    return all_data, balanced_train_data, balanced_test_data

def calculate_tokens(sample):
    """计算样本的token数量"""
    global tokenizer
    
    try:
        # 懒加载tokenizer
        if tokenizer is None:
            tokenizer = AutoTokenizer.from_pretrained(tokenizer_path, trust_remote_code=True)
        
        # 构建完整提示
        system = sample.get('system', '')
        instruction = sample.get('instruction', '')
        input_text = sample.get('input', '')
        
        # 计算指令部分的token数量
        instruction_token_count = len(tokenizer.encode(instruction))
        
        # 计算输入部分的token数量
        input_token_count = len(tokenizer.encode(input_text))
        
        # 计算总token数量（包括系统提示）
        if system:
            messages = [
                {"role": "system", "content": system},
                {"role": "user", "content": instruction + "\n" + input_text}
            ]
            prompt_text = tokenizer.apply_chat_template(messages, tokenize=False)
        else:
            prompt_text = instruction + "\n" + input_text
        
        total_token_count = len(tokenizer.encode(prompt_text))
        
        return {
            "instruction_tokens": instruction_token_count,
            "input_tokens": input_token_count,
            "total_tokens": total_token_count
        }
    except Exception as e:
        print(f"计算token数量时出错: {str(e)}")
        # 出错时返回默认值
        return {
            "instruction_tokens": 1000,
            "input_tokens": 2000,
            "total_tokens": 3000
        }

def update_dataset_info(train_filename, output_dir):
    """更新dataset_info.json文件中的信息，增强错误处理能力"""
    dataset_info_path = "/data/lhc/projects/LLaMA-Factory/data/dataset_info.json"
    backup_path = dataset_info_path + ".bak"
    
    # 备份原始文件
    try:
        shutil.copy2(dataset_info_path, backup_path)
        print(f"已创建备份文件: {backup_path}")
    except Exception as e:
        print(f"创建备份文件失败: {str(e)}")
    
    # 创建新的dataset_info数据
    train_key = train_filename.replace('.json', '')
    
    # 提取token数量
    token_count = 0
    match = re.search(r'tok(\d+)_', train_filename)
    if match:
        token_count = int(match.group(1))
    
    # 读取训练文件获取样本数量
    train_file_path = os.path.join(output_dir, train_filename)
    sample_count = 0
    try:
        if os.path.exists(train_file_path):
            with open(train_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                sample_count = len(data)
        else:
            print(f"警告：找不到训练文件 {train_file_path} 来计算样本数量")
    except Exception as e:
        print(f"读取训练文件获取样本数量时出错: {str(e)}")
    
    # 确定相对路径
    relative_path = os.path.relpath(output_dir, "/data/lhc/projects/LLaMA-Factory/data")
    
    # 新数据集信息
    new_info = {
        "script_url": relative_path,
        "file_name": train_filename,
        "columns": {
            "prompt": "instruction",
            "query": "input",
            "response": "output",
            "system": "system"
        },
        "token_count": token_count,
        "sample_count": sample_count
    }
    
    # 尝试读取并更新现有文件
    try:
        # 读取文件
        with open(dataset_info_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # 尝试解析JSON
        try:
            dataset_info = json.loads(content)
            # 更新数据
            dataset_info[train_key] = new_info
            # 写入更新后的文件
            with open(dataset_info_path, 'w', encoding='utf-8') as f:
                json.dump(dataset_info, f, ensure_ascii=False, indent=2)
            print(f"更新dataset_info.json文件成功，添加了 {train_key} 数据集")
            return
        except json.JSONDecodeError:
            print("JSON解析失败，将创建新文件")
    except Exception as e:
        print(f"读取文件失败: {str(e)}")
    
    # 如果上述步骤失败，创建新文件
    try:
        dataset_info = {train_key: new_info}
        with open(dataset_info_path, 'w', encoding='utf-8') as f:
            json.dump(dataset_info, f, ensure_ascii=False, indent=2)
        print(f"已创建新的dataset_info.json文件，包含 {train_key} 数据集")
    except Exception as e:
        print(f"创建新文件失败: {str(e)}")

# 全局变量
window_length_ms = 7500  

def print_stage_statistics(all_data, train_data, test_data):
    """打印各个睡眠阶段的详细样本统计信息
    
    参数:
        all_data: 所有收集到的数据
        train_data: 平衡训练集
        test_data: 测试集
    """
    # 定义睡眠阶段名称映射
    stage_names = {
        0: "清醒 (W)",
        1: "非快速眼动睡眠阶段1 (N1)",
        2: "非快速眼动睡眠阶段2 (N2)",
        3: "非快速眼动睡眠阶段3 (N3)",
        4: "非快速眼动睡眠阶段4 (N4)",
        5: "快速眼动睡眠 (REM)"
    }
    
    # 统计各数据集中各阶段的样本数
    def count_stages(data):
        stage_counts = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0}
        for item in data:
            stage = int(item["output"])
            stage_counts[stage] = stage_counts.get(stage, 0) + 1
        return stage_counts
    
    # 计算各数据集的阶段分布
    all_counts = count_stages(all_data)
    train_counts = count_stages(train_data)
    test_counts = count_stages(test_data)
    
    # 打印表头
    print("\n" + "="*100)
    print("各睡眠阶段样本数量统计报告")
    print("="*100)
    
    # 打印原始数据分布
    print("\n【原始数据】")
    print(f"{'睡眠阶段':^25} | {'样本数':^10} | {'占比':^10}")
    print("-"*50)
    total_samples = sum(all_counts.values())
    for stage in sorted(all_counts.keys()):
        count = all_counts[stage]
        percentage = (count / total_samples * 100) if total_samples > 0 else 0
        print(f"{stage}: {stage_names[stage]:20} | {count:^10} | {percentage:^8.2f}%")
    print("-"*50)
    print(f"{'总计':^25} | {total_samples:^10} | {'100.00':^8}%")
    
    # 打印平衡后的分布
    print("\n【平衡数据集分布】")
    print(f"{'睡眠阶段':^25} | {'训练集':^10} | {'训练集占比':^10} | {'测试集':^10} | {'测试集占比':^10} | {'总计':^10}")
    print("-"*85)
    for stage in sorted(train_counts.keys()):
        train_count = train_counts[stage]
        test_count = test_counts[stage]
        total_count = train_count + test_count
        
        train_total = sum(train_counts.values())
        test_total = sum(test_counts.values())
        
        train_percentage = (train_count / train_total * 100) if train_total > 0 else 0
        test_percentage = (test_count / test_total * 100) if test_total > 0 else 0
        
        print(f"{stage}: {stage_names[stage]:20} | {train_count:^10} | {train_percentage:^10.2f}% | {test_count:^10} | {test_percentage:^10.2f}% | {total_count:^10}")
    
    print("-"*85)
    train_total = sum(train_counts.values())
    test_total = sum(test_counts.values())
    print(f"{'总计':^25} | {train_total:^10} | {'100.00':^10}% | {test_total:^10} | {'100.00':^10}% | {train_total+test_total:^10}")
    
    # 阶段间样本比例分析
    print("\n【数据集平衡性分析】")
    
    # 原始数据
    max_stage_orig = max(all_counts.items(), key=lambda x: x[1])
    min_stage_orig = min(all_counts.items(), key=lambda x: x[1])
    imbalance_ratio_orig = max_stage_orig[1] / min_stage_orig[1] if min_stage_orig[1] > 0 else float('inf')
    
    # 平衡训练集
    max_stage_bal = max(train_counts.items(), key=lambda x: x[1])
    min_stage_bal = min(train_counts.items(), key=lambda x: x[1])
    imbalance_ratio_bal = max_stage_bal[1] / min_stage_bal[1] if min_stage_bal[1] > 0 else float('inf')
    
    print(f"原始数据最多/最少阶段比例: {imbalance_ratio_orig:.2f} ({max_stage_orig[0]}阶段:{max_stage_orig[1]}样本 vs {min_stage_orig[0]}阶段:{min_stage_orig[1]}样本)")
    print(f"平衡后训练集最多/最少阶段比例: {imbalance_ratio_bal:.2f} ({max_stage_bal[0]}阶段:{max_stage_bal[1]}样本 vs {min_stage_bal[0]}阶段:{min_stage_bal[1]}样本)")
    
    print("="*100)

def main():
    # 设置
    base_dir = '/data/lhc/datasets/sleep-edfx'
    
    # 修改输出目录路径
    train_output_dir = '/data/lhc/datasets_new/sleep/train'
    test_output_dir = '/data/lhc/datasets_new/sleep/test'
    all_output_dir = '/data/lhc/datasets_new/sleep/all'
    
    # 确保输出目录存在
    os.makedirs(train_output_dir, exist_ok=True)
    os.makedirs(test_output_dir, exist_ok=True)
    os.makedirs(all_output_dir, exist_ok=True)
    
    # 最大文件数 - 减少文件数量，降低内存需求
    max_files = 197
    
    # 采样率选择 (100Hz 或 200Hz)
    target_sampling_rate = 200  # 可以设置为100或200
    
    # 平衡数据集参数
    balance_strategy = "balanced"  # balanced 或 original
    balance_alpha = 0.1              # 平衡系数，0表示完全均衡，1表示保持原始分布
    weight_method = "sqrt_inverse" # inverse, sqrt_inverse, 或 log_inverse
    
    # CPU核心数 - 减少并行进程，降低内存压力
    n_jobs = max(1, int(multiprocessing.cpu_count() * 0.75))  # 从0.75减少到0.5
    print(f"使用 {n_jobs} 个CPU核心")
    
    # 处理文件
    print(f"处理文件，目标采样率: {target_sampling_rate}Hz...")
    print(f"平衡策略: {balance_strategy}, 平衡系数: {balance_alpha}, 权重方法: {weight_method}")
    
    # 处理目录，只接收新方法的返回值
    all_data, train_data, test_data = process_directory(
        base_dir, 
        {'train': train_output_dir, 'test': test_output_dir, 'all': all_output_dir}, 
        max_files, 
        n_jobs=n_jobs, 
        target_sfreq=target_sampling_rate,
        balance_strategy=balance_strategy, 
        balance_alpha=balance_alpha,
        weight_method=weight_method
    )
    
    # 打印各阶段样本数量统计
    print_stage_statistics(all_data, train_data, test_data)

if __name__ == "__main__":
    main()