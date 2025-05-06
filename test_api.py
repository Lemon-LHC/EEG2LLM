import os
import json
import time
import numpy as np
import pandas as pd
import traceback
import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm
from openai import OpenAI
from transformers.utils.versions import require_version
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
import re

require_version("openai>=1.5.0", "To fix: pip install openai>=1.5.0")

# 定义睡眠阶段标签，避免重复硬编码
SLEEP_STAGE_LABELS = [
    'Wake (W)', 
    'NREM Stage 1 (N1)', 
    'NREM Stage 2 (N2)', 
    'NREM Stage 3 (N3)', 
    'NREM Stage 4 (N4)', 
    'REM Sleep (R)'
]


def load_test_data(file_path, max_samples=None):
    """加载测试数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        test_data = json.load(f)
    
    # 如果设置了最大样本数，随机选择样本
    if max_samples and max_samples < len(test_data):
        indices = np.random.choice(len(test_data), max_samples, replace=False)
        test_data = [test_data[i] for i in indices]
        print(f"随机选择了 {len(test_data)} 条数据用于测试")

    return test_data


def get_api_prediction(client, messages, max_retries=5, retry_delay=2, max_tokens=60000):
    """通过API获取模型预测，包含重试机制

    Args:
        client: OpenAI 客户端
        messages: 发送给 API 的消息列表
        max_retries: 最大重试次数
        retry_delay: 重试间隔时间 (秒)
        max_tokens: API 请求的最大 token 数

    Returns:
        tuple: (response_content, prediction, inference_time)
               prediction 为提取到的数字 (0-5)，如果提取失败则为 None
               inference_time 为推理时间 (秒)
    """
    for attempt in range(max_retries):
        try:
            # 调用API
            start_time = time.time()
            result = client.chat.completions.create(
                messages=messages,
                model="test",
                temperature=0.1,
                max_tokens=max_tokens,
                timeout=30  # 设置超时时间为30秒
            )
            end_time = time.time()

            # 获取响应
            response_content = result.choices[0].message.content

            # 提取数字 (0-5)
            prediction = None
            match = re.search(r'\\b([0-5])\\b', response_content)
            if match:
                prediction = int(match.group(1))
            else:
                for char in response_content:
                    if char.isdigit() and int(char) in [0, 1, 2, 3, 4, 5]:
                        prediction = int(char)
                        break

            return response_content, prediction, end_time - start_time
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"API调用出错 (尝试 {attempt+1}/{max_retries}): {e}，等待 {retry_delay} 秒后重试...")
                time.sleep(retry_delay)
            else:
                print(f"API调用失败，已达到最大重试次数 ({max_retries}): {e}")
                return f"API调用失败: {e}", None, 0


def evaluate_model(client, test_data, print_interval=10, save_interval=500, save_dir='/data/lhc/results',
                model_name="unknown_model", test_set_name="unknown_dataset", max_tokens=60000):
    """评估模型性能
    (修改版：使用 system, instruction, input 构建 messages)

    Args:
        client: OpenAI客户端
        test_data: 测试数据 (list of dicts)
        print_interval: 打印中间结果的间隔
        save_interval: 保存中间结果的间隔
        save_dir: 保存结果的目录
        model_name: 模型名称
        test_set_name: 测试集名称
        max_tokens: API 请求的最大 token 数
    """
    # 创建一个包含模型名称和测试集名称的文件夹来保存结果
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    result_dir = os.path.join(save_dir, f"{model_name}_{test_set_name}_{timestamp}")
    os.makedirs(result_dir, exist_ok=True)

    # 保存中间结果的函数
    def save_intermediate_results(current_idx, current_metrics, current_results):
        """保存中间结果
           (需要更新以匹配新的 metrics 结构)
        """
        intermediate_dir = os.path.join(result_dir, f"intermediate_{current_idx}")
        os.makedirs(intermediate_dir, exist_ok=True)

        # 保存模型和测试集信息 (使用 metrics 中的信息)
        with open(os.path.join(intermediate_dir, "info.json"), "w") as f:
            json.dump({
                "model_name": model_name,
                "test_set_name": test_set_name,
                "samples_processed": current_idx,
                "total_samples_in_metrics": current_metrics.get("total_samples", "N/A"),
                "valid_predictions_in_metrics": current_metrics.get("valid_predictions", "N/A"),
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }, f, indent=4)

        # 保存指标 (确保 NumPy 数组能被 JSON 序列化)
        with open(os.path.join(intermediate_dir, "metrics.json"), "w", encoding='utf-8') as f:
            serializable_metrics = {k: v.tolist() if isinstance(v, np.ndarray) else v
                                     for k, v in current_metrics.items()}
            json.dump(serializable_metrics, f, indent=4, ensure_ascii=False)

        # 保存详细结果列表
        with open(os.path.join(intermediate_dir, "results.json"), "w", encoding='utf-8') as f:
            json.dump(current_results, f, indent=4, ensure_ascii=False)

        # 保存混淆矩阵图
        if 'confusion_matrix' in current_metrics and current_metrics['confusion_matrix'] is not None:
             cm_data = np.array(current_metrics['confusion_matrix'])
             cm_labels = current_metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])
             if cm_data.ndim == 2 and cm_data.shape == (len(cm_labels), len(cm_labels)):
                 try:
                     plt.figure(figsize=(10, 8))
                     sns.heatmap(cm_data, annot=True, fmt='d', cmap='Blues',
                                 xticklabels=cm_labels,
                                 yticklabels=cm_labels)
                     plt.xlabel('Predicted Label')
                     plt.ylabel('True Label')
                     plt.title(f'Intermediate Confusion Matrix (Samples Processed: {current_idx})')
                     plt.savefig(os.path.join(intermediate_dir, "confusion_matrix.png"))
                     plt.close()
                 except Exception as plot_err:
                     print(f"警告：绘制中间混淆矩阵图时出错: {plot_err}")
                     plt.close()
             else:
                 print(f"警告：中间混淆矩阵数据格式不正确，跳过绘图。Shape: {cm_data.shape}, Labels: {len(cm_labels)}")
        else:
             print("警告：中间指标中缺少混淆矩阵，跳过绘图。")

        # 保存Excel文件 (需要更新以处理新 metrics)
        save_results_to_excel(current_results, current_metrics, current_results, intermediate_dir, model_name, test_set_name)

        print(f"\n中间结果已保存至: {intermediate_dir}")

    print("开始评估模型性能...")
    results = []
    true_labels = []
    pred_labels = []
    responses = []

    # 打印每阶段标签名
    print("阶段顺序:", SLEEP_STAGE_LABELS)

    # 使用tqdm创建进度条
    pbar = tqdm(test_data, desc="Test Samples", ncols=100)

    # 用于实时显示性能指标
    correct_count = 0
    processed_count = 0
    total_time = 0.0

    for idx, data in enumerate(pbar):
        # 获取真实标签
        try:
            true_label = int(float(data["output"]))
            if not (0 <= true_label <= 5):
                 raise ValueError(f"True label {true_label} is outside the expected range 0-5.")
        except (KeyError, ValueError, TypeError) as e:
            error_msg = f"错误：样本 {idx} 缺少有效 'output' 字段或值无效: {e}"
            print(f"\n{error_msg}")
            results.append({
                "id": idx, "true_label": -99, "prediction": -2, "response": error_msg,
                "correct": False, "inference_time": 0
            })
            true_labels.append(-99)
            pred_labels.append(-2)
            responses.append(error_msg)
            continue

        try:
            # 从 data 构建 messages
            system_content = data.get("system", "")
            instruction_content = data.get("instruction", "")
            input_content = data.get("input", "")

            # Ensure we have content for the user role
            if not instruction_content and not input_content:
                 raise ValueError("样本必须包含 'instruction' 或 'input' 字段才能构建用户消息")

            # 构建 user content - prefer instruction if available, otherwise use input
            user_content = f"{instruction_content}\\n\\n{input_content}" if instruction_content and input_content else (instruction_content or input_content)

            # 构建 messages 列表
            messages = []
            if system_content:
                messages.append({"role": "system", "content": system_content})
            if user_content:
                messages.append({"role": "user", "content": user_content})
            else:
                # 此处检查理论上是多余的，因为前面的检查已经保证了 user_content 不会为空
                raise ValueError("无法构建用户消息，instruction 和 input 都为空")

            # 获取预测
            response, prediction, infer_time = get_api_prediction(client, messages, max_tokens=max_tokens)

            total_time += infer_time
            processed_count += 1

            # 处理预测结果 (prediction 是提取到的数字 0-5 或 None)
            current_prediction_label = -1
            correct = False
            if prediction is not None:
                current_prediction_label = prediction
                correct = (current_prediction_label == true_label)
                if correct:
                    correct_count += 1

            # 更新进度条状态 (基于已处理的样本)
            current_acc = correct_count / processed_count if processed_count > 0 else 0
            avg_time = total_time / processed_count if processed_count > 0 else 0

            pbar.set_postfix({
                'acc': f'{current_acc:.2%}',
                'avg_time': f'{avg_time:.2f}s',
                '预测': f'{current_prediction_label}',
                '真实': f'{true_label}',
                '正确': correct
            })

            result = {
                "id": idx,
                "true_label": true_label,
                "prediction": current_prediction_label,
                "response": response,
                "correct": correct,
                "inference_time": infer_time
            }

            results.append(result)
            true_labels.append(true_label)
            pred_labels.append(current_prediction_label)
            responses.append(response)

            # 每隔print_interval次测试打印一次相关的测试数据
            if (idx + 1) % print_interval == 0 or (idx + 1) == len(test_data):
                # 计算当前指标 (传入包含 -1 的 pred_labels)
                current_calculated_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                print(f"\n已处理 {idx + 1} 个样本 (其中 {processed_count} 个成功调用 API)")
                print(f"有效预测数: {current_calculated_metrics['valid_predictions']}")
                print(f"提取失败数 (-1): {current_calculated_metrics['unknown_predictions']}")
                print(f"处理错误数 (-2): {current_calculated_metrics['error_predictions']}")
                print(f"当前总体准确率 (基于有效预测): {current_calculated_metrics['accuracy']:.4f}")
                print(f"当前宏平均精确率: {current_calculated_metrics['precision_macro']:.4f}")
                print(f"当前宏平均召回率: {current_calculated_metrics['recall_macro']:.4f}")
                print(f"当前宏平均F1分数: {current_calculated_metrics['f1_macro']:.4f}")
                print(f"当前平均推理时间 (所有样本): {current_calculated_metrics['avg_time_per_sample']:.4f} 秒/样本")

                print("各阶段准确率 (基于有效预测):")
                class_accuracies = current_calculated_metrics.get('class_accuracies', {})
                for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
                    acc_val = class_accuracies.get(class_idx, 0.0)
                    print(f"  {label_name}: {acc_val:.4f}")

                print("各阶段F1分数 (基于有效预测):")
                class_f1s_metrics = current_calculated_metrics.get('class_metrics', {})
                short_labels = current_calculated_metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])
                for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
                     if class_idx < len(short_labels):
                         short_label = short_labels[class_idx]
                         f1_val = class_f1s_metrics.get(short_label, {}).get('f1', 0.0)
                         print(f"  {label_name}: {f1_val:.4f}")
                     else:
                         print(f"  {label_name}: N/A (标签索引超出范围)")

                print("混淆矩阵 (基于有效预测):")
                cm_data = current_calculated_metrics.get('confusion_matrix')
                cm_labels = current_calculated_metrics.get('confusion_matrix_labels')
                if cm_data is not None and cm_labels:
                     cm_data_np = np.array(cm_data)
                     if cm_data_np.shape == (len(cm_labels), len(cm_labels)):
                         print("    " + " ".join(f"{label:^5s}" for label in cm_labels))
                         for i, label in enumerate(cm_labels):
                             if i < cm_data_np.shape[0]:
                                 print(f"{label:^5s}" + " ".join(f"{cm_data_np[i, j]:^5d}" for j in range(len(cm_labels))))
                             else:
                                 print(f"{label:^5s} (无数据)")
                     else:
                         print(f"  (混淆矩阵维度 {cm_data_np.shape} 与标签数 {len(cm_labels)} 不匹配)")
                else:
                    print("  (无有效预测或标签，无法打印混淆矩阵)")

            # 每隔save_interval个样本保存一次中间结果
            if (idx + 1) % save_interval == 0:
                # 计算用于保存的中间指标
                intermediate_metrics = calculate_metrics(true_labels, pred_labels, total_time)
                # 传递计算出的指标和当前的详细结果列表
                save_intermediate_results(idx + 1, intermediate_metrics, results)

        except Exception as e:
            error_msg = f"处理样本 {idx} 时发生意外错误: {str(e)}"
            print(f"\n{error_msg}")
            print(traceback.format_exc())
            results.append({
                "id": idx, "true_label": true_label,
                "prediction": -2,
                "response": error_msg,
                "correct": False, "inference_time": 0
            })
            true_labels.append(true_label)
            pred_labels.append(-2)
            responses.append(error_msg)

    # 计算最终指标 (传入包含 -1 和 -2 的 pred_labels)
    final_metrics = calculate_metrics(true_labels, pred_labels, total_time)

    # 打印最终结果
    print("\n" + "=" * 50)
    print("睡眠分期测试结果")
    print("=" * 50)
    print(f"总样本数: {final_metrics.get('total_samples', 'N/A')}")
    print(f"有效预测样本数 (用于指标): {final_metrics.get('valid_predictions', 'N/A')}")
    print(f"提取失败样本数 (-1): {final_metrics.get('unknown_predictions', 'N/A')} ({final_metrics.get('unknown_ratio', 0.0):.2%})")
    print(f"处理错误样本数 (-2): {final_metrics.get('error_predictions', 'N/A')} ({final_metrics.get('error_ratio', 0.0):.2%})")
    print(f"总准确率 (基于有效预测): {final_metrics.get('accuracy', 0.0):.2%}")
    print(f"宏平均精确率 (基于有效预测): {final_metrics.get('precision_macro', 0.0):.2%}")
    print(f"宏平均召回率 (基于有效预测): {final_metrics.get('recall_macro', 0.0):.2%}")
    print(f"宏平均F1分数 (基于有效预测): {final_metrics.get('f1_macro', 0.0):.2%}")
    print(f"平均推理时间 (所有样本): {final_metrics.get('avg_time_per_sample', 0.0):.4f}秒/样本")

    # 打印每个标签的准确率
    print("\n各睡眠阶段准确率 (基于有效预测):")
    class_accuracies_final = final_metrics.get('class_accuracies', {})
    for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
        acc = class_accuracies_final.get(class_idx, 0.0)
        print(f"  {label_name}: {acc:.2%}")

    # 打印每个标签的 F1 分数
    print("\n各睡眠阶段F1分数 (基于有效预测):")
    class_metrics_final = final_metrics.get('class_metrics', {})
    short_labels_final = final_metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])
    for class_idx, label_name in enumerate(SLEEP_STAGE_LABELS):
         if class_idx < len(short_labels_final):
             short_label = short_labels_final[class_idx]
             f1 = class_metrics_final.get(short_label, {}).get('f1', 0.0)
             print(f"  {label_name}: {f1:.2%}")
         else:
             print(f"  {label_name}: N/A (标签索引超出范围)")

    # 打印混淆矩阵
    print("\n混淆矩阵 (基于有效预测):")
    cm_data_final = final_metrics.get('confusion_matrix')
    cm_labels_final = final_metrics.get('confusion_matrix_labels')
    if cm_data_final is not None and cm_labels_final:
         cm_data_final_np = np.array(cm_data_final)
         if cm_data_final_np.shape == (len(cm_labels_final), len(cm_labels_final)):
             print("    " + " ".join(f"{label:^5s}" for label in cm_labels_final))
             for i, label in enumerate(cm_labels_final):
                  if i < cm_data_final_np.shape[0]:
                      print(f"{label:^5s}" + " ".join(f"{cm_data_final_np[i, j]:^5d}" for j in range(len(cm_labels_final))))
                  else:
                      print(f"{label:^5s} (无数据)")
         else:
             print(f"  (最终混淆矩阵维度 {cm_data_final_np.shape} 与标签数 {len(cm_labels_final)} 不匹配)")
    else:
        print("  (无有效预测或标签，无法打印最终混淆矩阵)")

    print("=" * 50)

    # 保存结果到指定目录 (使用最终的 metrics 和 results)
    save_results(results, final_metrics, test_data, result_dir, model_name=model_name, test_set_name=test_set_name)

    return final_metrics


def calculate_metrics(true_labels, pred_labels, total_time):
    """计算评估指标
       (修改版：处理 -1 和 -2 预测)
    """
    # 过滤掉无效预测 (-1: unknown, -2: error)
    valid_indices = [i for i, pred in enumerate(pred_labels) if pred >= 0]
    unknown_predictions = sum(1 for pred in pred_labels if pred == -1)
    error_predictions = sum(1 for pred in pred_labels if pred == -2)
    num_total_samples = len(true_labels)
    num_classes = 6 # Sleep stages 0-5

    short_labels = ['W', 'N1', 'N2', 'N3', 'N4', 'R'] # Corresponds to 0-5

    # 处理没有有效预测的情况
    if not valid_indices:
        print("警告: 没有有效的预测可用于计算指标。")
        class_metrics_empty = {label: {'precision': 0.0, 'recall': 0.0, 'f1': 0.0} for label in short_labels}
        return {
            'accuracy': 0.0,
            'precision_macro': 0.0,
            'recall_macro': 0.0,
            'f1_macro': 0.0,
            'confusion_matrix': np.zeros((num_classes, num_classes), dtype=int).tolist(), # Use list for JSON
            'class_accuracies': {i: 0.0 for i in range(num_classes)},
            'class_metrics': class_metrics_empty,
            'confusion_matrix_labels': short_labels,
            'total_samples': num_total_samples,
            'valid_predictions': 0,
            'unknown_predictions': unknown_predictions,
            'error_predictions': error_predictions,
            'unknown_ratio': unknown_predictions / num_total_samples if num_total_samples > 0 else 0.0,
            'error_ratio': error_predictions / num_total_samples if num_total_samples > 0 else 0.0,
            'total_time': total_time,
            'avg_time_per_sample': total_time / num_total_samples if num_total_samples > 0 else 0.0
        }

    # 提取有效标签
    valid_true = [true_labels[i] for i in valid_indices if true_labels[i] >= 0]
    valid_pred_indices = [i for i in valid_indices if true_labels[i] >= 0]
    valid_pred = [pred_labels[i] for i in valid_pred_indices]

    # Recalculate valid count based on those with valid true labels
    num_valid_samples_for_metrics = len(valid_true)

    if num_valid_samples_for_metrics == 0:
        print("警告: 过滤后没有有效的真实标签和预测对可用于计算指标。")
        # Return empty/zeroed metrics similar to the 'not valid_indices' case
        class_metrics_empty = {label: {'precision': 0.0, 'recall': 0.0, 'f1': 0.0} for label in short_labels}
        return {
            'accuracy': 0.0, 'precision_macro': 0.0, 'recall_macro': 0.0, 'f1_macro': 0.0,
            'confusion_matrix': np.zeros((num_classes, num_classes), dtype=int).tolist(),
            'class_accuracies': {i: 0.0 for i in range(num_classes)},
            'class_metrics': class_metrics_empty, 'confusion_matrix_labels': short_labels,
            'total_samples': num_total_samples, 'valid_predictions': num_valid_samples_for_metrics,
            'unknown_predictions': unknown_predictions, 'error_predictions': error_predictions,
            'unknown_ratio': unknown_predictions / num_total_samples if num_total_samples > 0 else 0.0,
            'error_ratio': error_predictions / num_total_samples if num_total_samples > 0 else 0.0,
            'total_time': total_time,
            'avg_time_per_sample': total_time / num_total_samples if num_total_samples > 0 else 0.0
        }

    # --- 计算指标 (基于 valid_true 和 valid_pred) ---
    all_possible_labels = list(range(num_classes)) # 0 to 5

    accuracy = accuracy_score(valid_true, valid_pred)
    precision_macro = precision_score(valid_true, valid_pred, labels=all_possible_labels, average='macro', zero_division=0)
    recall_macro = recall_score(valid_true, valid_pred, labels=all_possible_labels, average='macro', zero_division=0)
    f1_macro = f1_score(valid_true, valid_pred, labels=all_possible_labels, average='macro', zero_division=0)
    # Ensure confusion matrix uses valid pairs
    confusion_matrix_result = confusion_matrix(valid_true, valid_pred, labels=all_possible_labels)

    # 计算每个类别的准确率 (基于混淆矩阵)
    class_accuracies = {}
    for i in all_possible_labels:
        # Ensure indices are within bounds if confusion_matrix_result isn't 6x6 (should be due to labels=)
        if i < confusion_matrix_result.shape[0] and i < confusion_matrix_result.shape[1]:
            tp = confusion_matrix_result[i, i]
            total_actual = confusion_matrix_result[i, :].sum()
            class_accuracies[i] = tp / total_actual if total_actual > 0 else 0.0
        else:
            class_accuracies[i] = 0.0 # Should not happen with labels=all_possible_labels

    # 计算每个类别的 Precision, Recall, F1
    precision_per_class = precision_score(valid_true, valid_pred, labels=all_possible_labels, average=None, zero_division=0)
    recall_per_class = recall_score(valid_true, valid_pred, labels=all_possible_labels, average=None, zero_division=0)
    f1_per_class = f1_score(valid_true, valid_pred, labels=all_possible_labels, average=None, zero_division=0)

    class_metrics = {}
    for i, label_name in enumerate(short_labels):
        # Ensure index i is valid for the score arrays
        if i < len(precision_per_class):
            class_metrics[label_name] = {
                'precision': float(precision_per_class[i]),
                'recall': float(recall_per_class[i]),
                'f1': float(f1_per_class[i])
            }
        else: # Should not happen if labels used correctly
             class_metrics[label_name] = {'precision': 0.0, 'recall': 0.0, 'f1': 0.0}

    return {
        'accuracy': accuracy,
        'precision_macro': precision_macro,
        'recall_macro': recall_macro,
        'f1_macro': f1_macro,
        'confusion_matrix': confusion_matrix_result.tolist(), # Use list for JSON
        'class_accuracies': class_accuracies,
        'class_metrics': class_metrics, # Contains P, R, F1 per class
        'confusion_matrix_labels': short_labels,
        'total_samples': num_total_samples,
        'valid_predictions': num_valid_samples_for_metrics, # Based on valid true/pred pairs
        'unknown_predictions': unknown_predictions,
        'error_predictions': error_predictions,
        'unknown_ratio': unknown_predictions / num_total_samples if num_total_samples > 0 else 0.0,
        'error_ratio': error_predictions / num_total_samples if num_total_samples > 0 else 0.0,
        'total_time': total_time,
        'avg_time_per_sample': total_time / num_total_samples if num_total_samples > 0 else 0.0 # Avg time over all attempted samples
    }


def save_results(results, metrics, test_data, result_dir, model_name="unknown_model", test_set_name="unknown_dataset"):
    """保存测试结果到指定目录 (特定运行的子目录)
       (更新版：使用新的 metrics 结构)
    """
    os.makedirs(result_dir, exist_ok=True)
    timestamp_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # 定义格式化指标的辅助函数
    def format_metric(value, fmt=".4f"):
        if value is None or value == 'N/A': return 'N/A'
        try: 
            return f"{float(value):{fmt}}"
        except (ValueError, TypeError): 
            return 'N/A'

    # 保存性能指标为JSON (使用新 metrics 结构)
    serializable_metrics = {k: v.tolist() if isinstance(v, np.ndarray) else v
                             for k, v in metrics.items()}
    metrics_path = os.path.join(result_dir, 'metrics.json')
    try:
        with open(metrics_path, 'w', encoding='utf-8') as f:
            json.dump(serializable_metrics, f, indent=4, ensure_ascii=False)
        print(f"Performance metrics saved to: {metrics_path}")
    except Exception as e:
        print(f"Error saving metrics JSON: {e}")
        print(traceback.format_exc())

    # 保存混淆矩阵为图片 (使用新 metrics 结构)
    if 'confusion_matrix' in metrics and metrics['confusion_matrix'] is not None:
         cm_data = np.array(metrics['confusion_matrix'])
         cm_labels = metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])
         if cm_data.ndim == 2 and cm_data.shape == (len(cm_labels), len(cm_labels)):
             try:
                 plt.figure(figsize=(10, 8))
                 sns.heatmap(cm_data, annot=True, fmt='d', cmap='Blues',
                             xticklabels=cm_labels, yticklabels=cm_labels)
                 plt.xlabel('Predicted Label')
                 plt.ylabel('True Label')
                 plt.title('Sleep Stage Classification Confusion Matrix')
                 plt.tight_layout()
                 cm_path = os.path.join(result_dir, 'confusion_matrix.png')
                 plt.savefig(cm_path)
                 plt.close()
                 print(f"混淆矩阵图已保存至: {cm_path}")
             except Exception as e:
                 print(f"Error saving confusion matrix plot: {e}")
                 print(traceback.format_exc())
                 plt.close()
         else:
             print(f"最终混淆矩阵数据格式不正确，跳过绘图。Shape: {cm_data.shape}, Labels: {len(cm_labels)}")
    else:
         print("最终指标中缺少混淆矩阵数据，跳过绘图。")

    # 保存测试样本分布统计 (基于 results 列表中的有效真实标签)
    class_counts = {}
    valid_true_labels_in_results = [r['true_label'] for r in results if r.get('true_label', -99) >= 0]
    total_valid_true_samples = len(valid_true_labels_in_results)

    for label in valid_true_labels_in_results:
        class_counts[label] = class_counts.get(label, 0) + 1

    # 绘制样本分布饼图
    try:
        plt.figure(figsize=(10, 8))
        all_possible_labels = list(range(6)) # 0-5
        sizes = [class_counts.get(i, 0) for i in all_possible_labels]
        pie_labels = [f"{SLEEP_STAGE_LABELS[i]} ({sizes[i]})" for i in all_possible_labels]

        filtered_sizes = [s for s in sizes if s > 0]
        filtered_labels = [pie_labels[i] for i, s in enumerate(sizes) if s > 0]

        if filtered_sizes:
             plt.pie(filtered_sizes, labels=filtered_labels, autopct='%1.1f%%', startangle=90)
             plt.axis('equal')
             plt.title(f'Test Sample Distribution (Valid True Labels: {total_valid_true_samples})')
             dist_path = os.path.join(result_dir, 'sample_distribution.png')
             plt.savefig(dist_path)
             plt.close()
             print(f"Sample distribution chart saved to: {dist_path}")
        else:
             print("没有有效的真实标签样本，无法生成分布图。")
             plt.close()
    except Exception as e:
        print(f"Error saving sample distribution plot: {e}")
        print(traceback.format_exc())
        plt.close()

    # 保存性能指标摘要为文本文件 (使用新 metrics 结构)
    summary_path = os.path.join(result_dir, 'summary.txt')
    try:
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("=" * 65 + "\n")
            f.write("Sleep Stage Classification Test Results Summary\n")
            f.write("=" * 65 + "\n\n")
            f.write(f"Test Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Model Name: {model_name}\n")
            f.write(f"Test Set Name: {test_set_name}\n\n")

            f.write("Dataset Statistics:\n")
            f.write(f"Total Samples Attempted: {metrics.get('total_samples', 'N/A')}\n")
            f.write(f"Valid Prediction Samples (for metrics): {metrics.get('valid_predictions', 'N/A')}\n")
            f.write(f"Extraction Failed Samples (-1): {metrics.get('unknown_predictions', 'N/A')}\n")
            f.write(f"Processing Error Samples (-2): {metrics.get('error_predictions', 'N/A')}\n\n")

            f.write("Sleep Stage Sample Distribution (Based on Valid True Labels):\n")
            for i in range(6):
                count = class_counts.get(i, 0)
                percentage = (count / total_valid_true_samples) * 100 if total_valid_true_samples > 0 else 0
                label_name = SLEEP_STAGE_LABELS[i]
                f.write(f"  {label_name:<20}: {count:<5} samples ({percentage:.2f}%)\n")
            f.write("\n")

            f.write("Performance Metrics (Based on Valid Predictions):\n")
            f.write(f"Overall Accuracy: {format_metric(metrics.get('accuracy'))}\n")
            f.write(f"Macro Precision: {format_metric(metrics.get('precision_macro'))}\n")
            f.write(f"Macro Recall: {format_metric(metrics.get('recall_macro'))}\n")
            f.write(f"Macro F1 Score: {format_metric(metrics.get('f1_macro'))}\n")
            f.write(f"Average Inference Time (All Samples): {format_metric(metrics.get('avg_time_per_sample'))} seconds/sample\n\n")

            f.write("Metrics by Sleep Stage:\n")
            # 使用字符串格式化直接写入表格
            header_fmt = "  {:<20} {:>10} {:>10} {:>10} {:>10}\n"
            row_fmt = "  {:<20} {:>10} {:>10} {:>10} {:>10}\n"
            
            f.write(header_fmt.format("Stage", "Accuracy", "Precision", "Recall", "F1"))
            f.write("  " + "-" * 63 + "\n") # 分隔线
            
            class_accuracies_final = metrics.get('class_accuracies', {})
            class_metrics_final = metrics.get('class_metrics', {})
            cm_labels_final = metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])
            
            for i, label_name in enumerate(SLEEP_STAGE_LABELS):
                # 获取每个指标
                short_label = cm_labels_final[i] if i < len(cm_labels_final) else f"Label_{i}"
                
                # 使用format_metric格式化指标值
                acc_str = format_metric(class_accuracies_final.get(i))
                p_str = format_metric(class_metrics_final.get(short_label, {}).get('precision'))
                r_str = format_metric(class_metrics_final.get(short_label, {}).get('recall'))
                f1_str = format_metric(class_metrics_final.get(short_label, {}).get('f1'))
                
                # 直接写入格式化的行
                f.write(row_fmt.format(label_name, acc_str, p_str, r_str, f1_str))

        print(f"Test summary saved to: {summary_path}")
    except Exception as e:
        print(f"Error saving summary text file: {e}")
        print(traceback.format_exc())

    # 调用函数保存结果为Excel表格 (使用新 metrics)
    save_results_to_excel(results, metrics, None, result_dir, model_name, test_set_name)

    # 更新总表格文件 (使用新 metrics)
    try:
         dir_timestamp = os.path.basename(result_dir).split('_')[-1]
         if not (len(dir_timestamp) == 15 and dir_timestamp[:8].isdigit() and dir_timestamp[9:].isdigit()):
              print(f"警告: 从目录 '{result_dir}' 提取的时间戳格式不符，使用当前时间替代。")
              dir_timestamp = timestamp_str
    except Exception:
         print(f"警告: 无法从目录 '{result_dir}' 提取时间戳，使用当前时间替代。")
         dir_timestamp = timestamp_str

    parent_save_dir = os.path.dirname(result_dir)
    update_summary_excel(results, metrics, test_data, parent_save_dir, model_name, test_set_name, dir_timestamp)


def save_results_to_excel(results, metrics, _, save_dir, model_name="unknown_model", test_set_name="unknown_dataset"):
    """保存测试结果为Excel表格文件
       (更新版：使用新的 metrics 结构, 忽略第三个参数, 保存到指定 save_dir)

    Args:
        results: 详细测试结果列表 (list of dicts)
        metrics: 计算出的性能指标字典
        _: Placeholder for the ignored third argument (previously test_data)
        save_dir: 保存 Excel 文件的目录 (通常是特定运行的 result_dir 或 intermediate_dir)
        model_name: 模型名称
        test_set_name: 测试集名称
    """
    try:
        # 导入必要的模块
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        # --- 样式定义 (移到函数内部) ---
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        unknown_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        error_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Helper function to format metric values safely (定义在函数内部)
        def format_metric(value, fmt=".4f"):
             if value is None or value == 'N/A': return 'N/A'
             try: return f"{float(value):{fmt}}"
             except (ValueError, TypeError): return 'N/A'

        # Excel 文件保存在指定的 save_dir 中
        excel_filename = f'{model_name}_{test_set_name}_detailed_results.xlsx'
        if 'intermediate' in os.path.basename(save_dir):
             excel_filename = f'{model_name}_{test_set_name}_intermediate_detailed_results.xlsx'
        excel_path = os.path.join(save_dir, excel_filename)

        wb = Workbook()

        # --- 1. 预测结果工作表 ---
        ws_results = wb.active
        ws_results.title = "预测结果"

        headers = ["样本ID", "真实标签", "预测标签", "标签状态", "是否正确", "预测时间(秒)", "预测文本"]
        ws_results.append(headers)

        # 设置表头样式
        for cell in ws_results[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # 添加数据
        for r in results:
             pred_label = r.get('prediction', -2)
             true_label = r.get('true_label', -99)
             is_correct = r.get('correct', False)

             prediction_display = str(pred_label)
             label_status = ""
             fill_color = None

             if pred_label == -1:
                 prediction_display = "提取失败"
                 label_status = "提取失败 (-1)"
                 fill_color = unknown_fill
             elif pred_label == -2:
                 prediction_display = "处理错误"
                 label_status = "处理错误 (-2)"
                 fill_color = error_fill
             elif pred_label < 0:
                 prediction_display = f"无效代码({pred_label})"
                 label_status = f"无效代码 ({pred_label})"
                 fill_color = error_fill
             else:
                 label_status = f"有效 ({pred_label})"
                 if true_label >= 0:
                      if is_correct:
                           fill_color = correct_fill
                      else:
                           fill_color = incorrect_fill
                 else:
                      fill_color = None

             correctness_display = is_correct if pred_label >= 0 and true_label >= 0 else 'N/A'

             inf_time = r.get('inference_time')
             inf_time_display = f"{inf_time:.4f}" if isinstance(inf_time, (int, float)) else 'N/A'

             row_data = [
                 r.get('id', 'N/A'),
                 true_label if true_label >= 0 else f"无效({true_label})",
                 prediction_display,
                 label_status,
                 correctness_display,
                 inf_time_display,
                 r.get('response', '')
             ]
             ws_results.append(row_data)

             # 应用行样式和条件格式
             current_row = ws_results.max_row
             for col_idx, cell in enumerate(ws_results[current_row], 1):
                 cell.border = border
                 cell.alignment = left_alignment if col_idx == 7 else center_alignment
                 if fill_color:
                      cell.fill = fill_color

        # --- 2. 总体指标工作表 ---
        ws_metrics = wb.create_sheet("总体指标")

        # 使用内部定义的 get_metric 辅助函数
        def get_metric_value(key, fmt=None):
             val = metrics.get(key)
             if val is None or val == 'N/A': return 'N/A'
             if fmt is None: return val
             try:
                 f_val = float(val)
                 if fmt == ".4f": return f"{f_val:.4f}"
                 if fmt == ".2%": return f"{f_val:.2%}"
                 return f_val
             except (ValueError, TypeError): return 'N/A'

        ws_metrics.append(["测试信息"])
        ws_metrics.append(["模型名称", model_name])
        ws_metrics.append(["测试集名称", test_set_name])
        ws_metrics.append(["测试时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_metrics.append([])
        ws_metrics.append(["数据集统计"])
        ws_metrics.append(["总样本数", get_metric_value('total_samples')])
        ws_metrics.append(["有效预测数 (用于指标计算)", get_metric_value('valid_predictions')])
        ws_metrics.append(["提取失败数 (-1)", get_metric_value('unknown_predictions')])
        ws_metrics.append(["处理错误数 (-2)", get_metric_value('error_predictions')])
        ws_metrics.append(["未知预测比例 (-1)", get_metric_value('unknown_ratio', ".4f")])
        ws_metrics.append(["错误预测比例 (-2)", get_metric_value('error_ratio', ".4f")])
        ws_metrics.append([])
        ws_metrics.append(["整体性能指标 (基于有效预测)"])
        ws_metrics.append(["总体准确率", get_metric_value('accuracy', ".4f")])
        ws_metrics.append(["宏平均精确率", get_metric_value('precision_macro', ".4f")])
        ws_metrics.append(["宏平均召回率", get_metric_value('recall_macro', ".4f")])
        ws_metrics.append(["宏平均F1分数", get_metric_value('f1_macro', ".4f")])
        ws_metrics.append(["平均推理时间 (所有样本)", get_metric_value('avg_time_per_sample', ".4f")])

        # 应用样式
        header_rows_metrics = [1, 6, 13]
        for row_idx in range(1, ws_metrics.max_row + 1):
             is_header_row = row_idx in header_rows_metrics
             cell_a = ws_metrics.cell(row=row_idx, column=1)
             cell_b = ws_metrics.cell(row=row_idx, column=2)

             cell_a.border = border
             cell_a.alignment = left_alignment
             if is_header_row:
                  cell_a.font = title_font
                  cell_a.fill = header_fill
                  if cell_b.value is None:
                      try: ws_metrics.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                      except: pass
             else:
                 cell_a.font = subheader_font

             cell_b.border = border
             cell_b.alignment = center_alignment
             if is_header_row:
                  cell_b.fill = header_fill

        # --- 3. 详细类别指标工作表 ---
        ws_class_metrics = wb.create_sheet("详细类别指标")

        headers_class = ["睡眠阶段", "准确率", "精确率", "召回率", "F1分数"]
        ws_class_metrics.append(headers_class)

        # 设置表头样式
        for cell in ws_class_metrics[1]:
             cell.font = header_font
             cell.fill = header_fill
             cell.alignment = center_alignment
             cell.border = border

        # 添加数据
        class_accuracies_final = metrics.get('class_accuracies', {})
        class_metrics_final = metrics.get('class_metrics', {})
        cm_labels_final = metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])

        for i, label_name in enumerate(SLEEP_STAGE_LABELS):
             acc = class_accuracies_final.get(i)
             short_label = cm_labels_final[i] if i < len(cm_labels_final) else f"Label_{i}"
             class_p_r_f1 = class_metrics_final.get(short_label, {})
             p = class_p_r_f1.get('precision')
             r = class_p_r_f1.get('recall')
             f1 = class_p_r_f1.get('f1')

             acc_str = format_metric(acc)
             p_str = format_metric(p)
             r_str = format_metric(r)
             f1_str = format_metric(f1)

             row_data = [label_name, acc_str, p_str, r_str, f1_str]
             ws_class_metrics.append(row_data)

             # 应用行样式
             current_row = ws_class_metrics.max_row
             for col_idx, cell in enumerate(ws_class_metrics[current_row], 1):
                 cell.border = border
                 cell.alignment = left_alignment if col_idx == 1 else center_alignment
                 if col_idx == 1: cell.font = subheader_font

        # --- 4. 混淆矩阵工作表 ---
        ws_cm = wb.create_sheet("混淆矩阵")
        ws_cm.append(["混淆矩阵 (基于有效预测)"])
        num_labels_cm = len(cm_labels_final)
        if num_labels_cm > 0:
             end_col_letter = get_column_letter(1 + num_labels_cm)
             try: ws_cm.merge_cells(f'A1:{end_col_letter}1')
             except: pass
        ws_cm['A1'].font = title_font
        ws_cm['A1'].alignment = center_alignment
        ws_cm['A1'].fill = header_fill

        headers_cm = ["真实 \\ 预测"] + cm_labels_final
        ws_cm.append(headers_cm)
        for cell in ws_cm[2]:
             cell.font = header_font
             cell.fill = header_fill
             cell.alignment = center_alignment
             cell.border = border

        cm_data = metrics.get('confusion_matrix')
        if cm_data is not None and isinstance(cm_data, list) and len(cm_data) == num_labels_cm and all(isinstance(row, list) and len(row) == num_labels_cm for row in cm_data):
             cm_data_np = np.array(cm_data)
             for i, label in enumerate(cm_labels_final):
                  row_data = [label] + [cm_data_np[i, j] for j in range(num_labels_cm)]
                  ws_cm.append(row_data)

             for row_idx in range(3, ws_cm.max_row + 1):
                 for col_idx in range(1, ws_cm.max_column + 1):
                     cell = ws_cm.cell(row=row_idx, column=col_idx)
                     cell.alignment = center_alignment
                     cell.border = border
                     if col_idx == 1:
                          cell.font = header_font
                     if row_idx > 2 and col_idx > 1 and (row_idx - 3) == (col_idx - 2):
                          cell.fill = correct_fill
        else:
             ws_cm.append(["混淆矩阵数据无效或缺失"])

        # --- 调整所有工作表的列宽 ---
        for ws in wb.worksheets:
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        cell_text = str(cell.value)
                        max_line_length = max(len(line) for line in cell_text.split('\n'))
                        dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), max_line_length)
            for col, value in dims.items():
                 ws.column_dimensions[col].width = max(12, value + 4)

        # --- 保存Excel文件 ---
        try:
            wb.save(excel_path)
            print(f"详细结果Excel表格已保存至: {excel_path}")
        except PermissionError:
            print(f"错误: 无法写入Excel文件 '{excel_path}'. 文件可能被其他程序占用。")
        except Exception as e_save:
            print(f"保存详细结果Excel时发生错误: {e_save}")
            print(traceback.format_exc())

    except ImportError:
         print("错误：需要 'openpyxl' 库来保存Excel文件。请安装：pip install openpyxl")
    except Exception as e:
        print(f"生成详细结果Excel时发生意外错误: {e}")
        print(traceback.format_exc())


def update_summary_excel(results, metrics, test_data, save_dir, model_name, test_set_name, timestamp):
    """更新实验结果总表格
       (更新版：使用新的 metrics 结构)
    """
    try:
        # Ensure openpyxl is available
        import openpyxl
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import os
        from datetime import datetime

        # 总表格文件路径 (保存在主 save_dir)
        summary_excel_path = os.path.join(save_dir, 'experiments_summary_sleep.xlsx')

        # 定义样式
        header_font = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # --- 准备新行数据 (使用新 metrics) ---
        try:
            test_time_str = datetime.strptime(timestamp, '%Y%m%d_%H%M%S').strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            print(f"警告: 无法解析时间戳 '{timestamp}' 用于总表, 使用当前时间。")
            test_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        experiment_id = f"{model_name}_{test_set_name}_{timestamp}"

        # Get per-class metrics safely using .get()
        class_accuracies = metrics.get('class_accuracies', {})
        class_metrics_dict = metrics.get('class_metrics', {})
        short_labels = metrics.get('confusion_matrix_labels', ['W', 'N1', 'N2', 'N3', 'N4', 'R'])

        # Helper function to format metric values safely for the summary sheet
        def format_summary_metric(value, fmt=".4f"):
            if value is None or value == 'N/A': return 'N/A'
            try: return f"{float(value):{fmt}}"
            except (ValueError, TypeError): return 'N/A'

        stage_accuracies_list = [format_summary_metric(class_accuracies.get(i)) for i in range(6)]
        stage_precisions_list = [format_summary_metric(class_metrics_dict.get(short_labels[i], {}).get('precision')) for i in range(6)]
        stage_recalls_list = [format_summary_metric(class_metrics_dict.get(short_labels[i], {}).get('recall')) for i in range(6)]
        stage_f1s_list = [format_summary_metric(class_metrics_dict.get(short_labels[i], {}).get('f1')) for i in range(6)]

        new_row_data = [
            experiment_id,
            test_time_str,
            model_name,
            test_set_name,
            metrics.get('total_samples', 'N/A'),
            metrics.get('valid_predictions', 'N/A'),
            metrics.get('unknown_predictions', 'N/A'),
            metrics.get('error_predictions', 'N/A'),
            format_summary_metric(metrics.get('accuracy')),
            format_summary_metric(metrics.get('precision_macro')),
            format_summary_metric(metrics.get('recall_macro')),
            format_summary_metric(metrics.get('f1_macro')),
            format_summary_metric(metrics.get('avg_time_per_sample')),
        ] + stage_accuracies_list + stage_precisions_list + stage_recalls_list + stage_f1s_list

        # --- 加载或创建工作簿 ---
        if os.path.exists(summary_excel_path):
            try:
                wb = load_workbook(summary_excel_path)
                if "实验结果总表" in wb.sheetnames:
                    ws = wb["实验结果总表"]
                else:
                    print(f"警告: 工作表 '实验结果总表' 不存在于 '{summary_excel_path}'. 使用活动工作表。")
                    ws = wb.active
                    ws.title = "实验结果总表"
            except Exception as e:
                print(f"无法加载现有总表格 '{summary_excel_path}': {e}. 将创建新文件。")
                wb = Workbook()
                ws = wb.active
                ws.title = "实验结果总表"
        else:
            print(f"总表格 '{summary_excel_path}' 不存在，将创建新文件。")
            wb = Workbook()
            ws = wb.active
            ws.title = "实验结果总表"

        # --- 添加表头（如果工作表是空的或只有一行） ---
        is_empty_or_new = ws.max_row <= 1
        has_headers = False
        if not is_empty_or_new:
             if ws.cell(row=1, column=1).value == "实验ID":
                  has_headers = True

        if is_empty_or_new or not has_headers:
             if not is_empty_or_new:
                  print("警告：总表第一行内容与预期表头不符，将重新写入表头。")
             headers = [
                 "实验ID", "实验时间", "模型名称", "测试集名称", "总样本数", "有效预测数",
                 "提取失败数", "处理错误数",
                 "准确率", "宏平均精确率", "宏平均召回率", "宏平均F1",
                 "平均推理时间(秒/样本)",
                 "W Acc", "N1 Acc", "N2 Acc", "N3 Acc", "N4 Acc", "R Acc",
                 "W P", "N1 P", "N2 P", "N3 P", "N4 P", "R P",
                 "W R", "N1 R", "N2 R", "N3 R", "N4 R", "R R",
                 "W F1", "N1 F1", "N2 F1", "N3 F1", "N4 F1", "R F1"
             ]
             ws.insert_rows(1)
             ws.append(headers)

             # 设置表头样式
             for cell in ws[1]:
                 cell.font = header_font
                 cell.alignment = center_alignment
                 cell.border = border

        # --- 追加新行数据 ---
        ws.append(new_row_data)

        # 设置新添加行的样式 (apply to the last row)
        last_row_idx = ws.max_row
        for cell in ws[last_row_idx]:
            cell.alignment = center_alignment
            cell.border = border

        # --- 调整列宽 ---
        for col_idx in range(1, ws.max_column + 1):
             max_length = 0
             col_letter = get_column_letter(col_idx)
             if ws.cell(row=1, column=col_idx).value:
                  max_length = max(max_length, len(str(ws.cell(row=1, column=col_idx).value)))
             if ws.cell(row=ws.max_row, column=col_idx).value:
                  max_length = max(max_length, len(str(ws.cell(row=ws.max_row, column=col_idx).value)))

             ws.column_dimensions[col_letter].width = max(12, max_length + 4)

        # --- 保存Excel文件 ---
        try:
            wb.save(summary_excel_path)
            print(f"实验结果总表已更新: {summary_excel_path}")
        except PermissionError:
            print(f"错误：无法写入总表 '{summary_excel_path}'. 文件可能被其他程序占用。")
        except Exception as e:
            print(f"保存总表时出错: {e}")
            print(traceback.format_exc())

    except ImportError:
         print("错误：需要 'openpyxl' 库来更新Excel总表。请安装：pip install openpyxl")
    except Exception as e:
        print(f"更新总表时发生意外错误: {e}")
        print(traceback.format_exc())


if __name__ == '__main__':
    # 设置API端口
    port = os.environ.get("API_PORT", 8001)
    
    # 初始化OpenAI客户端
    client = OpenAI(
        api_key="0",
        base_url=f"http://localhost:{port}/v1",
    )
    
    # 获取测试数据路径和模型名称
    default_test_path = "/data/lhc/datasets_new/emotion/test/sleep_st_44_100hz_eeg15s-step15s_emo2.0s-step1s_win_all_tok13101_bal0.2_sqrt_inverse_202504292208_test.json"
    test_data_path = os.environ.get("TEST_DATA_PATH", default_test_path)

    # 添加文件存在性检查
    if not os.path.exists(test_data_path):
        print(f"错误：测试数据文件未找到: {test_data_path}")
        print("请检查路径是否正确，或通过环境变量 TEST_DATA_PATH 指定有效路径。")
        exit(1)

    try:
        max_samples_env = os.environ.get("MAX_SAMPLES")
        max_samples = int(max_samples_env) if max_samples_env and max_samples_env.isdigit() else None
        test_data = load_test_data(test_data_path, max_samples=max_samples)
        if not test_data:
             print("错误：加载测试数据后列表为空。")
             exit(1)
    except Exception as e:
        print(f"加载测试数据时出错: {e}")
        print(traceback.format_exc())
        exit(1)

    # 从测试数据路径中提取测试集名称
    file_basename = os.path.basename(test_data_path).split('.')[0]
    test_set_name = file_basename
    suffixes_to_remove = [r'_test', r'_train', r'_val', r'_n\d+', r'_\d{8,}(?:_\d{4,})?', r'_tok\d+', r'bal[\d.]+', r'sqrt_inverse', r'_all', r'_win']
    for suffix in suffixes_to_remove:
         test_set_name = re.sub(suffix + '$', '', test_set_name, flags=re.IGNORECASE)
    test_set_name = re.sub(r'_+$', '', test_set_name)

    if not test_set_name:
        test_set_name = file_basename

    # 尝试从环境变量或配置中获取模型名称
    model_name = os.environ.get("MODEL_NAME", "emotion_st44")

    print(f"模型: {model_name}")
    print(f"测试集: {test_set_name} (来自: {os.path.basename(test_data_path)})")
    print(f"加载了 {len(test_data)} 个测试样本")

    # 创建保存结果的目录 (主目录，子目录在 evaluate_model 中创建)
    save_dir = os.environ.get("SAVE_DIR", "/data/lhc/results")
    os.makedirs(save_dir, exist_ok=True)

    # 设置打印和保存的间隔
    print_interval = int(os.environ.get("PRINT_INTERVAL", 10))
    # 计算默认保存间隔为总样本数的十分之一，确保至少为1
    default_save_interval = max(1, len(test_data) // 10) if test_data else 100
    save_interval = int(os.environ.get("SAVE_INTERVAL", default_save_interval))
    print(f"打印间隔: {print_interval}, 保存间隔: {save_interval} (默认为总样本数的1/10)")

    # 获取 max_tokens for API calls
    default_max_tokens = 60000
    max_tokens_env = os.environ.get("MAX_TOKENS")
    try:
         max_tokens = int(max_tokens_env) if max_tokens_env else default_max_tokens
    except ValueError:
         print(f"警告: 无效的 MAX_TOKENS 环境变量值 '{max_tokens_env}'. 使用默认值 {default_max_tokens}.")
         max_tokens = default_max_tokens
    print(f"API 最大 Token 数: {max_tokens}")

    # 评估模型
    metrics = evaluate_model(
        client,
        test_data,
        print_interval=print_interval,
        save_interval=save_interval,
        save_dir=save_dir,
        model_name=model_name,
        test_set_name=test_set_name,
        max_tokens=max_tokens
    )

    print("\n测试完成!")